import csv
import datetime
from django.contrib.auth.models import User
from django.db import transaction
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Sum
from django.http import HttpResponse
import glob
import logging
from numpy import *
import numpy as np
import openpyxl
import signal
import psutil
import pandas as pd
import shutil
import json
from pyper import *
import subprocess
from uuid import uuid4
import fnmatch
import zipfile
import tarfile


from database.models import Project, Reference, \
    Sample, Air, Human_Associated, Microbial, Soil, Water, UserDefined, \
    Kingdom, Phyla, Class, Order, Family, Genus, Species, OTU_99, Profile, \
    koOtuList, ko_entry, ko_lvl1, ko_lvl2, ko_lvl3, PICRUSt

from database import perms

import functions
from functions.utils.debug import debug
import h5py

stage = ''
last = ''
mothurStat = ""
perc = 0
rep_project = ''
pd.set_option('display.max_colwidth', -1)
LOG_FILENAME = 'error_log.txt'
pro = None


def dada2(dest, source):
    global pro
    try:
        global stage, perc, mothurStat
        stage = "Step 3 of 5: Running R..."
        perc = 0

        if os.name == 'nt':
            r = R(RCMD="R/R-Portable/App/R-Portable/bin/R.exe", use_pandas=True)
        else:
            r = R(RCMD="R/R-Linux/bin/R", use_pandas=True)

        r('.cran_packages <-  c("ggplot2", "gridExtra", "reshape2", "qiimer", "devtools")')
        r("new.packages <- .cran_packages[!(.cran_packages%in% installed.packages()[,'Package'])]")
        r("if (length(new.packages)) install.packages(new.packages, repos='http://cran.us.r-project.org', dependencies=T)")

        r('.bioc_packages <- c("rhdf5")')
        r("new.packages <- .bioc_packages[!(.bioc_packages %in% installed.packages()[,'Package'])]")
        r("if (length(new.packages)) source('http://bioconductor.org/biocLite.R')")
        r("if (length(new.packages)) biocLite(new.packages, type='source', suppressUpdate=T, dependencies=T)")

        r('.dev_tools <- c("biom")')
        r("new.packages <- .dev_tools[!(.dev_tools %in% installed.packages()[,'Package'])]")
        r('if (length(new.packages)) devtools::install_github("biom", "joey711")')

        r('.bioc_packages <- c("dada2", "phyloseq")')
        r("new.packages <- .bioc_packages[!(.bioc_packages %in% installed.packages()[,'Package'])]")
        r("if (length(new.packages)) source('http://bioconductor.org/biocLite.R')")
        r("if (length(new.packages)) biocLite(new.packages, type='source', suppressUpdate=T, dependencies=T)")

        if os.name == 'nt':
            cmd = "R\\R-Portable\\App\\R-Portable\\bin\\R.exe --no-save --no-restore < mothur\\temp\\dada2.R"
        else:
            cmd = "R/R-Linux/bin/R --no-save --no-restore < mothur/temp/dada2.R"

        try:
            f = open('mothur/temp/R.history', 'w')
            pro = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, bufsize=0)
            while True:
                line = pro.stdout.readline()
                if line != '':
                    mothurStat += line
                    f.write(line)
                else:
                    break

            if os.name == 'nt':
                cmd = "mothur\\mothur-win\\vsearch -usearch_global mothur\\temp\\dada.fasta -db mothur\\temp\\ref_trimmed.fa.gz --strand both --id 0.99 --fastapairs mothur\\temp\\pairs.fasta --notmatched mothur\\temp\\nomatch.fasta"
            else:
                cmd = "mothur/mothur-linux/vsearch -usearch_global mothur/temp/dada.fasta -db mothur/temp/ref_trimmed.fa.gz --strand both --id 0.99 --fastapairs mothur/temp/pairs.fasta --notmatched mothur/temp/nomatch.fasta"

            pro = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, bufsize=0)
            while True:
                line = pro.stdout.readline()
                if line != '':
                    mothurStat += line
                    f.write(line)
                else:
                    break
                line = pro.stderr.readline()
                if line != '':
                    mothurStat += line
                    f.write(line)
                else:
                    break
            mothurStat += '\n\nR processing is done! \n\n'
            f.close()

            f = open('mothur/temp/dada.cons.taxonomy', 'w')
            f.write("OTU\tSeq\tTaxonomy\n")
            pairDict = {}

            inFile1 = open('mothur/temp/pairs.fasta', 'r')
            counter = 1
            key = ''
            for line in inFile1:
                if counter == 1:
                    key = line.rstrip().replace('>', '')
                if counter == 3:
                    value = line.rstrip().replace('>', '')
                    pairDict[key] = value
                if counter == 5:
                    counter = 0
                counter += 1
            inFile1.close()

            inFile2 = open('mothur/temp/dada.fasta', 'r')
            counter = 1
            for line in inFile2:
                if counter == 1:
                    key = line.rstrip().replace('>', '')
                if counter == 2:
                    seq = line.rstrip()
                    if key in pairDict:
                        f.write(key)
                        f.write('\t')
                        f.write(seq)
                        f.write('\t')
                        f.write(pairDict[key])
                        f.write('\n')
                    else:
                        f.write(key)
                        f.write('\t')
                        f.write(seq)
                        f.write('\t')
                        if OTU_99.objects.filter(otuSeq=seq).exists():
                            otuName = OTU_99.objects.get(otuSeq=seq).otuName
                            f.write('k__unclassified;p__unclassified;c__unclassified;o__unclassified;f__unclassified;g__unclassified;s__unclassified;' + str(otuName) + ';')
                            f.write('\n')
                        else:
                            f.write('k__unclassified;p__unclassified;c__unclassified;o__unclassified;f__unclassified;g__unclassified;s__unclassified;otu__unclassified;')
                            f.write('\n')
                    counter = 0
                counter += 1

            inFile2.close()
            f.close()

            shutil.copy('mothur/temp/pairs.fasta', '% s/dada.vsearch_pairs.txt' % dest)
            shutil.copy('mothur/temp/dada.fasta', '% s/dada.rep_seqs.fasta' % dest)
            shutil.copy('mothur/temp/dada.cons.taxonomy', '% s/final.cons.taxonomy' % dest)
            shutil.copy('mothur/temp/dada.shared', '% s/final.tx.shared' % dest)
            shutil.copy('mothur/temp/dada2.R', '% s/dada2.R' % dest)
            shutil.copy('mothur/temp/R.history', '% s/R.history' % dest)
            shutil.rmtree('mothur/temp')

            dir = os.getcwd()

            path = os.path.join(dir, 'mothur', 'reference', 'align')
            stuff = os.listdir(path)
            for thing in stuff:
                if thing.endswith(".8mer"):
                    os.remove(os.path.join(path, thing))

            path = os.path.join(dir, 'mothur', 'reference', 'taxonomy')
            stuff = os.listdir(path)
            for thing in stuff:
                if thing.endswith(".numNonZero") or thing.endswith(".prob") or thing.endswith(".sum") or thing.endswith(".train"):
                    os.remove(os.path.join(path, thing))

            path = os.path.join(dir, 'mothur', 'reference', 'template')
            stuff = os.listdir(path)
            for thing in stuff:
                if thing.endswith(".8mer"):
                    os.remove(os.path.join(path, thing))

        except Exception as e:
            print "R failed: " + str(e)

    except Exception:
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)


def mothur(dest, source):
    global pro
    try:
        debug("Starting mothur: dest", dest, "source", source)
        global stage, perc, mothurStat
        stage = "Step 3 of 5: Running mothur..."
        perc = 0

        if not os.path.exists('mothur/reference/align'):
            os.makedirs('mothur/reference/align')

        if not os.path.exists('mothur/reference/taxonomy'):
            os.makedirs('mothur/reference/taxonomy')

        if not os.path.exists('mothur/reference/template'):
            os.makedirs('mothur/reference/template')

        if os.name == 'nt':
            filepath = "mothur\\mothur-win\\mothur.exe mothur\\temp\\mothur.batch"
        else:
            filepath = "mothur/mothur-linux/mothur mothur/temp/mothur.batch"
        try:
            pro = subprocess.Popen(filepath, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, bufsize=0)
            while True:
                line = pro.stdout.readline()
                if line != '':
                    mothurStat += line
                else:
                    break
            mothurStat = '\n\nmothur processing is done! \n\n'

        except Exception as e:
            print "mothur failed: " + str(e)

        if source == '454_sff':
            shutil.copy('mothur/temp/final.fasta', '% s/final.fasta' % dest)
            shutil.copy('mothur/temp/final.names', '% s/final.names' % dest)
            shutil.copy('mothur/temp/final.groups', '% s/final.groups' % dest)
            shutil.copy('mothur/temp/final.cons.taxonomy', '% s/final.cons.taxonomy' % dest)
            shutil.copy('mothur/temp/final.tx.shared', '% s/final.tx.shared' % dest)
            shutil.rmtree('mothur/temp')

        if source == '454_fastq':
            shutil.copy('mothur/temp/final.fasta', '% s/final.fasta' % dest)
            shutil.copy('mothur/temp/final.names', '% s/final.names' % dest)
            shutil.copy('mothur/temp/final.groups', '% s/final.groups' % dest)
            shutil.copy('mothur/temp/final.cons.taxonomy', '% s/final.cons.taxonomy' % dest)
            shutil.copy('mothur/temp/final.tx.shared', '% s/final.tx.shared' % dest)
            shutil.rmtree('mothur/temp')

        if source == 'miseq':
            shutil.copy('mothur/temp/final.fasta', '% s/final.fasta' % dest)
            shutil.copy('mothur/temp/final.names', '% s/final.names' % dest)
            shutil.copy('mothur/temp/final.groups', '% s/final.groups' % dest)
            shutil.copy('mothur/temp/final.cons.taxonomy', '% s/final.cons.taxonomy' % dest)
            shutil.copy('mothur/temp/final.tx.shared', '% s/final.tx.shared' % dest)
            shutil.rmtree('mothur/temp')

        for afile in glob.glob(r'*.logfile'):
            shutil.move(afile, dest)

        dir = os.getcwd()

        path = os.path.join(dir, 'mothur', 'reference', 'align')
        stuff = os.listdir(path)
        for thing in stuff:
            if thing.endswith(".8mer"):
                os.remove(os.path.join(path, thing))

        path = os.path.join(dir, 'mothur', 'reference', 'taxonomy')
        stuff = os.listdir(path)
        for thing in stuff:
            if thing.endswith(".numNonZero") or thing.endswith(".prob") or thing.endswith(".sum") or thing.endswith(".train"):
                os.remove(os.path.join(path, thing))

        path = os.path.join(dir, 'mothur', 'reference', 'template')
        stuff = os.listdir(path)
        for thing in stuff:
            if thing.endswith(".8mer"):
                os.remove(os.path.join(path, thing))

    except Exception:
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)


def termP():    # relies on global of pro because only one dataprocess should ever be running at a single time
    global pro  # pro as in process to be terminated
    try:
        if pro is not None:
            parent = psutil.Process(pro.pid)
            children = parent.children(recursive=True)
            for process in children:
                process.send_signal(signal.SIGTERM)
            os.kill(pro.pid, signal.SIGTERM)

            # clean up mothur files here
            dir = os.getcwd()
            stuff = os.listdir(dir)
            for thing in stuff:
                if thing.endswith(".num.temp") or thing.endswith(".logfile"):
                    os.remove(os.path.join(dir, thing))

            path = os.path.join(dir, 'mothur', 'reference', 'align')
            stuff = os.listdir(path)
            for thing in stuff:
                if thing.endswith(".8mer"):
                    os.remove(os.path.join(path, thing))

            path = os.path.join(dir, 'mothur', 'reference', 'taxonomy')
            stuff = os.listdir(path)
            for thing in stuff:
                if thing.endswith(".numNonZero") or thing.endswith(".prob") or thing.endswith(".sum") or thing.endswith(".train"):
                    os.remove(os.path.join(path, thing))

            path = os.path.join(dir, 'mothur', 'reference', 'template')
            stuff = os.listdir(path)
            for thing in stuff:
                if thing.endswith(".8mer"):
                    os.remove(os.path.join(path, thing))

    except Exception as e:
        print "Error with terminate: "+str(e)  # +", probably just not mothur"
        pass


def status(request):
    global last, mothurStat
    if request.is_ajax():
        RID = request.GET['all']
        queuePos = functions.datstat(RID)
        dict = {}
        if queuePos == 0:
            # check for duplicate output
            if last != mothurStat:
                dict['mothurStat'] = mothurStat
                last = mothurStat
            else:
                dict['mothurStat'] = ''

            dict['stage'] = stage
            dict['perc'] = "%.2f" % perc
            dict['project'] = rep_project
        else:
            if queuePos == -1024:  # need to distinguish between active and inactive PID    # gets stuck here? TODO 1.4
                dict['stage'] = "Stopping...please be patient while we restore the database!"   # jump
            elif queuePos == -512:  # for inactive processes, problem is when process DOES get stopped, queuepos
                                    #  switches to this and THEN loops anyways
                dict['stage'] = "Removing request from queue..."
                # need front end to catch this... need to get datfuncall to its page too (specific to function called)
            else:
                dict['stage'] = "In queue for processing, "+str(queuePos)+" requests in front of you"
                # TODO 1.3 getting stuck here with queuePos of None during fastq 5.9 GB .tar.gz upload

            dict['perc'] = 0    # TODO 1.4 could use this if given starting position from client (though jumps won't be accurate)
            dict['project'] = rep_project
            dict['mothurStat'] = ''

        mothurStat = ""
        json_data = json.dumps(dict)
        return HttpResponse(json_data, content_type='application/json')


def projectid(Document):
    try:
        global stage, perc
        perc = 0
        stage = "Step 1 of 5: Parsing project file..."
        wb = openpyxl.load_workbook(Document, data_only=True, read_only=False)  # treated as an archive?
        myDict = functions.excel_to_dict(wb, headerRow=5, nRows=1, sheet='Project')
        rowDict = myDict[0]
        pType = rowDict['projectType']
        projectid = rowDict['projectid']

        try:
            num_samp = int(rowDict['num_samp'])
        except:
            sampSheet = wb.worksheets[2]
            rowMax = sampSheet.max_row
            num_samp = 0
            for val in range(7, rowMax):
                samp = sampSheet['C'+str(val)].value
                if samp is not "" and samp is not None:
                    num_samp += 1

        if projectid is np.nan or projectid is None:
            p_uuid = uuid4().hex
        else:
            p_uuid = projectid

        return p_uuid, pType, num_samp

    except Exception as e:
        print "Error parsing meta project file: ", e
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)
        return None


def parse_project(Document, p_uuid, curUser):
    # no stoplist support because of how fast this is, and
    # cleanup varies based on at which point in this function the stop would be called (so just check before and after)

    # FUNCTION DESCRIPTION:
    # parse meta file (xlxs file) to create new project (or update an existing one)

    # check formatting and such, boxes have ranges of valid responses, check project type and permissions for now
    try:
        global stage, perc
        perc = 50
        wb = openpyxl.load_workbook(Document, data_only=True, read_only=False)
        myDict = functions.excel_to_dict(wb, headerRow=5, nRows=1, sheet='Project')
        rowDict = myDict[0]
        rowDict.pop('num_samp')

        # check project status
        if rowDict['status'] == "public" or rowDict['status'] == "private":
            pass
        else:
            return "status has options public and private only"

        # check project type
        if rowDict['projectType'] in {'air', 'human gut', 'human associated', 'microbial', 'soil', 'water'}:
            pass
        else:
            return "projectType options are air, human gut, human associated," \
                " microbial, soil, and water"

        newProjectStatus = rowDict['status']

        if not Project.objects.filter(projectid=p_uuid).exists():
            # if this project does not exist already, make a new one
            for key in rowDict.keys():
                if key == 'projectid':
                    rowDict[key] = p_uuid
            myProj = Project.objects.create(**rowDict)  # save pointer to this somehow (myProj = ?)
            # get project and set owner to current user (if project is new)
            myProj.owner = curUser  # might break if different user adds files later
            myProj.wip = True
            myProj.save()
        else:
            rowDict.pop('projectid')
            # if this project exists, update its values
            # since it exists already, lets check if the status (public or private) is changing with this update
            # if its changing, we need to update some permissions lists
            oldProject = Project.objects.get(projectid=p_uuid)
            oldProjectStatus = oldProject.status

            # actually update the project data
            Project.objects.filter(projectid=p_uuid).update(projectid=p_uuid, wip=True, **rowDict)

            # check if permissions change occurred, run updater if so
            if oldProjectStatus != newProjectStatus:
                # permissions are changing, get ready to check some lists
                if oldProjectStatus == "public" and newProjectStatus == "private":
                    perms.pubToPriv(p_uuid)
                if oldProjectStatus == "private" and newProjectStatus == "public":
                    perms.privToPub(p_uuid)

        stage = "Step 1 of 5: Parsing project file..."
        return "none"

    except Exception as ex:
        # need to make this more specific, give out specific row and column info + type of error
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)
        return str(ex)


def parse_reference(p_uuid, refid, path, raw, source, userid):
    try:
        author = User.objects.get(id=userid)
        if not Reference.objects.filter(path=path).exists():
            project = Project.objects.get(projectid=p_uuid)
            Reference.objects.create(
                refid=refid, projectid=project, path=path, source=source, raw=raw, author=author)
        else:
            refs = Reference.objects.all().filter(path=path)
            for ref in refs:
                refid = ref.refid
            ref = Reference.objects.get(refid=refid)
            ref.save()
        return refid

    except Exception:
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)


def parse_sample(Document, p_uuid, pType, num_samp, dest, raw, source, userID, stopList, RID, PID):
    # same meta file parsing as earlier, but now we go for the actual sample data instead of project info
    try:
        global stage, perc
        stage = "Step 2 of 5: Parsing sample file..."
        perc = 0

        project = Project.objects.get(projectid=p_uuid)
        wb = openpyxl.load_workbook(Document, data_only=True, read_only=False)
        perc = 25

        dict1 = functions.excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='MIMARKs')
        perc = 50

        if pType == 'air':
            dict2 = functions.excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='Air')

        elif pType == 'human associated' or pType == 'human gut':
            dict2 = functions.excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='Human Associated')

        elif pType == 'microbial':
            dict2 = functions.excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='Microbial')

        elif pType == 'soil':
            dict2 = functions.excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='Soil')

        elif pType == 'water':
            dict2 = functions.excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='Water')

        else:
            dict2 = list()

        dict3 = functions.excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='User')
        perc = 75

        tempid = uuid4().hex
        refid = parse_reference(p_uuid, tempid, dest, raw, source, userID)
        reference = Reference.objects.get(refid=refid)

        idList = []
        refDict = {}
        for i in xrange(num_samp):
            row = dict1[i]

            s_uuid = row['sampleid']
            if s_uuid is np.nan:
                s_uuid = uuid4().hex    # incredibly unlikely to produce an already existing ID
                # but going to check anyways to be absolutely certain
                while Sample.objects.filter(sampleid=s_uuid).exists():  # keep making a new ID until we get a unique ID
                    s_uuid = uuid4().hex
                row['sampleid'] = s_uuid

            # to be used in parse_profile()
            try:
                type = row['sample_type']
                row.pop('sample_type')
            except:
                type = np.nan

            if type is not np.nan:
                refDict[s_uuid] = type
            else:
                refDict[s_uuid] = 'replace'

            row.pop('seq_method')
            row.pop('geo_loc_name')
            row.pop('lat_lon')

            debug("Parsing sample: ", str(row['sample_name']))
            try:
                badChars = set('&')  # add to this whenever a problematic character shows up
                row['sample_name'] = "".join([c for c in row['sample_name'] if c not in badChars])
            except:
                pass

            idList.append(s_uuid)
            if Sample.objects.filter(sampleid=s_uuid).exists():
                Sample.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, **row)
            else:
                Sample.objects.create(projectid=project, refid=reference, **row)
                # error: create() got multiple values for keyword argument 'refid'

            sample = Sample.objects.get(sampleid=s_uuid)

            row = dict2[i]
            if 'sampleid' in row:
                row.pop('sampleid')
            if 'sample_name' in row:
                row.pop('sample_name')

            if pType == "air":
                if not Air.objects.filter(sampleid=s_uuid).exists():
                    Air.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
                else:
                    Air.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)

            elif pType == "human associated":
                if not Human_Associated.objects.filter(sampleid=s_uuid).exists():
                    Human_Associated.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
                else:
                    Human_Associated.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)

            elif pType == "microbial":
                if not Microbial.objects.filter(sampleid=s_uuid).exists():
                    Microbial.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
                else:
                    Microbial.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)

            elif pType == "soil":
                if not Soil.objects.filter(sampleid=s_uuid).exists():
                    Soil.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
                else:
                    Soil.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)

            elif pType == "water":
                if not Water.objects.filter(sampleid=s_uuid).exists():
                    Water.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
                else:
                    Water.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)
            else:
                pass

            row = dict3[i]
            if 'sampleid' in row:
                row.pop('sampleid')
            if 'sample_name' in row:
                row.pop('sample_name')

            if not UserDefined.objects.filter(sampleid=s_uuid).exists():
                UserDefined.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
            else:
                UserDefined.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)

            perc += 20/num_samp

            if stopList[PID] == RID:
                return None

        ### add myPhyloDB generated IDs to excel metafile
        wb = openpyxl.load_workbook(Document, data_only=False, read_only=False)
        ws = wb['Project']
        ws.cell(row=6, column=4).value = p_uuid

        ws = wb['MIMARKs']
        for i in xrange(len(idList)):
            j = i + 7
            ws.cell(row=j, column=1).value = idList[i]
            ws.cell(row=j, column=2).value = 'replace'  # all samples are automatically set to replace in meta-file
            perc += 5/len(idList)

        wb.save(Document)
        return refDict

    except Exception as e:
        print "Error parsing sample: ", e
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)
        return None


def parse_taxonomy(Document, stopList, PID, RID):
    # does this function assume use of sequence or count values for column 1 (id, ?, taxa)
    try:
        global stage, perc
        # get current number of unclassified seqs
        numUnclass = OTU_99.objects.filter(otuName__startswith='isv_').count()  # assert no gaps exist
        # although its fine if a gap does exist, just an isv number we won't ever use
        # the other problem being collisions if gaps exist, but we detect those now
        newOtuList = []

        # read files
        stage = "Step 4 of 5: Parsing taxonomy file..."
        perc = 0

        debug("Reading taxa:", Document)

        f = csv.reader(Document, delimiter='\t')
        f.next()
        total = 0.0
        for row in f:
            if row:
                total += 1.0

        Document.seek(0)
        f = csv.reader(Document, delimiter='\t')
        docLines = list(f)  # convert to list
        firstRow = docLines.pop(0)  # remove first row, save for later when rewriting file

        # do we have Sequences column?
        haveSeq = "Seq" in firstRow

        # which columns are at which indices?
        colLabels = {}
        colIndex = 0
        for label in firstRow:
            colLabels[label] = colIndex
            colIndex += 1

        step = 0.0
        curLine = 0
        editedLinesDict = {}
        debug("Parse taxonomy:")
        for row in docLines:
            # before parsing another row, check if stop command has been sent
            if stopList[PID] == RID:
                return
            if row:     # can confirm row object exists, necessary?
                if len(row) == 0:    # skip if row is an empty line
                    continue
                if len(row) < len(firstRow):       # in some cases, the tab between certain elements is read as a space
                    badIndex = 0
                    for col in row:
                        if len(col.split(" ")) > 1:  # Asserting 'space' is absent in all actual values of the file
                            subRow = col.split(" ")  # take the problematic element and split
                            row.pop(badIndex)  # remove old bad element
                            row.insert(badIndex, subRow[0])   # add its components back in
                            row.insert(badIndex+1, subRow[1])
                            debug("Fixed a row:", row)

                        badIndex += 1

                step += 1.0
                perc = int(step / total * 100)
                # prep taxonomy line by cleaning up tags and percentages (we don't use either here)
                subbed = re.sub(r'(\(.*?\)|k__|p__|c__|o__|f__|g__|s__|otu__)', '', row[colLabels["Taxonomy"]])
                subbed = subbed[:-1]

                taxon = subbed.split(';')

                if len(taxon) < 1:  # missing kingdom
                    taxon.append("unclassified")
                if len(taxon) < 2:  # missing phyla
                    taxon.append("unclassified")
                if len(taxon) < 3:  # missing class
                    taxon.append("unclassified")
                if len(taxon) < 4:  # missing order
                    taxon.append("unclassified")
                if len(taxon) < 5:  # missing family
                    taxon.append("unclassified")

                if len(taxon) < 6:  # missing genus
                    taxon.append("unclassified")

                if len(taxon) < 7:  # missing species
                    taxon.append("unclassified")

                if len(taxon) < 8:  # missing otu, unclassified the best thing to fill here?
                    taxon.append("unclassified")    # in case sequence is true but otu is not

                if not Kingdom.objects.filter(kingdomName=taxon[0]).exists():
                    kid = uuid4().hex
                    Kingdom.objects.create(kingdomid=kid, kingdomName=taxon[0])

                k = Kingdom.objects.get(kingdomName=taxon[0]).kingdomid
                if not Phyla.objects.filter(kingdomid_id=k, phylaName=taxon[1]).exists():
                    pid = uuid4().hex
                    Phyla.objects.create(kingdomid_id=k, phylaid=pid, phylaName=taxon[1])

                p = Phyla.objects.get(kingdomid_id=k, phylaName=taxon[1]).phylaid
                if not Class.objects.filter(kingdomid_id=k, phylaid_id=p, className=taxon[2]).exists():
                    cid = uuid4().hex
                    Class.objects.create(kingdomid_id=k, phylaid_id=p, classid=cid, className=taxon[2])

                c = Class.objects.get(kingdomid_id=k, phylaid_id=p, className=taxon[2]).classid
                if not Order.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderName=taxon[3]).exists():
                    oid = uuid4().hex
                    Order.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid=oid,
                                         orderName=taxon[3])

                o = Order.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderName=taxon[3]).orderid
                if not Family.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o,
                                             familyName=taxon[4]).exists():
                    fid = uuid4().hex
                    Family.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid=fid,
                                          familyName=taxon[4])

                f = Family.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o,
                                       familyName=taxon[4]).familyid
                if not Genus.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f,
                                            genusName=taxon[5]).exists():
                    gid = uuid4().hex
                    Genus.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f,
                                         genusid=gid, genusName=taxon[5])

                g = Genus.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f,
                                      genusName=taxon[5]).genusid
                if not Species.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o,
                                              familyid_id=f, genusid_id=g, speciesName=taxon[6]).exists():
                    sid = uuid4().hex
                    Species.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f,
                                           genusid_id=g, speciesid=sid, speciesName=taxon[6])
                s = Species.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f,
                                        genusid_id=g, speciesName=taxon[6]).speciesid
                # skip this row if the named otu already exists (we handle unclassifieds into ISV's later)
                if taxon[7] is not "unclassified":  # unclassified check is important
                    if OTU_99.objects.filter(otuName=taxon[7], kingdomid=k, phylaid=p, classid=c, orderid=o, familyid=f, genusid=g, speciesid=s).exists():
                        # if OTU with this name exists, matching parent taxa as well, we can safely skip since its already fully in the database
                        continue

                haveKing = True  # if we don't have a kingdom, jump to unclassified handling with sequence
                # if we have no kingdom AND no sequence, forget about this entry, we have no means of creating a key
                if taxon[0] is "unclassified" or taxon[0] is "unknown":  # mothur puts unknown for unclassified kingdoms
                    haveKing = False
                # row must contain a sequence for otu mapping to be worthwhile
                if haveSeq:
                    if taxon[7].startswith('gg'):
                        if taxon[7].startswith('gg_'):
                            pass
                        else:
                            taxon[7].replace('gg', 'gg_')

                    otuName = taxon[7]
                    # check if sequence is already in use, if not, make a new otu
                    if not OTU_99.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o,
                                                 familyid_id=f, genusid_id=g, speciesid_id=s,
                                                 otuSeq=row[colLabels["Seq"]]).exists():
                        # can assume this file has sequence data, so we check that first
                        # sequence-taxa combo does not exist yet, verify we have a valid otu name then make otu
                        if otuName is "unclassified":
                            otuName = 'isv_' + str(numUnclass)
                            numUnclass += 1
                            while OTU_99.objects.filter(otuName=otuName).exists():  # guarantee no duplicate names
                                otuName = 'isv_' + str(numUnclass)
                                numUnclass += 1
                        oid = uuid4().hex
                        while OTU_99.objects.filter(otuid=oid).exists():  # guarantee no duplicate ids
                            oid = uuid4().hex
                        OTU_99.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o,
                                              familyid_id=f, genusid_id=g, speciesid_id=s, otuid=oid,
                                              otuSeq=row[colLabels["Seq"]], otuName=otuName)
                        newOtuList.append(oid)
                    else:
                        # otu using this sequence-taxa combo exists, lets make sure our name matches
                        foundOtu = OTU_99.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o,
                                                      familyid_id=f, genusid_id=g, speciesid_id=s,
                                                      otuSeq=row[colLabels["Seq"]])
                        if otuName != foundOtu.otuName:
                            if foundOtu.otuName is "unclassified":
                                # update database otuname with name from file
                                foundOtu.otuName = otuName
                                foundOtu.save()
                            else:
                                # update name in file to match database
                                newRowList = []
                                # make and save line changes in dictionary value
                                for ind in range(
                                        colLabels["Taxonomy"]):  # range from 0 -> N-1  (does not include taxa)
                                    newRowList.append(row[ind])
                                rawTaxonList = row[colLabels["Taxonomy"]].split(";")
                                rawOtuName = rawTaxonList[7]
                                otuSplit = rawOtuName.split("(")
                                otuSplit.pop(0)
                                newOtuName = "" + otuName
                                for remaining in otuSplit:
                                    newOtuName += "(" + remaining
                                rawTaxonList[7] = newOtuName
                                newRow = ""
                                for rawTaxon in rawTaxonList:
                                    newRow += rawTaxon + ";"
                                newRowList.append(newRow)
                                editedLinesDict[curLine] = newRowList

                else:
                    # we do not have sequence data to work with
                    if not haveKing:
                        # no sequence, unknown kingdom. This is not a useful line
                        pass
                    else:
                        # we at least have a kingdom, no sequence though
                        # without sequence data, otu name is not particularly helpful
                        # check species existence, if the name + parent combo does not exist, make it
                        if not Species.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f,
                                                genusid_id=g, speciesName=taxon[6]).exists():
                            # species with this taxonomy does not currently exist in database, add it
                            s = Species.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f,
                                                genusid_id=g, speciesName=taxon[6])
                            s.save()
                        else:
                            # species already exists, nothing to do here
                            pass

                        # check if sequence is already in use, if not, make a new otu
                        #if not OTU_99.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o,
                                                     #familyid_id=f, genusid_id=g, speciesid_id=s,
                                                     #otuSeq=row[colLabels["Seq"]]).exists():

                # index for use with writing file back later
                curLine += 1


        ## write taxonomy to cons.taxonomy file
        # finished looping through csv, write it again so we can update unclassified otu's
        with open(Document.name, 'w') as writeFile:
            writer = csv.writer(writeFile, delimiter='\t')    # overwrite old file with updated names and such
            writer.writerow(firstRow)
            rowNum = 0
            for docRow in docLines:
                if rowNum in editedLinesDict.keys():
                    writer.writerow(editedLinesDict[rowNum])
                else:
                    writer.writerow(docRow)  # we kept lines in memory, could potentially have changed them
                rowNum += 1
            # writer.writerow(["Hey look we made it"])

        if newOtuList:
            updateKoOtuList(newOtuList, stopList, PID, RID)

    except Exception as e:
        print "Parse_taxonomy error: ", e
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)


def parse_biom(biom, taxa, sequence, project_id, refDict, stopList, PID, RID):
    # does this function assume use of sequence or count values for column 1 (id, ?, taxa)
    debug("Parse_biom: Starting")
    try:
        global stage, perc

        # get current number of unclassified seqs
        numUnclass = OTU_99.objects.filter(otuName__startswith='isv_').count()  # assert no gaps exist
        # although its fine if a gap does exist, just an isv number we won't ever use
        # the other problem being collisions if gaps exist, but we detect those now
        newOtuList = []

        # read files
        stage = "Step 4 of 5: Parsing biom file..."
        perc = 0

        debug("Reading taxa:", taxa)
        #print "AbundDict:", abundDict.keys()
        #for k in abundDict.keys():
        #    print "\tKey:", k, "taxa:", abundDict[k]

        # Now we have abundances of local taxa ids mapped to the sample names which contain them
        # so we need to read taxa file now, its a tab delimited file with three columns: id, taxonomy, and confidence

        # the taxonomy itself is a ';' delimited string, with each part headed by l__, where l is the first letter of the taxa level
        # so we'll build a dictionary where k:v is taxaid : levelDict (ie 2b67a:{k:v, k:v})
        # levelDict uses k:v of level : name (ie kingdom:bacteria)
        # currently no use for confidence values for our purposes, maybe have a minimum?
        minConfidence = 0.97    # TODO 1.3 get this value from the user

        # metrics for tracking taxonomy behavior with this upload
        mappedCount = 0
        madeCount = 0
        madeTotal = 0

        taxaFile = open(taxa, 'r')
        lines = taxaFile.readlines()    # load the lines of tsv taxa file into memory
        maxLine = len(lines)
        curLine = 0
        taxaDict = {}
        for line in lines:  # iterate through each line of taxa data and make sure each otu is represented in database
            parts = line.split("\t")
            myID = parts[0]
            myNames = parts[1].split("; ")
            myConfidence = parts[2]
            myNameList = []
            for name in myNames:
                nameSplit = name.split("__")
                if len(nameSplit) > 1:
                    myNameList.append(nameSplit[1])  # for otu support, this nameList needs to always go to otu level
                    # using unclassified whenever a level is missing

            depth = len(myNameList)

            if myConfidence < minConfidence:
                # not confident enough, likely just a problem with the lowest listed level, switch that to unclassified
                myNameList[depth-1] = "unclassified"

            while depth < 7:
                myNameList.append("unclassified")
                depth = len(myNameList)
            if depth < 8 or myNameList[7] == "unclassified":
                # append new ISV tag

                otuName = 'isv_' + str(numUnclass)  # TODO 1.4 can put this otuName maker into a function (parse_taxa)
                numUnclass += 1
                while OTU_99.objects.filter(otuName=otuName).exists():  # guarantee no duplicate names
                    otuName = 'isv_' + str(numUnclass)
                    numUnclass += 1

                # made the name, now either add it or replace the old one
                if depth < 8:
                    myNameList.append(otuName)
                if myNameList[7] == "unclassified":
                    myNameList[7] = otuName

            # now we have the ability to make or find this taxa in the main database, then map the database id to the file id
            # then use the file id to match the database id to the abundances of each sample
            lastTaxa = None
            foundTo = -1
            # how to do this accurately, need to know exactly which level to stop at, if any level doesn't exist, make the whole set (top down)
            try:
                king = Kingdom.objects.get(kingdomName=myNameList[0])
                lastTaxa = king.kingdomid
                foundTo += 1
                phyl = Phyla.objects.get(phylaName=myNameList[1], kingdomid=king)   # TODO 1.3 seriously need to refactor this, kingdomid is a kingdom
                lastTaxa = phyl.phylaid
                foundTo += 1
                clas = Class.objects.get(className=myNameList[2], phylaid=phyl, kingdomid=king)
                lastTaxa = clas.classid
                foundTo += 1
                orde = Order.objects.get(orderName=myNameList[3], classid=clas, phylaid=phyl, kingdomid=king)
                lastTaxa = orde.orderid
                foundTo += 1
                fami = Family.objects.get(familyName=myNameList[4], orderid=orde, classid=clas, phylaid=phyl, kingdomid=king)
                lastTaxa = fami.familyid
                foundTo += 1
                genu = Genus.objects.get(genusName=myNameList[5], familyid=fami, orderid=orde, classid=clas, phylaid=phyl, kingdomid=king)
                lastTaxa = genu.genusid
                foundTo += 1
                spec = Species.objects.get(speciesName=myNameList[6], genusid=genu, familyid=fami, orderid=orde, classid=clas, phylaid=phyl, kingdomid=king)
                lastTaxa = spec.speciesid
                foundTo += 1
                otu9 = OTU_99.objects.get(otuName=myNameList[7], speciesid=spec, genusid=genu, familyid=fami, orderid=orde, classid=clas, phylaid=phyl, kingdomid=king)
                lastTaxa = otu9.otuid
                foundTo += 1
            except ObjectDoesNotExist:
                # this exception catches missing taxa, use foundTo to start building new taxa, keep going down to depth
                # then set lastTaxa to the id of the lowest new value
                if foundTo <= -1:
                    # new kingdom, I don't find this case likely but heck, we'll cover it
                    king = Kingdom.objects.create(kingdomName=myNameList[0], kingdomid=uuid4().hex)
                    king.save()
                    lastTaxa = king.kingdomid
                if foundTo <= 0:
                    # new phyla
                    phyl = Phyla.objects.create(phylaName=myNameList[1], phylaid=uuid4().hex, kingdomid=king)
                    phyl.save()
                    lastTaxa = phyl.phylaid
                if foundTo <= 1:
                    # new class
                    clas = Class.objects.create(className=myNameList[2], classid=uuid4().hex, kingdomid=king, phylaid=phyl)
                    clas.save()
                    lastTaxa = clas.classid
                if foundTo <= 2:
                    # new order
                    orde = Order.objects.create(orderName=myNameList[3], orderid=uuid4().hex, kingdomid=king, phylaid=phyl, classid=clas)
                    orde.save()
                    lastTaxa = orde.orderid
                if foundTo <= 3:
                    # new family
                    fami = Family.objects.create(familyName=myNameList[4], familyid=uuid4().hex, kingdomid=king, phylaid=phyl, classid=clas, orderid=orde)
                    fami.save()
                    lastTaxa = fami.familyid
                if foundTo <= 4:
                    # new genus
                    genu = Genus.objects.create(genusName=myNameList[5], genusid=uuid4().hex, kingdomid=king, phylaid=phyl, classid=clas, orderid=orde, familyid=fami)
                    genu.save()
                    lastTaxa = genu.genusid
                if foundTo <= 5:
                    # new species
                    spec = Species.objects.create(speciesName=myNameList[6], speciesid=uuid4().hex, kingdomid=king, phylaid=phyl, classid=clas, orderid=orde, familyid=fami, genusid=genu)
                    spec.save()
                    lastTaxa = spec.speciesid
                if foundTo <= 6:
                    # new otu
                    oid = uuid4().hex
                    while OTU_99.objects.filter(otuid=oid).exists():  # guarantee no duplicate ids, a bit overkill since DB will error if id matches on create
                        oid = uuid4().hex
                    otu9 = OTU_99.objects.create(otuName=myNameList[7], otuid=oid, kingdomid=king, phylaid=phyl, classid=clas, orderid=orde, familyid=fami, genusid=genu, speciesid=spec)
                    otu9.save()
                    lastTaxa = otu9.otuid
                    newOtuList.append(oid)

            taxaDict[myID] = lastTaxa
            if depth-1 == foundTo:
                mappedCount += 1
            else:
                madeCount += 1
                madeTotal += depth-foundTo-1

            # update perc for progress tracking, this is the first major step so we'll track out of .33 for this loop
            # spending most of upload at 20, then a jump to 33, 66, snag there until done (loops not iterating?)
            curLine += 1
            perc = curLine*1.0 / maxLine * 60.0

        # TODO 1.4 do we want this to print always? like debug controls printouts that tell us where things crashed, this is very useful for database size tracking
        debug("Finished taxonomy handling, found", mappedCount, "existing entries and made", madeCount, "new ones (", madeTotal, " including parents)")

        debug("Reading biom:", biom)

        # lets try h5py instead, skip R entirely
        biomDict = h5py.File(biom, 'r')
        obs = biomDict['observation']
        # obs_group = obs['group-metadata']  # group, 0 members
        obs_ids = obs['ids'][:]  # dataset, type "|0"
        obs_matrix = obs['matrix']  # group, 3 members
        # obs_meta = obs['metadata']  # group, 0 members

        samp = biomDict['sample']
        # samp_group = samp['group-metadata'] # group, 0 members
        samp_ids = samp['ids'][:]  # dataset, type "|0"
        samp_matrix = samp['matrix']  # group, 3 members
        # samp_meta = samp['meta']    # group, 0 members

        # so obs ids are taxa ids, to be used with taxa.tsv to find taxa names)
        # obs_matrix contains idptr, indices, and data
        # idptr has one row for each taxa, with a value listing the starting index of that taxon's listing in indices and data
        # (stop at next rows index, if next row does not exist, this entry is the end of the file, skip)
        # indices point to a sample number, for a given sample containing the taxa in idptr
        # data gives abundances of a given taxa in a given sample (the sample can be found from indices, using samp_ids)

        # need to get data grouped by sample id
        # so we make a dictionary, k:v of samp_id:taxa_dict
        # taxa_dict k:v is taxa_id (from obs_ids) : count_for_sample
        # samp_id is found by taking the index from obs_indices and accessing samp_ids with it
        # count_for_sample is found by iterating through this taxa's index range in obs_matrix['data'], match with sample index
        # ( so this loop will go from obs_matrix['indptr'], iterating over each TAXA, pointing back towards the sample when saving )

        # This section parses the hdf5 data into a dictionary, keyed by sample name (which should exist in the meta file)
        # also describes taxon abundances, using a local taxaid (not matching database, instead pairing with taxa and
        # sequence files, which we have at this point as 'taxa' and 'sequence' (simple enough)
        abundDict = {}

        taxCount = 0
        taxSize = 0
        maxTaxa = len(obs_ids)
        for ptr in range(maxTaxa):
            myTaxID = obs_ids[ptr]
            # get the real taxaid from taxaDict
            myTaxID = taxaDict[myTaxID]
            # get start and end points of the range for this taxon
            startInd = obs_matrix['indptr'][ptr]
            endInd = obs_matrix['indptr'][ptr + 1]
            for taxptr in range(startInd, endInd):
                mySampID = samp_ids[obs_matrix['indices'][taxptr]]
                myAbund = obs_matrix['data'][taxptr]
                if mySampID not in abundDict.keys():
                    abundDict[mySampID] = {}
                abundDict[mySampID][myTaxID] = myAbund
                taxSize += 1
            taxCount += 1
            perc = 60 + taxCount*1.0 / maxTaxa * 10.0

        debug("Finished parsing HDF5 biom, found", taxCount, "taxa with avg", float(taxSize) / float(taxCount),
              "entries per taxa")

        # So now we have taxaDict for pairing biom taxaids with database ids that definitely exist
        # and we can pair that with the sample_name keyed dictionary containing all biom taxaids with their abundances in specified samples
        # The goal is now to create profile objects, which map taxa in the myphylo database with abundances in specific samples in the same database
        # We should have samples in the database courtesy of parsing the meta file, we can find the right ones using abundDict.keys() and the project info
        myProj = Project.objects.get(projectid=project_id)
        mySamples = Sample.objects.filter(projectid=myProj)
        missingSamples = []
        sampCount = len(mySamples)
        curSamp = 0
        for samp in mySamples:
            # for each sample, get the name, use it to access abundDict
            sampName = samp.sample_name
            # verify this sample is in the biom data, skip if its not, after saving in missingSamples to report back
            if sampName not in abundDict.keys():
                missingSamples.append(sampName)
                continue
            myTaxa = abundDict[sampName]    # k:v = taxaid:count
            # create profile object for each taxa in this sample
            # we don't have otu data with this, at least not always
            myReads = 0
            for tax in myTaxa.keys():
                # get otu object by id
                myOTU = OTU_99.objects.get(otuid=tax)
                count = myTaxa[tax]
                replaceType = refDict[samp.sampleid]
                # Parse_profile code section, we should respect replaceType
                if replaceType == 'replace':
                    Profile.objects.filter(sampleid=samp).delete()
                    # TODO 1.4 refactor the ids, either use the actual id or change the label to just sample (since its
                    # a foreign key to a sample object)

                if replaceType == 'new' or replaceType == 'replace':
                    Profile.objects.create(projectid=myProj, sampleid=samp, kingdomid=myOTU.kingdomid,
                    phylaid=myOTU.phylaid, classid=myOTU.classid, orderid=myOTU.orderid, familyid=myOTU.familyid,
                    genusid=myOTU.genusid, speciesid=myOTU.speciesid, otuid=myOTU, count=count)

                if replaceType == 'append':
                    if Profile.objects.filter(projectid=myProj, sampleid=samp, kingdomid=myOTU.kingdomid,
                    phylaid=myOTU.phylaid, classid=myOTU.classid, orderid=myOTU.orderid, familyid=myOTU.familyid,
                    genusid=myOTU.genusid, speciesid=myOTU.speciesid, otuid=myOTU).exists():
                        t = Profile.objects.get(projectid=myProj, sampleid=samp, kingdomid=myOTU.kingdomid,
                        phylaid=myOTU.phylaid, classid=myOTU.classid, orderid=myOTU.orderid, familyid=myOTU.familyid,
                        genusid=myOTU.genusid, speciesid=myOTU.speciesid, otuid=myOTU)
                        old = t.count
                        count += old
                        t.count = count
                        t.save()
                    else:
                        Profile.objects.create(projectid=myProj, sampleid=samp, kingdomid=myOTU.kingdomid,
                        phylaid=myOTU.phylaid, classid=myOTU.classid, orderid=myOTU.orderid, familyid=myOTU.familyid,
                        genusid=myOTU.genusid, speciesid=myOTU.speciesid, otuid=myOTU, count=count)

                myReads += count

            # update reads field for sample to this parse
            samp.reads = myReads
            samp.save()

            curSamp += 1
            perc = 70 + curSamp*1.0 / sampCount * 28.0

        # find samples which were present in biom but not in meta (we have the reverse from earlier)
        missingBiom = []
        for sampName in abundDict.keys():
            found = False
            for samp in mySamples:
                if sampName == samp.sample_name:
                    found = True
            if not found:
                missingBiom.append(sampName)
        perc = 99.0
        # update koOTUList takes new OTU from this upload and attempts to pair them with gene lists (ko) for faster
        # queries during taxonomic profile table searches
        if newOtuList:
            updateKoOtuList(newOtuList, stopList, PID, RID)
        perc = 100.0
        # report back with missingSample and missingBiom lists so user knows which data is missing where (or mismatched)
        return missingSamples, missingBiom

    except Exception as e:
        print "Parse_biom error: ", e
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)


def updateKoOtuList(otuList, stopList, PID, RID):  # for during a new upload
    # get otuList from project data (argument): Should be a list of all  NEW otu's in project
    # get all ko genelists and search for matching otus (out of new otus from related upload)
    # add matching otus to related koOtuList entries
    try:
        # get every single ko_entry object (or equivalent data from levels 1-3)
        koList = []
        records = ko_lvl1.objects.using('picrust').all()
        for record in records:
            koList.extend(record.ko_entry_set.values_list('ko_orthology', flat=True))

        records = ko_lvl2.objects.using('picrust').all()
        for record in records:
            koList.extend(record.ko_entry_set.values_list('ko_orthology', flat=True))

        records = ko_lvl3.objects.using('picrust').all()
        for record in records:
            koList.extend(record.ko_entry_set.values_list('ko_orthology', flat=True))

        koList.extend(ko_entry.objects.all())

        if stopList[PID] == RID:
            return

        curOtu = 0
        if koList:
            koOtuDict = {}
            for otu in otuList:
                curTotal = 0
                creTotal = 0
                if PICRUSt.objects.using('picrust').filter(otuid=otu).exists():
                    qs = PICRUSt.objects.using('picrust').filter(otuid=otu)
                    geneList = qs[0].geneList
                    geneList = geneList.replace("[", "")
                    geneList = geneList.replace("]", "")
                    geneList = geneList.replace("'", "")
                    geneList = geneList.replace(" ", "")
                    genes = geneList.split(",")
                    for gene in genes:
                        if gene in koList:
                            # get or make koOtuList object for this ko
                            if gene not in koOtuDict.keys():
                                koOtuDict[gene] = []
                                creTotal += 1
                            koOtuDict[gene].append(otu)
                            curTotal += 1

                    curOtu += 1
                    if stopList[PID] == RID:
                        return

            for ko in koOtuDict.keys():
                if not koOtuList.objects.filter(koID=ko).exists():
                    # koOtuList did not exist from main pop function, can just save list on new entry
                    newKo = koOtuList.objects.create(koID=ko)
                    newKo.otuList = koOtuDict[ko]
                    newKo.save()
                else:
                    myKo = koOtuList.objects.get(koID=ko)
                    # slightly trickier, as existing entry likely has a unicode comma split string, add comma then list
                    oldList = myKo.otuList  # other thing to consider is duplicate entries, need to rebuild entirely
                    # need to convert oldList from full string
                    newList = koOtuDict[ko]
                    myKo.otuList = oldList + "," + str(newList)
                    myKo.save()


    except Exception as er:
        print "Problem with updateKOL (M) ", er
    return


def parse_profile(file3, file4, p_uuid, refDict, stopList, PID, RID):   # TODO 1.4 this pipeline needs refactored
    debug("Parsing profile!")
    try:
        global stage, perc
        stage = "Step 5 of 5: Parsing shared file..."
        perc = 0

        # taxonomy file
        data1 = genfromtxt(file3, delimiter='\t', dtype=None, autostrip=True, encoding=None)  # getting empty dataset...
        df1 = pd.DataFrame(data1[1:, 1:], index=data1[1:, 0], columns=data1[0, 1:])  # too many indices crash
        df1 = df1[df1.index != 'False']
        a = df1.columns.values.tolist()
        a = [x for x in a if x != 'False']
        df1 = df1[a]
        # file3.close() # this line breaks when using a filename instead of a file object

        # shared file
        data2 = genfromtxt(file4, delimiter='\t', dtype=None, autostrip=True, encoding=None)
        arr2 = data2.T
        arr2 = np.delete(arr2, 0, axis=0)
        arr2 = np.delete(arr2, 1, axis=0)
        df2 = pd.DataFrame(arr2[1:, 1:], index=arr2[1:, 0], columns=arr2[0, 1:])
        df2 = df2[df2.index != 'False']
        b = df2.columns.values.tolist()
        b = [x for x in b if x != 'False']
        df2 = df2[b]
        file4.close()

        df3 = pd.merge(df1, df2, left_index=True, right_index=True, how='inner')
        df3['Taxonomy'].replace(to_replace='(\(.*?\)|k__|p__|c__|o__|f__|g__|s__|otu__)', value='', regex=True,
                                inplace=True)
        df3.reset_index(drop=True, inplace=True)
        del df1, df2

        total, columns = df3.shape
        sampleList = df3.columns.values.tolist()
        try:
            sampleList.remove('Seq')
        except:
            pass
        try:
            sampleList.remove('Taxonomy')
        except:
            pass
        try:
            sampleList.remove('Size')   # do we need this?
        except:
            pass

        # need to get rid of any old taxonomy data if sample type is set to replace
        for name in sampleList:
            sample = Sample.objects.filter(projectid=p_uuid).get(sample_name=name)
            sampid = sample.sampleid
            replaceType = refDict[sampid]
            if replaceType == 'replace':
                Profile.objects.filter(sampleid=sampid).delete()
            if stopList[PID] == RID:
                return


        idList = []
        step = 0.0
        for index, row in df3.iterrows():
            perc = int(step / total * 100)
            taxon = str(row['Taxonomy'])
            taxon = taxon[:-1]
            taxaList = taxon.split(';')

            k = taxaList[0]
            p = taxaList[1]
            c = taxaList[2]
            o = taxaList[3]
            f = taxaList[4]
            g = taxaList[5]

            if len(taxaList) < 7:
                taxaList.append('unclassified')
            if len(taxaList) < 8:
                taxaList.append('unclassified')

            s = taxaList[6]

            otu = taxaList[7]

            try:
                t_kingdom = Kingdom.objects.get(kingdomName=k)
                t_phyla = Phyla.objects.get(kingdomid=t_kingdom, phylaName=p)
                t_class = Class.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, className=c)
                t_order = Order.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderName=o)
                t_family = Family.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order,
                                              familyName=f)
                t_genus = Genus.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order,
                                            familyid=t_family, genusName=g)
                t_species = Species.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order,
                                                familyid=t_family, genusid=t_genus, speciesName=s)
            except Exception as e:
                print "Error finding taxa: ", e     # currently getting here via "unknown" kingdom
                print taxaList

            t_otu = ''
            try:
                if k == 'unclassified':
                    t_otu = OTU_99.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order,
                                               familyid=t_family, genusid=t_genus, speciesid=t_species, otuSeq=row['Seq'])
                else:
                    t_otu = OTU_99.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order,
                                               familyid=t_family, genusid=t_genus, speciesid=t_species, otuName=otu)
            except Exception as e:
                # otuID needs to be unique, apparently not specifying allows duplicates? based on name maybe
                oid = uuid4().hex
                t_otu = OTU_99.objects.create(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order,
                                              familyid=t_family, genusid=t_genus, speciesid=t_species,
                                              otuName=otu, otuid=oid)
                t_otu.save()

            if stopList[PID] == RID:
                return

            for name in sampleList:
                try:
                    count = int(row[name])
                    project = Project.objects.get(projectid=p_uuid)
                    sample = Sample.objects.filter(projectid=p_uuid).get(sample_name=name)
                    sampid = sample.sampleid
                    idList.append(sampid)
                    replaceType = refDict[sampid]


                    if count > 0:
                        if replaceType == 'new' or replaceType == 'replace':
                            Profile.objects.create(projectid=project, sampleid=sample, kingdomid=t_kingdom,
                                                   phylaid=t_phyla, classid=t_class, orderid=t_order, familyid=t_family,
                                                   genusid=t_genus, speciesid=t_species, otuid=t_otu, count=count)

                        if replaceType == 'append':
                            if Profile.objects.filter(projectid=project, sampleid=sample, kingdomid=t_kingdom,
                                                      phylaid=t_phyla, classid=t_class, orderid=t_order,
                                                      familyid=t_family, genusid=t_genus, speciesid=t_species,
                                                      otuid=t_otu).exists():
                                t = Profile.objects.get(projectid=project, sampleid=sample, kingdomid=t_kingdom,
                                                        phylaid=t_phyla, classid=t_class, orderid=t_order,
                                                        familyid=t_family, genusid=t_genus, speciesid=t_species,
                                                        otuid=t_otu)
                                old = t.count
                                new = old + int(count)
                                t.count = new
                                t.save()
                            else:
                                Profile.objects.create(projectid=project, sampleid=sample, kingdomid=t_kingdom,
                                                       phylaid=t_phyla, classid=t_class, orderid=t_order,
                                                       familyid=t_family, genusid=t_genus, speciesid=t_species,
                                                       otuid=t_otu, count=count)
                except Exception as exc:
                    print 'Sample could not be added: ', str(name), " because ", exc
                if stopList[PID] == RID:
                    return

            step += 1.0

        idUnique = list(set(idList))
        for ID in idUnique:
            sample = Sample.objects.get(sampleid=ID)
            reads = Profile.objects.filter(sampleid=sample).aggregate(count=Sum('count'))
            sample.reads = reads['count']
            sample.save()
            if stopList[PID] == RID:
                return

        stage = "Step 1 of 5: Parsing project file..."

    except Exception as e:
        print "Error with shared file parsing: ", e
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)

    debug("Done parsing profile")


def repStop(request):
    debug("Repstop!")
    try:
        # cleanup in repro project!
        # reset WIP flag
        ids = request.POST["ids"]
        if isinstance(ids, list):
            refList = Reference.objects.all().filter(refid__in=ids)
        else:
            refList = Reference.objects.all().filter(refid=ids)
        for ref in refList:
            curProj = Project.objects.get(projectid=ref.projectid)
            curProj.wip = False
            curProj.save()
    except Exception as e:
        print "Repstop errored:", e
    return "Stopped"


def mothurRestore(dest):
    # if reprocess is cancelled / fails, replace new mothur file with old
    try:
        shutil.copyfile(str(dest)+'/mothur.batch.old', str(dest)+'/mothur.batch')
    except:
        pass
    try:
        shutil.copyfile(str(dest)+'/dada2.R.old', str(dest)+'/dada2.R')
    except:
        pass


def bubbleFiles(dest):
    # recursively find all files in dest directory
    # move each file to dest, delete subdirectories
    for root, dirnames, filenames in os.walk(dest):
        for filename in fnmatch.filter(filenames, '*'):
            shutil.move(os.path.join(root, filename), os.path.join(dest, filename))
    return


def reanalyze(request, stopList):
    ### create savepoint
    sid = transaction.savepoint()

    try:
        global rep_project

        mothurdest = 'mothur/temp/'
        if not os.path.exists(mothurdest):
            os.makedirs(mothurdest)

        ids = request.POST["ids"]
        processors = request.POST["processors"]
        platform = request.POST["platform"]
        RID = request.POST['RID']
        PID = 0  # change if adding additional data threads

        replaceBatch = False
        mFile = None

        try:
            mFile = request.FILES['mothurFile']
            replaceBatch = True
        except Exception as e:
            pass  # file not found, use default mothur batch

        if stopList[PID] == RID:
            return repStop(request)

        if isinstance(ids, list):
            refList = Reference.objects.all().filter(refid__in=ids)
        else:
            refList = Reference.objects.all().filter(refid=ids)

        if stopList[PID] == RID:
            return repStop(request)

        for ref in refList:
            curProj = Project.objects.get(projectid=ref.projectid.projectid)
            curProj.wip = True
            curProj.save()

            rep_project = 'myPhyloDB is currently reprocessing project: ' + str(ref.projectid.project_name)
            dest = ref.path
            source = ref.source
            origAuthor = ref.author

            try:
                if platform == 'mothur':
                    # copy old mothur file with new name
                    try:
                        shutil.copyfile(str(dest)+'/mothur.batch', str(dest)+'/mothur.batch.old')
                    except:
                        pass
                    if replaceBatch:
                        # drop new mothur file onto old
                        functions.handle_uploaded_file(mFile, dest, 'mothur.batch')
                else:
                    # copy old R file with new name
                    try:
                        shutil.copyfile(str(dest)+'/dada2.R', str(dest)+'/dada2.R.old')
                    except:
                        pass
                    if replaceBatch:
                        # drop new mothur file onto old
                        functions.handle_uploaded_file(mFile, dest, 'dada2.R')
            except:
                pass

            if stopList[PID] == RID:
                mothurRestore(dest)
                return repStop(request)

            if source == '454_sff':
                ls_dir = os.listdir(dest)
                for afile in ls_dir:
                    srcStr = str(dest) + '/' + str(afile)
                    destStr = str(mothurdest) + '/' + str(afile)
                    shutil.copyfile(srcStr, destStr)

            if source == '454_fastq':
                file_list = []
                for afile in glob.glob(r'% s/*.fna' % dest):
                    file_list.append(afile)

                tempList = []
                if file_list.__len__() > 1:
                    for oldPath in file_list:
                        name = os.path.basename(oldPath)
                        newPath = 'mothur/temp/' + str(name)
                        shutil.copy(oldPath, newPath)
                        tempList.append(newPath)
                    inputList = "-".join(tempList)

                    if stopList[PID] == RID:
                        mothurRestore(dest)
                        return repStop(request)

                    if os.name == 'nt':
                        os.system('"mothur\\mothur-win\\mothur.exe \"#merge.files(input=%s, output=mothur\\temp\\temp.fasta)\""' % inputList)
                    else:
                        os.system("mothur/mothur-linux/mothur \"#merge.files(input=%s, output=mothur/temp/temp.fasta)\"" % inputList)

                else:
                    fPath = 'mothur/temp' + os.path.basename(file_list[0])
                    shutil.copyfile(file_list[0], fPath)

                file_list = []
                for afile in glob.glob(r'% s/*.qual' % dest):
                    file_list.append(afile)

                tempList = []
                if file_list.__len__() > 1:
                    for oldPath in file_list:
                        name = os.path.basename(oldPath)
                        newPath = 'mothur/temp/' + str(name)
                        shutil.copy(oldPath, newPath)
                        tempList.append(newPath)
                    inputList = "-".join(tempList)

                    if stopList[PID] == RID:
                        mothurRestore(dest)
                        return repStop(request)

                    if os.name == 'nt':
                        os.system('"mothur\\mothur-win\\mothur.exe \"#merge.files(input=%s, output=mothur\\temp\\temp.qual)\""' % inputList)
                    else:
                        os.system("mothur/mothur-linux/mothur \"#merge.files(input=%s, output=mothur/temp/temp.qual)\"" % inputList)

                else:
                    fPath = 'mothur/temp' + os.path.basename(file_list[0])
                    shutil.copyfile(file_list[0], fPath)

                for oldPath in glob.glob(r'% s/*.oligos' % dest):
                    name = os.path.basename(oldPath)
                    newPath = 'mothur/temp/' + str(name)
                    shutil.copyfile(oldPath, newPath)

                for oldPath in glob.glob(r'% s/*.batch' % dest):
                    name = os.path.basename(oldPath)
                    newPath = 'mothur/temp/' + str(name)
                    shutil.copyfile(oldPath, newPath)

            if source == 'miseq':
                ls_dir = os.listdir(dest)
                for afile in ls_dir:
                    srcStr = str(dest) + '/' + str(afile)
                    destStr = str(mothurdest) + '/' + str(afile)
                    shutil.copyfile(srcStr, destStr)

                for subdir, dirs, files in os.walk(mothurdest):
                    for zipp in files:
                        filepath= os.path.join(subdir, zipp)
                        try:
                            zip_ref = zipfile.ZipFile(filepath, 'r')
                            zip_ref.extractall(mothurdest)
                            zip_ref.close()
                        except:
                            try:
                                tar = tarfile.open(filepath, 'r')
                                tar.extractall(path=mothurdest)
                                tar.close()
                            except:
                                pass

            if stopList[PID] == RID:
                mothurRestore(dest)
                return repStop(request)

            if platform == 'mothur':
                try:
                    with open("% s/mothur.batch" % dest, 'r+') as bat:
                        with open("% s/mothur.batch" % mothurdest, 'wb+') as destination:
                            for line in bat:
                                if "processors" in line:
                                    line = line.replace('processors=X', 'processors='+str(processors))
                                destination.write(line)
                except Exception as e:
                    print("Error with batch file: ", e)
            else:
                shutil.copy('% s/dada2.R' % dest, '% s/dada2.R' % mothurdest)

            shutil.copy('% s/final_meta.xlsx' % dest, '% s/final_meta.xlsx' % mothurdest)

            if stopList[PID] == RID:
                mothurRestore(dest)
                return repStop(request)

            if not os.path.exists(dest):
                os.makedirs(dest)

            shutil.copy('% s/final_meta.xlsx' % mothurdest, '% s/final_meta.xlsx' % dest)
            if stopList[PID] == RID:
                mothurRestore(dest)
                transaction.savepoint_rollback(sid)
                return repStop(request)

            bubbleFiles(mothurdest)

            # subQueue()
            metaName = 'final_meta.xlsx'
            metaFile = '/'.join([dest, metaName])

            with open(metaFile, 'rb') as f:
                p_uuid, pType, num_samp = projectid(f)

            curUser = origAuthor.username
            parse_project(metaFile, p_uuid, curUser)

            if stopList[PID] == RID:
                mothurRestore(dest)
                transaction.savepoint_rollback(sid)
                return repStop(request)

            raw = True
            userID = str(origAuthor.id)
            refDict = parse_sample(metaFile, p_uuid, pType, num_samp, dest, raw, source, userID, stopList, RID, PID)

            if stopList[PID] == RID:
                mothurRestore(dest)
                transaction.savepoint_rollback(sid)
                return repStop(request)

            try:
                if platform == 'mothur':
                    mothur(dest, source)
                else:
                    dada2(dest, source)

            except Exception as e:
                transaction.savepoint_rollback(sid)
                print("Error with mothur/R: " + str(e))
                return HttpResponse(
                    json.dumps({"error": "yes"}),
                    content_type="application/json"
                )

            if stopList[PID] == RID:
                mothurRestore(dest)
                transaction.savepoint_rollback(sid)  # this does not seem to cancel properly
                return repStop(request)

            with open('% s/final.cons.taxonomy' % dest, 'rb') as file3:
                parse_taxonomy(file3, stopList, PID, RID)

            if stopList[PID] == RID:
                mothurRestore(dest)
                transaction.savepoint_rollback(sid)  # this does not seem to cancel properly
                return repStop(request)

            with open('% s/final.cons.taxonomy' % dest, 'rb') as file3:
                with open('% s/final.tx.shared' % dest, 'rb') as file4:
                    parse_profile(file3, file4, p_uuid, refDict, stopList, PID, RID)

            if stopList[PID] == RID:
                mothurRestore(dest)
                transaction.savepoint_rollback(sid)  # this does not seem to cancel properly
                return repStop(request)

            # reprocess completed, delete old mothur
            try:
                if platform == 'mothur':
                    os.remove(os.path.join(dest, "mothur.batch.old"))
                else:
                    os.remove(os.path.join(dest, "dada2.R.old"))
            except:
                pass

            curProj = Project.objects.get(projectid=ref.projectid.projectid)
            curProj.wip = False
            curProj.save()

        return None

    except Exception as e:
        print "Error during reprocessing: ", e
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        transaction.savepoint_rollback(sid)
        logging.exception(myDate)

        try:
            mothurRestore(dest)
        except Exception:
            pass

        return repStop(request)

