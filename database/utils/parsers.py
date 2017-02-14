import csv
import datetime
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
import glob
from io import BytesIO
import logging
from numpy import *
import numpy as np
import openpyxl
import os
import signal
import psutil
import pandas as pd
import re
import shutil
import json
import subprocess
from uuid import uuid4

from database.models import Project, Reference, Sample, Air, Human_Associated, Microbial, Soil, Water, UserDefined, \
    Kingdom, Phyla, Class, Order, Family, Genus, Species, OTU_99, Profile

from utils_df import handle_uploaded_file, excel_to_dict
import database.queue


stage = ''
last = ''
mothurStat = ""
perc = 0
rep_project = ''
pd.set_option('display.max_colwidth', -1)
LOG_FILENAME = 'error_log.txt'

p = None


def mothur(dest, source):
    global p
    try:
        global stage, perc, mothurStat
        stage = "Step 3 of 5: Running mothur..."
        perc = 0

        if not os.path.exists('mothur/reference/align'):
            os.makedirs('mothur/reference/align')

        if not os.path.exists('mothur/reference/taxonomy'):
            os.makedirs('mothur/reference/taxonomy')

        if not os.path.exists('mothur/reference/template'):
            os.makedirs('mothur/reference/template')

        if os.name == 'nt':
            filepath = "mothur\\mothur-win\\mothur.exe mothur\\temp\\mothur.batch"
        else:
            filepath = "mothur/mothur-linux/mothur mothur/temp/mothur.batch"
        try:
            p = subprocess.Popen(filepath, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            for line in iter(p.stdout.readline, ''):
                print line
                mothurStat += line
            mothurStat = 'Previous mothur output:\n'
        except Exception as e:
            print "Mothur failed: " + str(e)

        if source == '454_sff':
            shutil.copy('mothur/temp/final.fasta', '% s/final.fasta' % dest)
            shutil.copy('mothur/temp/final.names', '% s/final.names' % dest)
            shutil.copy('mothur/temp/final.groups', '% s/final.groups' % dest)
            shutil.copy('mothur/temp/final.taxonomy', '% s/final.taxonomy' % dest)
            shutil.copy('mothur/temp/final.shared', '% s/final.shared' % dest)
            shutil.rmtree('mothur/temp')

        if source == '454_fastq':
            shutil.copy('mothur/temp/final.fasta', '% s/final.fasta' % dest)
            shutil.copy('mothur/temp/final.names', '% s/final.names' % dest)
            shutil.copy('mothur/temp/final.groups', '% s/final.groups' % dest)
            shutil.copy('mothur/temp/final.taxonomy', '% s/final.taxonomy' % dest)
            shutil.copy('mothur/temp/final.shared', '% s/final.shared' % dest)
            shutil.rmtree('mothur/temp')

        if source == 'miseq':
            shutil.copy('mothur/temp/final.fasta', '% s/final.fasta' % dest)
            shutil.copy('mothur/temp/final.names', '% s/final.names' % dest)
            shutil.copy('mothur/temp/final.groups', '% s/final.groups' % dest)
            shutil.copy('mothur/temp/final.taxonomy', '% s/final.taxonomy' % dest)
            shutil.copy('mothur/temp/final.shared', '% s/final.shared' % dest)
            shutil.rmtree('mothur/temp')

        for afile in glob.glob(r'*.logfile'):
            shutil.move(afile, dest)

    except Exception:
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)


def termP():
    global p
    try:
        parent = psutil.Process(p.pid)
        children = parent.children(recursive=True)
        for process in children:
            process.send_signal(signal.SIGTERM)
        os.kill(p.pid, signal.SIGTERM)
    except Exception as e:
        print "Error with terminate: "+str(e)
        pass


def status(request):
    global last, mothurStat
    if request.is_ajax():
        RID = request.GET['all']
        queuePos = database.queue.dataqueue.datstat(RID)
        dict = {}
        if queuePos == 0:
            # check for duplicate output
            if last != mothurStat:
                dict['mothurStat'] = mothurStat
                last = mothurStat
            else:
                dict['mothurStat'] = ''

            dict['stage'] = stage
            dict['perc'] = perc
            dict['project'] = rep_project
        else:
            if (queuePos == -1024):
                dict['stage'] = "Stopping...please be patient while we restore the database!"
            else:
                dict['stage'] = "In queue for processing, "+str(queuePos)+" requests in front of you"
            dict['perc'] = 0
            dict['project'] = rep_project
            dict['mothurStat'] = ''
        mothurStat = ""
        json_data = json.dumps(dict)
        return HttpResponse(json_data, content_type='application/json')


def projectid(Document):
    try:
        global stage, perc
        perc = 0
        stage = "Step 1 of 5: Parsing project file..."
        wb = openpyxl.load_workbook(BytesIO(Document.read()), data_only=True, read_only=False)
        ws = wb.get_sheet_by_name('Project')
        pType = ws.cell(row=6, column=3).value
        projectid = ws.cell(row=6, column=4).value
        num_samp = ws.cell(row=6, column=1).value

        if projectid:
            p_uuid = projectid
        else:
            p_uuid = uuid4().hex

        return p_uuid, pType, num_samp

    except Exception:
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)
        return None


def parse_project(Document, p_uuid, curUser):
    try:
        global stage, perc
        perc = 50
        wb = openpyxl.load_workbook(Document, data_only=True, read_only=False)
        myDict = excel_to_dict(wb, headerRow=5, nRows=1, sheet='Project')
        rowDict = myDict[0]
        rowDict.pop('num_samp')

        if not Project.objects.filter(projectid=p_uuid).exists():
            for key in rowDict.keys():
                if key == 'projectid':
                    rowDict[key] = p_uuid
            myProj = Project.objects.create(**rowDict)  # save pointer to this somehow (myProj = ?)
            # get project and set owner to current user (if project is new)
            myProj.owner = curUser
            myProj.save()
        else:
            rowDict.pop('projectid')
            Project.objects.filter(projectid=p_uuid).update(projectid=p_uuid, **rowDict)
        stage = "Step 1 of 5: Parsing project file...done"
        return "none"
    except Exception as ex:
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)
        return str(ex)


def parse_reference(p_uuid, refid, path, batch, raw, source, userid):
    try:
        author = User.objects.get(id=userid)
        align_ref = ''
        template_ref = ''
        taxonomy_ref = ''

        if raw:
            batch.seek(0)
            for row in batch:
                if "align.seqs" in row:
                    for item in row.split(','):
                        if "reference=" in item:
                            string = item.split('=')
                            align_ref = string[1].replace('mothur/reference/align/', '')
                if "classify.seqs" in row:
                    for item in row.split(','):
                        if "template=" in item:
                            string = item.split('=')
                            template_ref = string[1].replace('mothur/reference/template/', '')
                        if "taxonomy=" in item:
                            string = item.split('=')
                            taxonomy_ref = string[1].replace('mothur/reference/taxonomy/', '')
        else:
            align_ref = 'null'
            template_ref = 'null'
            taxonomy_ref = 'null'

        if not Reference.objects.filter(refid=refid).exists():
            project = Project.objects.get(projectid=p_uuid)
            Reference.objects.create(refid=refid, projectid=project, path=path, source=source, raw=raw, alignDB=align_ref, templateDB=template_ref, taxonomyDB=taxonomy_ref, author=author)
        else:
            project = Project.objects.get(projectid=p_uuid)
            Reference.objects.filter(refid=refid).update(refid=refid, projectid=project, path=path, source=source, raw=raw, alignDB=align_ref, templateDB=template_ref, taxonomyDB=taxonomy_ref, author=author)

    except Exception:
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)


def parse_sample(Document, p_uuid, pType, num_samp, dest, batch, raw, source, userID):
    try:
        global stage, perc
        stage = "Step 2 of 5: Parsing sample file..."
        perc = 0

        project = Project.objects.get(projectid=p_uuid)
        wb = openpyxl.load_workbook(Document, data_only=True, read_only=False)
        perc = 25

        dict1 = excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='MIMARKs')
        perc = 50

        if pType == 'air':
            dict2 = excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='Air')

        elif pType == 'human associated' or pType == 'human gut':
            dict2 = excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='Human Associated')

        elif pType == 'microbial':
            dict2 = excel_to_dict(wb,headerRow=6, nRows=num_samp, sheet='Microbial')

        elif pType == 'soil':
            dict2 = excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='Soil')

        elif pType == 'water':
            dict2 = excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='Water')

        else:
            dict2 = list()

        dict3 = excel_to_dict(wb, headerRow=6, nRows=num_samp, sheet='User')
        perc = 75

        idList = []
        refid = ''
        refDict = {}
        newRefID = uuid4().hex
        for i in xrange(num_samp):
            row = dict1[i]
            pathid = row['refid']

            if isinstance(pathid, unicode):
                refid = pathid
            else:
                refid = newRefID

            s_uuid = row['sampleid']
            row.pop('refid')
            row.pop('seq_method')
            row.pop('geo_loc_name')
            row.pop('lat_lon')

            parse_reference(p_uuid, refid, dest, batch, raw, source, userID)
            reference = Reference.objects.get(refid=refid)

            if Sample.objects.filter(sampleid=s_uuid).exists():
                idList.append(s_uuid)
                Sample.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, **row)
            else:
                s_uuid = uuid4().hex
                row['sampleid'] = s_uuid
                idList.append(s_uuid)
                Sample.objects.create(projectid=project, refid=reference, **row)

            refDict[s_uuid] = refid
            sample = Sample.objects.get(sampleid=s_uuid)

            row = dict2[i]
            if pType == "air":
                if not Air.objects.filter(sampleid=s_uuid).exists():
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Air.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
                else:
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Air.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)

            elif pType == "human associated":
                if not Human_Associated.objects.filter(sampleid=s_uuid).exists():
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Human_Associated.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
                else:
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Human_Associated.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)

            elif pType == "microbial":
                if not Microbial.objects.filter(sampleid=s_uuid).exists():
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Microbial.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
                else:
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Microbial.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)

            #TODO: remove extras once they are added to the excel metafile
            elif pType == "soil":
                if not Soil.objects.filter(sampleid=s_uuid).exists():
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Soil.objects.create(projectid=project, refid=reference, sampleid=sample, soil_water_cap=None,
                        soil_surf_hard=None, soil_subsurf_hard=None, soil_agg_stability=None, soil_ACE_protein=None,
                        soil_active_C=None, **row)
                else:
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Soil.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample,
                        soil_water_cap=None, soil_surf_hard=None, soil_subsurf_hard=None, soil_agg_stability=None,
                        soil_ACE_protein=None, soil_active_C=None, **row)

            elif pType == "water":
                if not Water.objects.filter(sampleid=s_uuid).exists():
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Water.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
                else:
                    row.pop('sampleid')
                    row.pop('sample_name')
                    Water.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)
            else:
                pass

            row = dict3[i]
            if not UserDefined.objects.filter(sampleid=s_uuid).exists():
                row.pop('sampleid')
                row.pop('sample_name')
                UserDefined.objects.create(projectid=project, refid=reference, sampleid=sample, **row)
            else:
                row.pop('sampleid')
                row.pop('sample_name')
                UserDefined.objects.filter(sampleid=s_uuid).update(projectid=project, refid=reference, sampleid=sample, **row)

            perc += 20/num_samp

        ### add myPhyloDB generated IDs to excel metafile
        wb = openpyxl.load_workbook(Document, data_only=False, read_only=False)
        ws = wb.get_sheet_by_name('Project')
        ws.cell(row=6, column=4).value = p_uuid

        ws = wb.get_sheet_by_name('MIMARKs')
        for i in xrange(len(idList)):
            j = i + 7
            ws.cell(row=j, column=1).value = refid
            ws.cell(row=j, column=2).value = idList[i]
            perc += 5/len(idList)

        wb.save(Document)
        return refDict

    except Exception as e:
        print e
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)
        return None


def parse_taxonomy(Document):
    try:
        global stage, perc
        stage = "Step 4 of 5: Parsing taxonomy file..."
        perc = 0
        print "Parsing taxonomy!"
        f = csv.reader(Document, delimiter='\t')
        f.next()
        total = 0.0
        for row in f:
            if row:
                total += 1.0

        Document.seek(0)
        f = csv.reader(Document, delimiter='\t')
        f.next()
        step = 0.0
        for row in f:
            if row:
                step += 1.0
                perc = int(step / total * 100)
                subbed = re.sub(r'(\(.*?\)|k__|p__|c__|o__|f__|g__|s__|otu__)', '', row[2])
                subbed = subbed[:-1]
                taxon = subbed.split(';')

                if not Kingdom.objects.filter(kingdomName=taxon[0]).exists():
                    kid = uuid4().hex
                    Kingdom.objects.create(kingdomid=kid, kingdomName=taxon[0])

                k = Kingdom.objects.get(kingdomName=taxon[0]).kingdomid
                if not Phyla.objects.filter(kingdomid_id=k, phylaName=taxon[1]).exists():
                    pid = uuid4().hex
                    Phyla.objects.create(kingdomid_id=k, phylaid=pid, phylaName=taxon[1])

                p = Phyla.objects.get(kingdomid_id=k, phylaName=taxon[1]).phylaid
                if not Class.objects.filter(kingdomid_id=k, phylaid_id=p, className=taxon[2]).exists():
                    cid = uuid4().hex
                    Class.objects.create(kingdomid_id=k, phylaid_id=p, classid=cid, className=taxon[2])

                c = Class.objects.get(kingdomid_id=k, phylaid_id=p, className=taxon[2]).classid
                if not Order.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderName=taxon[3]).exists():
                    oid = uuid4().hex
                    Order.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid=oid, orderName=taxon[3])

                o = Order.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderName=taxon[3]).orderid
                if not Family.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyName=taxon[4]).exists():
                    fid = uuid4().hex
                    Family.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid=fid, familyName=taxon[4])

                f = Family.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyName=taxon[4]).familyid
                if not Genus.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusName=taxon[5]).exists():
                    gid = uuid4().hex
                    Genus.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid=gid, genusName=taxon[5])

                try:
                    g = Genus.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusName=taxon[5]).genusid
                    if not Species.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesName=taxon[6]).exists():
                        sid = uuid4().hex
                        Species.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesid=sid, speciesName=taxon[6])
                except Exception:
                    g = Genus.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusName=taxon[5]).genusid
                    if not Species.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesName='unclassified').exists():
                        sid = uuid4().hex
                        Species.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesid=sid, speciesName='unclassified')

                try:
                    s = Species.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesName=taxon[6]).speciesid
                    if not OTU_99.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesid_id=s, otuName=taxon[7]).exists():
                        oid = uuid4().hex
                        OTU_99.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesid_id=s, otuid=oid, otuName=taxon[7])
                except Exception:
                    s = Species.objects.get(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesName=taxon[6]).speciesid
                    if not OTU_99.objects.filter(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesid_id=s, otuName='unclassified').exists():
                        oid = uuid4().hex
                        OTU_99.objects.create(kingdomid_id=k, phylaid_id=p, classid_id=c, orderid_id=o, familyid_id=f, genusid_id=g, speciesid_id=s, otuid=oid, otuName='unclassified')

    except Exception:
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)


def parse_profile(file3, file4, p_uuid, refDict):
    try:
        global stage, perc
        stage = "Step 5 of 5: Parsing shared file..."
        perc = 0
        print "Parsing profile!"
        data1 = genfromtxt(file3, delimiter='\t', dtype=None, autostrip=True)
        arr1 = np.delete(data1, 1, axis=1)
        df1 = pd.DataFrame(arr1[1:, 1:], index=arr1[1:, 0], columns=arr1[0, 1:])
        df1 = df1[df1.index != 'False']
        a = df1.columns.values.tolist()
        a = [x for x in a if x != 'False']
        df1 = df1[a]
        file3.close()

        data2 = genfromtxt(file4, delimiter='\t', dtype=None, autostrip=True)
        arr2 = data2.T
        arr2 = np.delete(arr2, 0, axis=0)
        arr2 = np.delete(arr2, 1, axis=0)
        df2 = pd.DataFrame(arr2[1:, 1:], index=arr2[1:, 0], columns=arr2[0, 1:])
        df2 = df2[df2.index != 'False']
        b = df2.columns.values.tolist()
        b = [x for x in b if x != 'False']
        df2 = df2[b]
        file4.close()

        df3 = pd.merge(df1, df2, left_index=True, right_index=True, how='inner')
        df3['Taxonomy'].replace(to_replace='(\(.*?\)|k__|p__|c__|o__|f__|g__|s__|otu__)', value='', regex=True, inplace=True)
        df3.reset_index(drop=True, inplace=True)
        del df1, df2

        total, columns = df3.shape
        sampleList = df3.columns.values.tolist()
        sampleList.remove('Taxonomy')

        step = 0.0
        for index, row in df3.iterrows():
            perc = int(step / total * 100)
            taxon = str(row['Taxonomy'])
            taxon = taxon[:-1]
            taxaList = taxon.split(';')

            k = taxaList[0]
            p = taxaList[1]
            c = taxaList[2]
            o = taxaList[3]
            f = taxaList[4]
            g = taxaList[5]
            try:
                s = taxaList[6]
            except Exception:
                s = 'unclassified'
            try:
                otu = taxaList[7]
            except Exception:
                otu = 'unclassified'

            t_kingdom = Kingdom.objects.get(kingdomName=k)
            t_phyla = Phyla.objects.get(kingdomid=t_kingdom, phylaName=p)
            t_class = Class.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, className=c)
            t_order = Order.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderName=o)
            t_family = Family.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order, familyName=f)
            t_genus = Genus.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order, familyid=t_family, genusName=g)
            t_species = Species.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order, familyid=t_family, genusid=t_genus, speciesName=s)
            t_otu = OTU_99.objects.get(kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order, familyid=t_family, genusid=t_genus, speciesid=t_species, otuName=otu)

            for name in sampleList:
                count = int(row[name])
                if count > 0:
                    project = Project.objects.get(projectid=p_uuid)
                    sample = Sample.objects.filter(projectid=p_uuid).get(sample_name=name)
                    sampid = sample.sampleid
                    refid = refDict[sampid]
                    reference = Reference.objects.get(refid=refid)

                    if Profile.objects.filter(projectid=project, refid=reference, sampleid=sample, kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order, familyid=t_family, genusid=t_genus, speciesid=t_species, otuid=t_otu).exists():
                        t = Profile.objects.get(projectid=project, refid=reference, sampleid=sample, kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order, familyid=t_family, genusid=t_genus, speciesid=t_species, otuid=t_otu)
                        old = t.count
                        new = old + int(count)
                        t.count = new
                        t.save()
                    else:
                        Profile.objects.create(projectid=project, refid=reference, sampleid=sample, kingdomid=t_kingdom, phylaid=t_phyla, classid=t_class, orderid=t_order, familyid=t_family, genusid=t_genus, speciesid=t_species, otuid=t_otu, count=count)

            step += 1.0

        for ID in sampleList:
            sample = Sample.objects.get(sampleid=ID)
            reads = Profile.objects.filter(sampleid=ID).aggregate(count=Sum('count'))
            sample.reads = reads['count']
            sample.save()

    except Exception:
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)


def repStop(request):
    # cleanup in repro project!
    print "Stopping!"
    return "Stopped"

@transaction.atomic
def reanalyze(request, stopList):
    ### create savepoint
    sid = transaction.savepoint()

    try:
        global rep_project

        mothurdest = 'mothur/temp/'
        if not os.path.exists(mothurdest):
            os.makedirs(mothurdest)
        ids = request.POST["ids"]
        processors = request.POST["processors"]
        RID = request.POST['RID']
        PID = 0  # change if adding additional data threads

        if stopList[PID] == RID:
            return repStop(request)

        new_align = 'reference=mothur/reference/align/' + str(request.POST['alignDB'])
        new_taxonomy = 'taxonomy=mothur/reference/taxonomy/' + str(request.POST['taxonomyDB'])
        new_tax = str(request.POST['taxonomyDB'])
        new_tax_tag = str(new_tax.split('.')[-2:-1][0])
        new_template = 'template=mothur/reference/template/' + str(request.POST['templateDB'])

        if isinstance(ids, list):
            projects = Reference.objects.all().filter(refid__in=ids)
        else:
            projects = Reference.objects.all().filter(refid=ids)

        if stopList[PID] == RID:
            return repStop(request)

        for project in projects:
            rep_project = 'myPhyloDB is currently reprocessing project: ' + str(project.projectid.project_name)
            dest = project.path
            source = project.source

            if stopList[PID] == RID:
                return repStop(request)

            if source == '454_sff':
                ls_dir = os.listdir(dest)
                for afile in ls_dir:
                    if stopList[PID] == RID:
                        return repStop(request)
                    srcStr = str(dest) + '/' + str(afile)
                    destStr = str(mothurdest) + '/' + str(afile)
                    shutil.copyfile(srcStr, destStr)

            if source == '454_fastq':
                file_list = []
                for afile in glob.glob(r'% s/*.fna' % dest):
                    if stopList[PID] == RID:
                        return repStop(request)
                    file_list.append(afile)

                tempList = []
                if file_list.__len__() > 1:
                    for each in file_list:
                        if stopList[PID] == RID:
                            return repStop(request)
                        file = each
                        handle_uploaded_file(file, mothurdest, each)
                        handle_uploaded_file(file, dest, each)
                        if os.name == 'nt':
                            myStr = "mothur\\temp\\" + str(file.name)
                        else:
                            myStr = "mothur/temp/" + str(file.name)
                        tempList.append(myStr)
                    inputList = "-".join(tempList)

                    if stopList[PID] == RID:
                        return repStop(request)

                    if os.name == 'nt':
                        os.system('"mothur\\mothur-win\\mothur.exe \"#merge.files(input=%s, output=mothur\\temp\\temp.fasta)\""' % inputList)
                    else:
                        os.system("mothur/mothur-linux/mothur \"#merge.files(input=%s, output=mothur/temp/temp.fasta)\"" % inputList)

                else:
                    for each in file_list:
                        if stopList[PID] == RID:
                            return repStop(request)

                        file = each
                        fasta = 'temp.fasta'
                        handle_uploaded_file(file, mothurdest, fasta)
                        handle_uploaded_file(file, dest, each)

                file_list = []
                for afile in glob.glob(r'% s/*.qual' % dest):
                    file_list.append(afile)

                tempList = []
                if file_list.__len__() > 1:
                    for each in file_list:
                        file = each
                        handle_uploaded_file(file, mothurdest, each)
                        handle_uploaded_file(file, dest, each)
                        if os.name == 'nt':
                            myStr = "mothur\\temp\\" + str(file.name)
                        else:
                            myStr = "mothur/temp/" + str(file.name)
                        tempList.append(myStr)
                    inputList = "-".join(tempList)
                    if os.name == 'nt':
                        os.system('"mothur\\mothur-win\\mothur.exe \"#merge.files(input=%s, output=mothur\\temp\\temp.qual)\""' % inputList)
                    else:
                        os.system("mothur/mothur-linux/mothur \"#merge.files(input=%s, output=mothur/temp/temp.qual)\"" % inputList)
                else:
                    for each in file_list:
                        file = each
                        qual = 'temp.qual'
                        handle_uploaded_file(file, mothurdest, qual)
                        handle_uploaded_file(file, dest, each)

                for afile in glob.glob(r'% s/*.oligos' % dest):
                    srcStr = str(dest) + '/' + str(afile)
                    destStr = str(mothurdest) + '/' + str(afile)
                    shutil.copyfile(srcStr, destStr)

                for afile in glob.glob(r'% s/*.batch' % dest):
                    srcStr = str(dest) + '/' + str(afile)
                    destStr = str(mothurdest) + '/' + str(afile)
                    shutil.copyfile(srcStr, destStr)

            if source == 'miseq':
                ls_dir = os.listdir(dest)
                for afile in ls_dir:
                    srcStr = str(dest) + '/' + str(afile)
                    destStr = str(mothurdest) + '/' + str(afile)
                    shutil.copyfile(srcStr, destStr)

            if stopList[PID] == RID:
                return repStop(request)

            orig_align = 'reference=mothur/reference/align/' + str(project.alignDB)
            orig_taxonomy = 'taxonomy=mothur/reference/taxonomy/' + str(project.taxonomyDB)
            orig_tax = str(project.taxonomyDB)
            orig_tax_tag = str(orig_tax.split('.')[-2:-1][0])
            orig_template = 'template=mothur/reference/template/' + str(project.templateDB)

            if stopList[PID] == RID:
                return repStop(request)

            try:
                with open("% s/mothur.batch" % dest, 'r+') as bat:
                    with open("% s/mothur.batch" % mothurdest, 'wb+') as destination:
                        method = 'wang'
                        foundClassify = False
                        orig_tag_meth = ''
                        new_tag_meth = ''
                        for line in bat:
                            if "processors" in line:
                                line = line.replace('processors=X', 'processors='+str(processors))
                            if "align.seqs" in line:
                                line = line.replace(str(orig_align), str(new_align))
                            if "classify.seqs" in line:
                                line = line.replace(str(orig_template), str(new_template))
                                line = line.replace(str(orig_taxonomy), str(new_taxonomy))
                                cmds = line.split(',')
                                for item in cmds:
                                    if "method=" in item:
                                        method = item.split('=')[1]
                                orig_tag_meth = str(orig_tax_tag) + "." + str(method)
                                new_tag_meth = str(new_tax_tag) + "." + str(method)
                                foundClassify = True
                            if (str(orig_tag_meth) in line) and foundClassify:
                                line = line.replace(str(orig_tag_meth), str(new_tag_meth))
                            destination.write(line)
            except Exception as e:
                print("Error with batch file: ", e)

            shutil.copy('% s/final_meta.xlsx' % dest, '% s/final_meta.xlsx' % mothurdest)

            if stopList[PID] == RID:
                return repStop(request)

            if not os.path.exists(dest):
                os.makedirs(dest)

            shutil.copy('% s/final_meta.xlsx' % mothurdest, '% s/final_meta.xlsx' % dest)
            if stopList[PID] == RID:
                transaction.savepoint_rollback(sid)
                return repStop(request)

            try:
                mothur(dest, source)
            except Exception as e:
                transaction.savepoint_rollback(sid)
                print("Error with mothur: " + str(e))
                return HttpResponse(
                    json.dumps({"error": "yes"}),
                    content_type="application/json"
                )

            if stopList[PID] == RID:
                transaction.savepoint_rollback(sid)
                return repStop(request)

            # subQueue()
            metaName = 'final_meta.xlsx'
            metaFile = '/'.join([dest, metaName])

            with open(metaFile, 'rb') as f:
                p_uuid, pType, num_samp = projectid(f)

            curUser = User.objects.get(username=request.user.username)
            parse_project(metaFile, p_uuid, curUser)

            if stopList[PID] == RID:
                transaction.savepoint_rollback(sid)
                return repStop(request)

            with open('% s/mothur.batch' % dest, 'rb') as file7:
                raw = True
                userID = str(request.user.id)

                refDict = parse_sample(metaFile, p_uuid, pType, num_samp, dest, file7, raw, source, userID)

            with open('% s/final.taxonomy' % dest, 'rb') as file3:
                parse_taxonomy(file3)

            with open('% s/final.taxonomy' % dest, 'rb') as file3:
                with open('% s/final.shared' % dest, 'rb') as file4:
                    parse_profile(file3, file4, p_uuid, refDict)

        return None

    except Exception as e:
        print " e = ", e
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        transaction.savepoint_rollback(sid)
        logging.exception(myDate)
        return repStop(request)
