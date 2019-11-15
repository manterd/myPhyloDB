import datetime
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.hashers import check_password
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.signals import user_logged_in
from django.db import transaction
from django.db.models import Q, Value
from django.db.models.functions import Concat
from django.http import *
from django.shortcuts import render
import fileinput
import json
import logging
import multiprocessing as mp
import openpyxl
import os
import pandas as pd
import pickle
import shutil
import subprocess
import tarfile
import time
import ujson
from uuid import uuid4
import zipfile
import fnmatch

from forms import UploadForm1, UploadForm2, UploadForm4, UploadForm5, \
    UploadForm6, UploadForm7, UploadForm8, UploadForm9, UploadForm10, UserUpdateForm, BulkForm

from database.models import Project, Reference, Sample, Air, Human_Associated, Microbial, Soil, Water, UserDefined, \
    OTU_99, PICRUSt, UserProfile, \
    ko_lvl1, ko_lvl2, ko_lvl3, ko_entry, \
    nz_lvl1, nz_lvl2, nz_lvl3, nz_lvl4, nz_entry, \
    addQueue, getQueue, subQueue, koOtuList

import functions
from functions.utils.debug import debug

from database import perms

from django.contrib.admin.models import LogEntry

maxSamples = 500  # change this value for maximum sample count changes (more RAM available or somesuch)
rep_project = ''
pd.set_option('display.max_colwidth', -1)
LOG_FILENAME = 'error_log.txt'
pro = None

# when adding new file types, just add name used in html (and for directory) to subDirList
subDirList = 'meta', 'shared', 'taxa', 'sequence', 'script', \
                         'sff', 'oligos', 'files', 'fna', 'qual', 'contig', 'fastq'


def home(request, errorText=""):
    if errorText == "":
        functions.log(request, "PAGE", "HOME")
    else:
        functions.log(request, "ERR", "HOME")
    return render(
        request,
        'home.html',
        {
            'error': errorText  # TODO 1.3 home page doesn't check error text, add to base?
        }
    )


@login_required(login_url='/myPhyloDB/accounts/login/')
def admin_console(request):  # console is a page with ongoing data, atm cannot think of a use for onload error alerts
    functions.log(request, "PAGE", "CONSOLE")
    if request.user.is_superuser:
        return render(
            request,
            'console.html'
        )
    else:
        # this is an odd case of someone trying to get to a page they shouldn't know about and/or be able to see
        # flagging it in console log for crime and punishment reasons (can we ban people? ips?)
        functions.log(request, "ILLEGAL", "ILL_CONSOLE")
        # this error message is not displayed to user, they are simply sent the home page instead
        return home(request, errorText="Invalid permissions")


@login_required(login_url='/myPhyloDB/accounts/login/')
def history(request, errorText=""):   # TODO 1.4 support error case: RID not find? request out of date, etc
    # page for viewing previous analyses (can add ongoing and shared results next)
    # the page itself makes all needed calls, so just render the template and go
    functions.log(request, "PAGE", "HISTORY")
    return render(
        request,
        'history.html',
        {
            'error': errorText
        }
    )


@login_required(login_url='/myPhyloDB/accounts/login/')
def files(request, errorText=""):
    if errorText == "":
        functions.log(request, "PAGE", "FILES")
    else:
        functions.log(request, "ERR", "FILES")
    # script files are unusable unless uploaded by admin (literal account name, not role) for security purposes
    # scriptVis flag doesn't prevent user from uploading scripts, just makes it harder
    # supers can see it, but only the ones uploaded by admin can be used
    scriptVis = "false"
    if request.user.is_superuser:
        scriptVis = "true"

    return render(
        request,
        'files.html',
        {'BulkForm': BulkForm,
         'form2': UploadForm2,
         'scriptVis': scriptVis,
         'error': errorText}
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def process(request, errorText=""):
    if errorText == "":
        functions.log(request, "PAGE", "PROCESS")

    projects = Reference.objects.none()
    if request.user.is_superuser:
        projects = Reference.objects.all().order_by('projectid__project_name', 'path')
    elif request.user.is_authenticated():
        projects = Reference.objects.all().order_by('projectid__project_name', 'path').filter(author=request.user)

    return render(
        request,
        'process.html',
        {'projects': projects,
         'form1': UploadForm1,
         'form2': UploadForm2,
         'error': errorText}
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def download(request):
    functions.log(request, "PAGE", "DOWNLOAD")
    projects = Reference.objects.none()
    if request.user.is_superuser:
        projects = Reference.objects.all().order_by('projectid__project_name', 'path')
    elif request.user.is_authenticated():
        projects = Reference.objects.all().order_by('projectid__project_name', 'path').filter(author=request.user)

    return render(
        request,
        'download.html',
        {'projects': projects,
         'type': "GET",
         'paths': ''}
    )


def getProjectFiles(request):
    all = request.GET['all']
    selKeys = json.loads(all)

    zip_file = os.path.join(os.getcwd(), 'myPhyloDB', 'media', 'usr_temp', request.user.username, 'final_data.zip')
    zf = zipfile.ZipFile(zip_file, "w", zipfile.ZIP_DEFLATED)

    foundCount = 0
    errorMsg = "none"
    debug("GET PROJECT", selKeys)
    visibleProjects = perms.getViewProjects(request)
    visibleIDs = []
    # get list of project ID's we are okay with this user viewing / downloading from
    for proj in visibleProjects:
        visibleIDs.append(proj.projectid)
    for path in selKeys:
        debug("PATH:", path)

        # Three main security requirements for user-submitted paths: (can be safe if these are followed)
        # REQUIREMENT 1:
        #   path must include 'uploads' as first directory (absolute paths start with / so first would be blank there)
        #   this also makes certain our relative path starts in the right location for requirement 2
        # REQUIREMENT 2:
        #   path must include projectid as second directory, with projectid belonging to visibleProjects for this user
        # REQUIREMENT 3:
        #   path cannot include ".." anywhere, to prevent jumping back up the file tree

        if path.split("/")[0] != "uploads":
            print "Security error!", request.user.username, "attempted to download from a directory other than uploads!"
            functions.securityLog(request, "FILES", "download traversal attempt")
            zf.close()
            results = {'error': "Invalid path"}
            myJson = json.dumps(results, ensure_ascii=False)
            return HttpResponse(myJson)
        if ".." in path:
            print "Security error!", request.user.username, "attempted use '..' to escape pathing!"
            functions.securityLog(request, "FILES", "download escape attempt")
            zf.close()
            results = {'error': "Invalid path"}
            myJson = json.dumps(results, ensure_ascii=False)
            return HttpResponse(myJson)
        # verify this path is in the list of allowed paths
        if path.split("/")[1] in visibleIDs: #or path.split("\\")[1] in visibleIDs:
            # atm assuming linux style file paths (forward slashes)
            # now we verify the path exists (like the actual file, in theory this always works, unless the user throws .. into the path or something...
            if os.path.exists(path):
                foundCount += 1
                try:
                    debug("Found")
                    zf.write(path)
                except Exception:
                    debug("Exc")
                    pass
            else:
                debug("Not found")
        else:
            # they aren't supposed to be able to see this project, the tree shouldn't even show it (HAX!!!)
            print "Security error!", request.user.username, "attempted to download files without permission!"
            functions.securityLog(request, "FILES", "download permissions abuse attempt")
            zf.close()
            results = {'error': "Invalid permissions"}
            myJson = json.dumps(results, ensure_ascii=False)
            return HttpResponse(myJson)
    zf.close()

    if foundCount != len(selKeys):
        errorMsg = "Failed to locate project files"

    results = {'files': zip_file, 'error': errorMsg}
    myJson = json.dumps(results, ensure_ascii=False)
    return HttpResponse(myJson)


def upStop(request):  # upStop is not cleaning up uploaded files (directory, selection) # TODO 1.3 procStop refactor
    # this function exists for logging sake basically
    functions.log(request, "STOP", "PROCESS")
    process(request, errorText="Processing stopped")


# TODO 1.3 html check
def upErr(msg, request, dest, sid):  # TODO 1.3 refactor to procErr
    logException()
    debug("upErr: user", request.user.username, "msg", msg)
    functions.log(request, "ERROR", "UPLOAD")
    projects = Reference.objects.none()
    try:
        functions.remove_proj(dest, request.user)
        transaction.savepoint_rollback(sid)
    except Exception as e:
        print "Error during upload: ", e

    try:
        if request.user.is_superuser:
            projects = Reference.objects.all().order_by('projectid__project_name', 'path')
        elif request.user.is_authenticated():
            projects = Reference.objects.all().order_by('projectid__project_name', 'path').filter(author=request.user)
    except:
        pass
    # need to strip msg of newline and single quotes so javascript doesn't throw a fit (since django direct inserts)
    msg = msg.translate(None, "\n\'")

    return render(
        request,
        'process.html',
        {'projects': projects,
         'form1': UploadForm1,
         'form2': UploadForm2,
         'error': msg   # TODO 1.3 verify security of msg, can user custom write? can it inject?
         }
    )


# need to clean this up, make like 6 helper functions, I don't want to add another 1k line function to the codebase
# needs dynamic input support, ie make it as simple as possible to add another data type to the upload chain
def cleanInput(key, data, text, has, good):
    curText = str(data[key])
    found = False
    if curText != "None":
        curText = ""
        curSection = 1
        for section in data[key]:
            if curSection < len(data[key]):
                curText += str(section) + ","
            else:
                curText += str(section)
            curSection += 1
        found = True

    # verify no malicious characters are in input, verify input exists, get string form to hand back
    if curText.find("..") != -1:
        good = False

    text[key] = curText
    has[key] = found

    return text, has, good


def checkFilePresence(source, has, ref):
    missingFiles = False
    filesNeeded = ""

    if not has['meta']:  # we always need a meta file
        missingFiles = True
        filesNeeded += "Meta; "

    if source == 'mothur':
        if not has['shared']:  # we always need a meta file
            missingFiles = True
            filesNeeded += "Shared; "
        raw = False
        if ref == 'yes':
            if not has['sequence']:
                missingFiles = True
                filesNeeded += "Sequence; "
        else:
            if not has['taxa']:
                missingFiles = True
                filesNeeded += "Taxa; "
    elif source == '454_sff':
        raw = True

        if not has['sff']:
            missingFiles = True
            filesNeeded += "SFF; "
        if not has['oligos']:
            missingFiles = True
            filesNeeded += "Oligos; "
        if not has['files']:
            missingFiles = True
            filesNeeded += "Files; "
        if not has['script']:
            missingFiles = True
            filesNeeded += "Script; "

    elif source == '454_fastq':
        raw = True

        if not has['fna']:
            missingFiles = True
            filesNeeded += "FNA; "
        if not has['qual']:
            missingFiles = True
            filesNeeded += "Qual; "
        if not has['oligos']:
            missingFiles = True
            filesNeeded += "Oligos; "
        if not has['script']:
            missingFiles = True
            filesNeeded += "Script; "

    elif source == 'miseq':
        raw = True

        if not has['contig']:
            missingFiles = True
            filesNeeded += "Contig; "
        if not has['fastq']:
            missingFiles = True
            filesNeeded += "Fastq; "
        if not has['script']:
            missingFiles = True
            filesNeeded += "Script; "

    elif source == 'biom':
        raw = False
        # shortcutting via taxa file type, since biom contains taxa here anyways
        if not has['taxa']:
            missingFiles = True
            filesNeeded += "Taxa; "

    else:
        # they sent a value we weren't expecting, cancel!!!
        raw = False
        missingFiles = True
        filesNeeded += "Upload format not supported\t"
    return missingFiles, filesNeeded, raw


def copyFromUpload(source, dest, name):
    # use shutil.copy to get file from source directory and copy to dest
    # make sure source and dest are both valid locations (no '..' charsets, make needed directories)
    # as in, this DOES NOT clean inputs, that's what cleanInput function is for
    try:
        if not os.path.exists(dest):
            try:
                os.makedirs(dest)
            except Exception as errr:
                print "Error with making new", name, "directory:", errr
        trueDest = os.path.join(dest, name)
        # copy file to working project directory
        shutil.copy2(source, trueDest)
    except Exception as er:
        print "Error with", name, "shutil:", er


# TODO 1.3 html check
def cancelUploadEarly(request, projects, reason):
    functions.log(request, "UP_FAIL", reason)
    return render(
        request,
        'process.html',
        {'projects': projects,
         'form1': UploadForm1,
         'form2': UploadForm2,
         'error': reason}
    )


def checkCompsGetNames(compList, user, key):
    keyName = ""
    selcomps = compList.split('/')
    selpart = ""
    curComp = 0
    lastComp = len(selcomps) - 1
    if len(selcomps) > 3:
        print user.username, "attempted to send invalid node data"  # log this in admin?
        # this is legitimately the most sketchy thing someone can do on this site atm, why its being checked for
        return "Illegal tree size", "error"
    myProf = UserProfile.objects.get(user=user)
    myPerms = myProf.hasPermsFrom.split(';')
    for comp in selcomps:
        if curComp == 0:
            if key == 'script':
                if comp != 'admin' and comp != "None":  # TODO 1.4 this is where the script admin is listed, revise
                    return "Illegal file access", "error"
            else:
                # can bypass with superuser, or by whitelist from permissions/user_profile system
                if comp != user.username and comp != "None":
                    # print user.username, "attempted to access files outside of their permissions"
                    if not user.is_superuser and comp not in myPerms:
                        return "Illegal file access", "error"

        if curComp == lastComp:
            selpart += key + "/" + comp
            keyName = comp
        else:
            selpart += comp + "/"
        curComp += 1

    selectedPath = os.path.join("user_uploads", selpart)
    return selectedPath, keyName


def getSelectedPath(key, textDict, user):
    # this function takes a key (type of file), textDict (actual strings given by user)
    # also given user to check super status and username (could query for profile later)
    # take given file type and text input, generate actual path from user_uploads to file, as well as filenames
    # return list of paths and names

    # need to do this tree length and first level match check with ALL file types involved
    paths = []
    filenames = []

    for filepath in textDict[key].split(','):
        retPath, retName = checkCompsGetNames(filepath, user, key)
        if retName == "error":
            print "Error!", retPath, "during checkComps. case:", filepath
            return retPath, "error"
        paths.append(retPath)
        filenames.append(retName)
        # store this in a list
        # return the full list of selected files
    return paths, filenames


def logException():
    # get crash info and send it to error_log
    logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG, )
    myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
    logging.exception(myDate)


def handleMothurRefData(request, nameDict, selDict, dest):  # no samples with this method, need different fasta?
    # make taxa file for upload1 style using fasta data
    debug("Handling mothur reference data")
    mothurdest = 'mothur/temp'
    # wipe old temp mothur files, recreate directory for new process
    if os.path.exists(mothurdest):
        shutil.rmtree(mothurdest)

    # write dada file to mothur
    try:
        if not os.path.exists(mothurdest):
            try:
                os.makedirs(mothurdest)
            except Exception as errr:
                print "Error with making new sequence directory:", errr
                # should return and cancel upload if this exception occurs
                return errr
        trueDest = os.path.join(mothurdest, 'dada.fasta')
        # copy file to working project directory
        shutil.copy2(selDict['sequence'], trueDest)
    except Exception as er:
        return "-reference handling: " + str(er)

    copyFromUpload(selDict['meta'], dest, nameDict['meta'])
    cmd = ''
    if os.name == 'nt':
        if request.POST['ref_var'] == 'V4':
            cmd = "mothur\\mothur-win\\vsearch -usearch_global mothur\\temp\\dada.fasta -db mothur\\reference\\dada2\\gg_13_5_99.V4.fa.gz --strand both --id 0.99 --fastapairs mothur\\temp\\pairs.fasta --notmatched mothur\\temp\\nomatch.fasta"
        if request.POST['ref_var'] == 'V34':
            cmd = "mothur\\mothur-win\\vsearch -usearch_global mothur\\temp\\dada.fasta -db mothur\\reference\\dada2\\gg_13_5_99.V3_V4.fa.gz --strand both --id 0.99 --fastapairs mothur\\temp\\pairs.fasta --notmatched mothur\\temp\\nomatch.fasta"
        if request.POST['ref_var'] == 'V13':
            cmd = "mothur\\mothur-win\\vsearch -usearch_global mothur\\temp\\dada.fasta -db mothur\\reference\\dada2\\gg_13_5_99.V1_V3.fa.gz --strand both --id 0.99 --fastapairs mothur\\temp\\pairs.fasta --notmatched mothur\\temp\\nomatch.fasta"
        if request.POST['ref_var'] == 'V19':
            cmd = "mothur\\mothur-win\\vsearch -usearch_global mothur\\temp\\dada.fasta -db mothur\\reference\\dada2\\gg_13_5_99.V1_V9.fa.gz --strand both --id 0.99 --fastapairs mothur\\temp\\pairs.fasta --notmatched mothur\\temp\\nomatch.fasta"
    else:
        if request.POST['ref_var'] == 'V4':
            cmd = "mothur/mothur-linux/vsearch -usearch_global mothur/temp/dada.fasta -db mothur/reference/dada2/gg_13_5_99.V4.fa.gz --strand both --id 0.99 --fastapairs mothur/temp/pairs.fasta --notmatched mothur/temp/nomatch.fasta"
        if request.POST['ref_var'] == 'V34':
            cmd = "mothur/mothur-linux/vsearch -usearch_global mothur/temp/dada.fasta -db mothur/reference/dada2/gg_13_5_99.V3_V4.fa.gz --strand both --id 0.99 --fastapairs mothur/temp/pairs.fasta --notmatched mothur/temp/nomatch.fasta"
        if request.POST['ref_var'] == 'V13':
            cmd = "mothur/mothur-linux/vsearch -usearch_global mothur/temp/dada.fasta -db mothur/reference/dada2/gg_13_5_99.V1_V3.fa.gz --strand both --id 0.99 --fastapairs mothur/temp/pairs.fasta --notmatched mothur/temp/nomatch.fasta"
        if request.POST['ref_var'] == 'V19':
            cmd = "mothur/mothur-linux/vsearch -usearch_global mothur/temp/dada.fasta -db mothur/reference/dada2/gg_13_5_99.V1_V9.fa.gz --strand both --id 0.99 --fastapairs mothur/temp/pairs.fasta --notmatched mothur/temp/nomatch.fasta"

    pro = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                           bufsize=0)   # this should run mothur, does subprocess ever get started though?
    debug("Made subprocess, reading lines")
    while True:
        line = pro.stdout.readline()    # does this work?
        print "Line:", line
        if line != "":  # changed '' to ""
            print line
            # have not found an example of this working fully, probably because of this step being blank each time I try
        else:
            break
    debug("Read lines")

    f = open('mothur/temp/dada.cons.taxonomy', 'w')
    f.write("OTU\tSeq\tTaxonomy\n")
    pairDict = {}

    inFile1 = open('mothur/temp/pairs.fasta', 'r')
    counter = 1
    key = ''
    for line in inFile1:
        if counter == 1:
            key = line.rstrip().replace('>', '')
        if counter == 3:
            value = line.rstrip().replace('>', '')
            pairDict[key] = value
        if counter == 5:
            counter = 0
        counter += 1
    inFile1.close()

    inFile2 = open('mothur/temp/dada.fasta', 'r')
    counter = 1
    for line in inFile2:
        if counter == 1:
            key = line.rstrip().replace('>', '')
        if counter == 2:
            seq = line.rstrip()
            if key in pairDict:
                f.write(key)
                f.write('\t')
                f.write(seq)
                f.write('\t')
                f.write(pairDict[key])
                f.write('\n')
            else:
                f.write(key)
                f.write('\t')
                f.write(seq)
                f.write('\t')
                f.write(
                    'k__unclassified;p__unclassified;c__unclassified;o__unclassified;f__unclassified;g__unclassified;s__unclassified;otu__unclassified;')
                f.write('\n')
            counter = 0
        counter += 1

    inFile2.close()
    f.close()

    shutil.copy('mothur/temp/pairs.fasta', '% s/dada.vsearch_pairs.txt' % dest)
    shutil.copy('mothur/temp/dada.fasta', '% s/dada.rep_seqs.fasta' % dest)
    shutil.copy('mothur/temp/dada.cons.taxonomy', '% s/final.cons.taxonomy' % dest)
    return open('% s/final.cons.taxonomy' % dest)


def handleArchive(dest, name, request):  # take archive 'name' at location 'dest', extract all contents, move contents to pwd
    # this function performed a security check by making sure all contents of the archive will end up where their name
    # implies, throwing a security error back up if something doesn't match (absolute paths in archive possibly)
    try:
        unArchived = False
        debug("Attempting to extract archive at:", dest, "named:", name)
        zip = zipfile.ZipFile(os.path.join(dest, name))
        for member in zip.infolist():
            abspath = os.path.abspath(os.path.join(dest, member.filename))
            if abspath.startswith(os.path.abspath(dest)):
                zip.extract(member, dest)
            else:
                print "Security error: absolute path does not match in ZIP"
                functions.securityLog(request, "HandleArchive", "absolute path mismatch (ZIP)")
        zip.close()
        unArchived = True
    except Exception as e:
        debug("handleArchive: exception during zip:", e)
        try:
            tar = tarfile.open(os.path.join(dest, name))
            for member in tar.getmembers():
                abspath = os.path.abspath(os.path.join(dest, member.name))
                if abspath.startswith(os.path.abspath(dest)):
                    tar.extract(member, dest)
                else:
                    print "Security error: absolute path does not match in TAR"
                    functions.securityLog(request, "HandleArchive", "absolute path mismatch (TAR)")
            tar.close()
            unArchived = True
        except Exception as ex:
            debug("handleArchive: exeption during tar:", ex)
            return False  # file is unlikely to be an archive, as unzip and untar both failed
        if unArchived:
            # walk through the newly unarchived files and make sure everything is on the same level
            try:
                bubbleFiles(dest)
            except Exception as exc:
                debug("Error with move: ", exc)
    return True  # return true if file was successfully extracted as an archive, false otherwise



def makeSamplesErrorText(missingMeta, missingSecondary, secondaryFileType):
    errText = ""
    if len(missingMeta) != 0 or len(missingSecondary) != 0:
        if len(missingMeta) > 0:
            errText += "Found samples in meta that were missing from "+str(secondaryFileType)+":"
            for err in missingMeta:
                errText += err + ", "
            errText += "\n"
        if len(missingSecondary) > 0:
            errText += "Found samples in "+str(secondaryFileType)+" that were missing from meta:"
            for err in missingSecondary:
                errText += err + ", "
            errText += "\n"
    return errText

# TODO 1.3 sid usage in uploads (for transaction rollbacks, variable goes unused)
def uploadWithMothur(request, nameDict, selDict, refDict, p_uuid, dest, stopList, PID, RID):

    try:
        copyFromUpload(selDict['shared'], dest, nameDict['shared'])
        file4 = open(os.path.join(dest, nameDict['shared']), 'r')
        # get sample names from shared file, then send to comparison function (with p_uuid)
        sampNames = []
        lineNum = 0
        for line in file4:
            parts = line.split("\t")
            if lineNum != 0 and len(parts) > 1:
                sampNames.append(parts[1])
            lineNum += 1
        missingSamples, missingShared = functions.validateSamples(p_uuid, sampNames)
        errText = makeSamplesErrorText(missingSamples, missingShared, "shared")
        if errText != "":
            return errText
    except Exception:
        return "Cannot open shared file"

    if stopList[PID] == RID:
        return "Stop"

    try:
        if request.POST['ref_data'] == 'yes':  # sequence plus R script
            file3 = handleMothurRefData(request, nameDict, selDict, dest)
            if type(file3) is str:
                return file3   # not getting samples from this method atm
        else:
            # directly added taxa file
            taxaname = nameDict['taxa']
            copyFromUpload(selDict['taxa'], dest, taxaname)
            file3 = open(os.path.join(dest, taxaname), 'r')
    except Exception as e:
        return "Problem with taxa file: " + str(e)

    try:
        functions.parse_taxonomy(file3, stopList, PID, RID)
        file3 = file3.name
    except Exception:
        return "Failed parsing your taxonomy file:" + nameDict['taxa']

    if stopList[PID] == RID:
        return "Stop"

    try:
        functions.parse_profile(file3, file4, p_uuid, refDict, stopList, PID, RID)  # taxafile vs file3
    except Exception:
        return "Failed parsing your shared file:" + str(file4.name)

    if stopList[PID] == RID:
        return "Stop"

    return "None"


def uploadWithSFF(request, nameDict, selDict, refDict, p_uuid, dest, stopList, PID, RID, sid, processors):
    mothurdest = 'mothur/temp'

    if os.path.exists(mothurdest):
        shutil.rmtree(mothurdest)

    if not os.path.exists(mothurdest):
        os.makedirs(mothurdest)

    if stopList[PID] == RID:
        return "Stop"

    # unzip/tar uploaded files. Needs to work recursively, pulling everything back to the top
    file_list = selDict['sff']
    name_list = nameDict['sff']
    # entries from these two dicts are lists, order SHOULD be synced via map (their lists were populated simultaneously)
    for filepath, name in map(None, file_list, name_list):
        copyFromUpload(filepath, mothurdest, name)
        handleArchive(mothurdest, name, request)

    if stopList[PID] == RID:
        return "Stop"

    file_list = selDict['oligos']
    name_list = nameDict['oligos']
    # entries from these two dicts are lists, order SHOULD be synced, need to verify (so far works)
    for filepath, name in map(None, file_list, name_list):
        copyFromUpload(filepath, mothurdest, name)
        handleArchive(mothurdest, name, request)

    if stopList[PID] == RID:
        return "Stop"

    copyFromUpload(selDict['files'], mothurdest, 'temp.txt')
    copyFromUpload(selDict['files'], dest, 'temp.txt')  # backup to upload folder for downloading later
    bubbleFiles(mothurdest)

    # function to check if mothur and metafiles samples match
    error = checkSamples(selDict['meta'], '454_sff', 'mothur/temp/temp.txt')

    if error != "":
        return "-checkSamples: " + str(error)

    batch = 'mothur.batch'
    batchPath = selDict['script']

    avail_proc = mp.cpu_count()
    use_proc = min(avail_proc, processors)
    actual_proc = 'processors=' + str(use_proc)

    copyFromUpload(batchPath, mothurdest, batch)

    for line in fileinput.input('mothur/temp/mothur.batch', inplace=1):
        line = line.rstrip('\r\n')
        line = line.replace("processors=X", actual_proc)
        print line

    copyFromUpload(batchPath, dest, batch)  # copy into project dest for archive/download purposes

    if stopList[PID] == RID:
        return "Stop"

    # check queue for mothur availability, then run or wait
    # if mothur is available, run it
    # else sleep, loop back to check again
    addQueue()
    while getQueue() > 1:
        time.sleep(5)
    try:
        functions.mothur(dest, "454_sff")
    except Exception as er:
        logException()
        return "mothur batch file:", str(nameDict['script']) + str(er)

    if stopList[PID] == RID:
        return "Stop"

    subQueue()

    try:
        with open('% s/final.cons.taxonomy' % dest, 'rb') as file3:
            functions.parse_taxonomy(file3, stopList, PID, RID)
    except Exception as er:
        logException()
        return "parsing your taxonomy file: final.cons.taxonomy" + str(er)

    if stopList[PID] == RID:
        return "Stop"

    try:
        with open('% s/final.cons.taxonomy' % dest, 'rb') as file3:
            with open('% s/final.tx.shared' % dest, 'rb') as file4:
                functions.parse_profile(file3, file4, p_uuid, refDict, stopList, PID, RID)

    except Exception as er:
        logException()
        return "parsing your shared file: final.tx.shared" + str(er)

    if stopList[PID] == RID:
        return "Stop"

    return "None"


def uploadWithFastq(request, nameDict, selDict, refDict, p_uuid, dest, stopList, PID, RID, sid, processors):
    mothurdest = 'mothur/temp'

    if os.path.exists(mothurdest):
        shutil.rmtree(mothurdest)

    if not os.path.exists(mothurdest):
        os.makedirs(mothurdest)

    file_list = selDict['fna']
    tempList = prepMultiFiles(file_list, dest, mothurdest, 'temp.fasta')
    inputList = "-".join(tempList)
    if os.name == 'nt':
        os.system(
            '"mothur\\mothur-win\\mothur.exe \"#merge.files(input=%s, output=mothur\\temp\\temp.fasta)\""' % inputList)
    else:
        os.system(
            "mothur/mothur-linux/mothur \"#merge.files(input=%s, output=mothur/temp/temp.fasta)\"" % inputList)

    if stopList[PID] == RID:
        return "Stop"

    file_list = selDict['qual']
    tempList = prepMultiFiles(file_list, dest, mothurdest, 'temp.qual')
    inputList = "-".join(tempList)
    if os.name == 'nt':
        os.system(
            '"mothur\\mothur-win\\mothur.exe \"#merge.files(input=%s, output=mothur\\temp\\temp.qual)\""' % inputList)
    else:
        os.system("mothur/mothur-linux/mothur \"#merge.files(input=%s, output=mothur/temp/temp.qual)\"" % inputList)

    if stopList[PID] == RID:
        return "Stop"

    oligo = 'temp.oligos'
    file6 = selDict['oligos'][0]    # remove list component when not a list? oligo vs oligos though
    copyFromUpload(file6, mothurdest, oligo)

    if stopList[PID] == RID:
        return "Stop"

    # TODO 1.4 split single and multi oligos formats (single doesn't work for multi and visa versa) (verify problem)
    # can't just go oligo and oligos, the 'in' checks get confused (could redo those to fix this?)
    # could try oligos and 0ligos or some mispelling, variant name, etc
    copyFromUpload(file6, dest, nameDict['oligos'][0])
    # function to check if mothur and metafiles samples match
    error = checkSamples(selDict['meta'], "454_fastq", 'mothur/temp/temp.oligos')

    if error != "":
        return "-sample check: " + str(error)

    if stopList[PID] == RID:
        return "Stop"

    batch = 'mothur.batch'
    file7 = selDict['script']

    avail_proc = mp.cpu_count()
    use_proc = min(avail_proc, processors)
    actual_proc = 'processors=' + str(use_proc)

    copyFromUpload(file7, mothurdest, batch)
    copyFromUpload(file7, dest, batch)

    if stopList[PID] == RID:
        return "Stop"

    for line in fileinput.input('mothur/temp/mothur.batch', inplace=True):
        line = line.rstrip('\r\n')
        line = line.replace("processors=X", actual_proc)
        print line

    if stopList[PID] == RID:
        return "Stop"

    try:
        functions.mothur(dest, "454_fastq")

    except Exception as er:
        logException()
        return "Encountered problem while running mothur:" + str(er)

    if stopList[PID] == RID:
        return "Stop"

    try:
        # final.fasta isn't being made before this step?
        with open('% s/final.cons.taxonomy' % dest, 'rb') as file3:  # crashing on open, why isn't the file here?
            functions.parse_taxonomy(file3, stopList, PID, RID)
    except Exception as er:
        logException()
        return "Encountered problem while parsing taxonomy:" + str(er)

    if stopList[PID] == RID:
        return "Stop"

    try:
        with open('% s/final.cons.taxonomy' % dest, 'rb') as file3:
            with open('% s/final.tx.shared' % dest, 'rb') as file4:
                functions.parse_profile(file3, file4, p_uuid, refDict, stopList, PID, RID)
    except Exception as er:
        logException()
        return "Encountered problem while parsing shared:" + str(er)

    if stopList[PID] == RID:
        return "Stop"

    return "None"


def uploadWithMiseq(request, nameDict, selDict, refDict, p_uuid, dest, stopList, PID, RID, sid, processors, platform):
    mothurdest = 'mothur/temp'

    if os.path.exists(mothurdest):
        shutil.rmtree(mothurdest)

    if not os.path.exists(mothurdest):
        os.makedirs(mothurdest)

    fastq = 'temp.files'
    file13 = selDict['contig']   # contig
    copyFromUpload(file13, mothurdest, fastq)

    # function to check if mothur and metafiles samples match
    error = checkSamples(selDict['meta'], "miseq", 'mothur/temp/temp.files')

    if error != "":
        return "-sample check: " + str(error)

    if stopList[PID] == RID:
        return "Stop"

    # functions.handle_uploaded_file(file13, dest, fastq)
    copyFromUpload(file13, dest, fastq)

    if stopList[PID] == RID:
        return "Stop"

    # file_list = request.FILES.getlist('fastq_files')  # fastq unzipper, make recursive pool (do NOT unzip more)
    file_list = selDict['fastq']
    name_list = nameDict['fastq']

    for file, name in map(None, file_list, name_list):  # TODO 1.3 this step needs a progress bar and a performance pass
        copyFromUpload(file, dest, name)    # copy for working with the files, then move to mothurdest (why)
        handleArchive(dest, name, request)
    for maindir, subdirs, filenames in os.walk(dest):
        for filename in filenames:
            # print "Moving ", filename
            try:
                if filename == "final_meta.xlsx":
                    debug("Copying meta file")
                    shutil.copy(os.path.join(maindir, filename), '% s/final_meta.xlsx' % mothurdest)    # TODO 1.3 verify
                else:
                    shutil.move(os.path.join(maindir, filename), mothurdest)
            except Exception as exc:
                debug("Error with move: ", exc)
        # need to move to mothurdest AFTER all handleArchive calls
        # this prevents missing files from archives in the map
        #copyFromUpload(file, mothurdest, name)

        if stopList[PID] == RID:
            return "Stop"
    # TODO 1.3 trying to track down the invalid literal error page, should be from in here somewhere, how to recreate...

    bubbleFiles(mothurdest)

    if platform == 'mothur':
        batch = 'mothur.batch'
        # file7 = request.FILES['docfile7']
        file7 = selDict['script']

        avail_proc = mp.cpu_count()
        use_proc = min(avail_proc, processors)
        actual_proc = 'processors=' + str(use_proc)

        copyFromUpload(file7, mothurdest, batch)
        copyFromUpload(file7, dest, batch)

        if stopList[PID] == RID:
            return "Stop"

        for line in fileinput.input('mothur/temp/mothur.batch', inplace=1):
            line = line.rstrip('\r\n')
            line = line.replace("processors=X", actual_proc)
            print line

        if stopList[PID] == RID:
            return "Stop"

        try:
            functions.mothur(dest, "miseq")
        except Exception as er:
            logException()
            return "Encountered problem while running mothur:" + str(er)

        if stopList[PID] == RID:
            return "Stop"

    elif platform == 'dada2':
        batch = 'dada2.R'
        # file7 = request.FILES['docfile7']
        file7 = selDict['script']    # repeated from above, move up in scope

        avail_proc = mp.cpu_count()
        use_proc = min(avail_proc, processors)
        actual_proc = 'multithread=' + str(use_proc)

        # functions.handle_uploaded_file(file7, mothurdest, batch)
        copyFromUpload(file7, mothurdest, batch)    # also repeated, batch is the only difference

        if stopList[PID] == RID:
            return "Stop"

        for line in fileinput.input('mothur/temp/dada2.R', inplace=1):
            line = line.rstrip('\r\n')
            line.replace("multithread=TRUE", actual_proc)
            print line

        # functions.handle_uploaded_file(file7, dest, batch)
        copyFromUpload(file7, dest, batch)

        if stopList[PID] == RID:
            return "Stop"

        try:
            functions.dada2(dest, "miseq")
        except Exception as er:
            logException()
            return "Encountered problem while processing data2:" + str(er)    # probably not an accurate error message

        if stopList[PID] == RID:
            return "Stop"

    try:
        with open('% s/final.cons.taxonomy' % dest, 'rb') as file3:
            functions.parse_taxonomy(file3, stopList, PID, RID)
    except Exception as er:
        logException()
        return "Encountered problem while parsing taxonomy:" + str(er)

    if stopList[PID] == RID:
        return "Stop"

    try:
        with open('% s/final.cons.taxonomy' % dest, 'rb') as file3:
            with open('% s/final.tx.shared' % dest, 'rb') as file4:
                functions.parse_profile(file3, file4, p_uuid, refDict, stopList, PID, RID)
    except Exception as er:
        logException()
        return "Encountered problem while parsing profile:" + str(er)

    if stopList[PID] == RID:
        return "Stop"
    return "None"


def uploadWithBiom(request, nameDict, selDict, refDict, p_uuid, dest, minConfidence, stopList, PID, RID, sid, processors):
    debug("UPLOAD WITH BIOM")
    try:
        # directly added taxa file (in biom format)
        # for .tsv with taxonomies, and sequence for .fasta with sequences
        biomName = nameDict['shared']
        taxaName = nameDict['taxa']
        sequenceName = nameDict['sequence']

        copyFromUpload(selDict['shared'], dest, biomName)
        copyFromUpload(selDict['taxa'], dest, taxaName)
        copyFromUpload(selDict['sequence'], dest, sequenceName)
        biomFile = os.path.join(dest, biomName)
        taxaFile = os.path.join(dest, taxaName)
        sequenceFile = os.path.join(dest, sequenceName)
    except Exception as e:
        return "Problem with taxa file: " + str(e)

    try:
        errSamp, errBiom = functions.parse_biom(biomFile, taxaFile, sequenceFile, p_uuid, refDict, minConfidence, stopList, PID, RID)
        # check errSamp and errBiom for names of samples which were in one file but not the other
        errText = ""
        if len(errSamp) > 0:
            # we had samples in meta but not in biom
            errText += "Found samples in meta that were missing from biom:"
            for err in errSamp:
                errText += err + ", "
            errText += "\n"
        if len(errBiom) > 0:
            # we had biom samples not in meta
            errText += "Found samples in biom that were missing from meta:"
            for err in errBiom:
                errText += err + ", "
            errText += "\n"
        if errText != "":
            return errText
    except Exception as exc:
        return "Error while parsing your biom files:" + str(exc)

    if stopList[PID] == RID:
        return "Stop"

    return "None"


# TODO 1.3 html check
def processFunc(request, stopList):

    # validation process involves checking settings paired with actual files sent

    # TODO 1.3 make sure all project files wind up in the p_uuid folder, including raw data files (oof thats a lot of space)
    # TODO 1.3 Also need to automatically cleanup raw files (in user_uploads) after a certain amount of time has passed (6 months for now)
    # don't cleanup the admin's scripts, maybe need a timer for projects as well? there's just a whole lot of files here
    # can have auto cleanup script go by folder name (since most are dates)
    # only need the script to run once a day or maybe even once a week

    projects = Reference.objects.none()

    start = datetime.datetime.now()

    RID = ''
    try:
        RID = request.POST['RID']
    except Exception as ers:
        print "Error with RID:", ers
        pass

    userID = str(request.user.id)
    PID = 0  # change if adding additional data threads

    if stopList[PID] == RID:
        return upStop(request)

    source = ''
    try:
        source = str(request.POST['source'])
    except Exception as ers:
        print "Error with source:", ers
        pass

    platform = ''
    try:
        platform = str(request.POST['platform'])
    except Exception as ers:
        print "Error with platform:", ers
        pass

    processors = 2
    try:
        processors = int(request.POST['processors'])
    except Exception as ers:
        print "Error with processors:", ers
        pass


    good = True   # verification flag, if anything goes wrong in security checks or we're missing data, make this False

    # do this process of checking for files via helper function in loop?
    # cleanInput function?
    # dict for pre's, dict for has
    # ie preDict['meta'] = "None" while hasDict['meta'] = False
    # vs preDict['meta'] = "Example1.xlsx" while hasDict['meta'] = True

    textDict = {}
    hasDict = {}

    # sff, oligos, and a couple others need multi select support

    allTheThings = request.POST
    debug("processFunc: allTheThings", allTheThings.keys())
    curData = json.loads(allTheThings['data'])

    for subDir in subDirList:
        textDict, hasDict, good = cleanInput(subDir, curData, textDict, hasDict, good)

    if not good:
        print request.user.username, "attempted to use illegal characters"
        return cancelUploadEarly(request, projects, "Illegal characters in file path")  # .. in filename or hacking
        # should probably make a separate potential hackers log
        # could inform user which file and which characters, in case they just misnamed a file

    # should probably include extra check steps on script if present
    # point being, only scripts the host of a particular server have approved or personally uploaded will be usable
    # thus we can skip screening scripts as only the highest permissions of users will be able to add them
    # can R or mothur be used for traversal? or other attacks? probably, need safeguards for these incoming scripts

    # verify all needed files are present based on source value
    # turn source value checking into helper function, given source and hasDict, return false if missing anything
    missingFiles, filesNeeded, raw = checkFilePresence(source, hasDict, request.POST['ref_data'])

    if missingFiles:
        return cancelUploadEarly(request, projects, "Missing files: "+filesNeeded)

    # need helper function to cancel upload (ie return with error message), log failed upload attempt (w/reason)
    # Search through each incoming node, filter and flag all .. occurrences, drop the request if found
    # otherwise just double check each section has permissions

    selDict = {}
    nameDict = {}
    for subDir in subDirList:
        selPaths, selNames = getSelectedPath(subDir, textDict, request.user)
        debug("For subDir:", subDir, "found paths", selPaths, "and names", selNames)
        if selNames == "error":
            return cancelUploadEarly(request, projects, selPaths)
        if subDir.lower() in "meta sequence taxa shared files script contig":
            # insert other always singular file types here  OLIGOS IS WEIRD OKAY?
            selPaths = selPaths[0]
            selNames = selNames[0]
        selDict[subDir] = selPaths
        nameDict[subDir] = selNames

    try:
        p_uuid, pType, num_samp = functions.projectid(selDict['meta'])  # crashes here if actual parse fails

        if not num_samp:
            reallyLongError = "Error: Please open and save your meta file in Excel/OpenOffice" \
                              " in order to perform the necessary calculations..."
            return cancelUploadEarly(request, projects, reallyLongError)    # last pre upload error
            # past this point, use upErr because we're making a project

    except Exception:
        return upErr("There was an error parsing your meta file:" + str(selDict['meta']), request, None, None)

    sid = transaction.savepoint()
    # got puuid by here, next?
    # actually run project parsing, file locating, etc
    date = datetime.date.today().isoformat()
    hour = datetime.datetime.now().hour
    minute = datetime.datetime.now().minute
    second = datetime.datetime.now().second
    timestamp = ".".join([str(hour), str(minute), str(second)])
    datetimestamp = "_".join([str(date), str(timestamp)])
    dest = "/".join(["uploads", str(p_uuid), str(datetimestamp)])
    metaName = 'final_meta.xlsx'
    metaFile = '/'.join([dest, metaName])   # metafile vs metaFile, should refactor
    curUser = User.objects.get(username=request.user.username)

    try:
        copyFromUpload(selDict['meta'], dest, metaName)
        res = functions.parse_project(metaFile, p_uuid, curUser)
        if res != "none":
            print "Encountered error while parsing meta file: " + res
            return upErr("Encountered error while parsing meta file: " + res, request, dest, sid)
    except Exception:
        return upErr("There was an error parsing your meta file:" + str(selDict['meta']), request, dest, sid)
    # end of meta parsing

    if stopList[PID] == RID:
        functions.remove_proj(dest, request.user)
        transaction.savepoint_rollback(sid)
        return upStop(request)

    # get sample info from selected meta file
    try:
        refDict = functions.parse_sample(metaFile, p_uuid, pType, num_samp, dest, raw, source, userID, stopList, RID, PID)
        debug("processFunc: refDict")#, refDict)
        # stop will force an early return, full stop will occur when detected again right after here
    except Exception as e:
        print "Error parsing sample: ", e
        logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,)
        myDate = "\nDate: " + str(datetime.datetime.now()) + "\n"
        logging.exception(myDate)
        return upErr("There was an error parsing your meta file:" + str(e), request, dest, sid)

    if stopList[PID] == RID:
        functions.remove_proj(dest, request.user)
        transaction.savepoint_rollback(sid)
        return upStop(request)

    # if using shared+taxa or shared+sequence mothur combo
    debug("Upload finished shared section")
    errorText = "Invalid source type for processing"
    if source == 'mothur':
        errorText = uploadWithMothur(request, nameDict, selDict, refDict, p_uuid, dest, stopList, PID, RID)  # TODO 1.3 sid on mothur upload

    if source == '454_sff':
        errorText = uploadWithSFF(request, nameDict, selDict, refDict, p_uuid, dest, stopList, PID, RID, sid, processors)

    if source == "454_fastq":
        errorText = uploadWithFastq(request, nameDict, selDict, refDict, p_uuid, dest, stopList, PID, RID, sid, processors)

    if source == "miseq":
        errorText = uploadWithMiseq(request, nameDict, selDict, refDict, p_uuid, dest, stopList, PID, RID, sid, processors, platform)

    if source == "biom":
        minConfidence = curData['cutoff']
        errorText = uploadWithBiom(request, nameDict, selDict, refDict, p_uuid, dest, minConfidence, stopList, PID, RID, sid, processors)

    if errorText == "Stop":  # upload subfuncs all return "Stop" when told to end early, saving cleanup for here
        functions.remove_proj(dest, request.user)
        transaction.savepoint_rollback(sid)
        return upStop(request)

    if errorText != "None":
        return upErr("Error during " + str(source) + ":" + str(errorText), request, dest, sid)

    end = datetime.datetime.now()
    debug('Total time for ' + str(request.user.username) + '\'s upload:', end - start)
    # TODO 1.3 verify files are winding up in the right place after upload is done
    myProject = Project.objects.get(projectid=p_uuid)
    myProject.wip = False
    myProject.save()

    # completely finished the upload!
    # update permissions lists, parse_project only handles this on status change during update (in case of reset)
    if myProject.status == "private":
        perms.newPriv(p_uuid)
    if myProject.status == "public":
        perms.newPub(p_uuid)

    if request.user.is_superuser:
        projects = Reference.objects.all().order_by('projectid__project_name', 'path')
    elif request.user.is_authenticated():
        projects = Reference.objects.all().order_by('projectid__project_name', 'path').filter(
            author=request.user)

    return render(
        request,
        'process.html',
        {'projects': projects,
         'form1': UploadForm1,
         'form2': UploadForm2,
         'error': "Upload complete"}
    )


def bubbleFiles(dest):
    # recursively find all files in dest directory
    # move each file to dest, delete subdirectories
    for root, dirnames, filenames in os.walk(dest):
        for filename in fnmatch.filter(filenames, '*'):
            shutil.move(os.path.join(root, filename), os.path.join(dest, filename))
    return


def prepMultiFiles(file_list, dest, mothurdest, outFileName):   # outFileName is 'temp.qual' var
    tempList = []
    for file in file_list:
        splitFile = file.split('/')
        end = len(splitFile) - 1
        try:
            copyFromUpload(file, dest, splitFile[end])  # into project directory with ye
            tar = tarfile.open(os.path.join(dest, splitFile[end]))   # try to untar the file
            tarParts = tar.getmembers()
            tar.extractall(path=mothurdest)
            for tarp in tarParts:
                if os.name == 'nt':
                    splitTar = tarp.name.split('\\')
                    splitEnd = splitTar[len(splitTar)-1]
                    splitSplit = splitEnd.split('.')
                    if splitSplit[len(splitSplit)-1] == "qual" or splitSplit[len(splitSplit)-1] == "fna":
                        myStr = "mothur\\temp\\" + str(splitTar[len(splitTar)-1])
                        tempList.append(myStr)
                else:
                    splitTar = tarp.name.split('/')
                    splitEnd = splitTar[len(splitTar)-1]
                    splitSplit = splitEnd.split('.')
                    if splitSplit[len(splitSplit)-1] == "qual" or splitSplit[len(splitSplit)-1] == "fna":
                        myStr = "mothur/temp/" + str(splitEnd)
                        tempList.append(myStr)
            tar.close()
        except Exception:
            try:
                zip = zipfile.ZipFile(os.path.join(dest, splitFile[end]))    # attempt to unzip
                zipParts = zip.namelist()
                zip.extractall(mothurdest)
                for zips in zipParts:
                    if os.name == 'nt':
                        splitZip = zips.split('\\')
                        splitEnd = splitZip[len(splitZip)-1]
                        splitSplit = splitEnd.split('.')
                        if splitSplit[len(splitSplit)-1] == "qual" or splitSplit[len(splitSplit)-1] == "fna":
                            myStr = "mothur\\temp\\" + str(splitZip[len(splitZip)-1])
                            tempList.append(myStr)
                    else:
                        splitZip = zips.split('/')
                        splitEnd = splitZip[len(splitZip)-1]
                        splitSplit = splitEnd.split('.')
                        if splitSplit[len(splitSplit)-1] == "qual" or splitSplit[len(splitSplit)-1] == "fna":
                            myStr = "mothur/temp/" + str(splitZip[len(splitZip)-1])
                            tempList.append(myStr)
                zip.close()
            except Exception:
                copyFromUpload(file, mothurdest, splitFile[end])  # copy to mothurdest because tar and zip failed
                if os.name == 'nt':
                    myStr = "mothur\\temp\\" + str(splitFile[end])
                else:
                    myStr = "mothur/temp/" + str(splitFile[end])
                tempList.append(myStr)
                if len(file_list) == 1:
                    # post as outfile name if only one
                    copyFromUpload(file, mothurdest, outFileName)

    bubbleFiles(mothurdest)  # bubble up subdirected files, templist already has file names chopped
    return tempList


def remProjectFiles(request):
    if request.is_ajax():
        allJson = request.GET['all']
        data = json.loads(allJson)
        refList = data['paths']
        functions.remove_list(refList, request.user)    # TODO 1.3 fetch errors from function calls
        results = {'error': 'none'}
        myJson = json.dumps(results)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "REM_FILES")


def removeUploads(request):  # removes files from user_upload directory based on request from upload page
    # currently assumes timestamp on folders to be unique (as only one batch of files can be uploaded at a time)
    # so asserting dataqueue is implemented and working properly, attempt to delete files at path+reflist
    # use try except, if it fails try it on another folder set with permissions (later)
    if request.is_ajax():
        allJson = request.GET['all']
        data = json.loads(allJson)
        folderList = data['folders']
        fileList = data['files']

        print "Deleting uploaded folders:", folderList, "and files:", fileList

        path = str(os.getcwd()) + "/user_uploads/"

        # SECURITY CONCERN: Traversal attacks, deleting files outside of folder using /../
        # Verify that path given is part of natural tree?
        # could just screen all incoming strings for the .. substring, flag em and drop em

        # tree gets populated by permissions, should verify when using though
        # need to get user profile permissions as well as screen each top level for being correct
        # AFTER erroring the heck out when someone sends a .. or any other sensitive character combinations
        username = request.user.username

        # check if username is request.user.username OR contained in user's UserProfile.hasPermsFrom.split(';') list
        # if ANY permissions issues come up, cancel the whole process
        for curFolder in folderList:
            if curFolder.find("..") != -1:
                print "Security!!! Folders w/ malicious characters,", username, "did it!"
                results = {'error': 'Illegal characters in request'}
                myJson = json.dumps(results)
                return HttpResponse(myJson)
            else:
                firstDir = curFolder.split('/')[0]
                if firstDir == username or request.user.is_superuser:
                    # firstDir in UserProfile.objects.get(user=request.user).hasPermsFrom.split(';'): # deletion by perm
                    curPath = os.path.join(path, curFolder)
                    try:
                        shutil.rmtree(curPath)
                    except Exception as e:
                        print "Error while deleting path:", curFolder, ":", e
                else:
                    print "Security!!! Folders outside of given permissions,", username, "did it!"
                    results = {'error': 'Illegal access attempt'}
                    myJson = json.dumps(results)
                    return HttpResponse(myJson)

        for curFile in fileList:
            if curFile.find("..") != -1:
                print "Security!!! Filepaths w/ malicious characters,", username, "did it!"
                results = {'error': 'Illegal characters in request'}
                myJson = json.dumps(results)
                return HttpResponse(myJson)
            else:
                firstDir = curFile.split('/')[0]
                if firstDir == username or firstDir in UserProfile.objects.get(user=request.user).hasPermsFrom.split(';'):
                    curPath = os.path.join(path, curFile)
                    try:
                        os.remove(curPath)
                    except Exception as e:
                        print "Error while deleting file:", curFile, ":", e  # TODO 1.4 error often triggers due to parent directory being deleted first
                else:
                    print "Security!!! Files outside of given permissions,", username, "did it!"
                    results = {'error': 'Illegal access attempt'}
                    myJson = json.dumps(results)
                    return HttpResponse(myJson)

        results = {'error': 'none'}
        myJson = json.dumps(results)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "REM_UPLOADS")


def projectTableJSON(request):
    if request.is_ajax():
        jsonSamples = request.GET['key']
        selSamples = json.loads(jsonSamples)

        if selSamples:
            if request.user.is_superuser:
                qs = Sample.objects.filter(sampleid__in=selSamples)
            elif request.user.is_authenticated():
                path_list = Sample.objects.filter(refid__author=request.user).values_list('sampleid', flat=True)
                qs = Sample.objects.filter( Q(sampleid__in=path_list) | Q(projectid__status='public') ).filter(sampleid__in=selSamples)
            else:
                qs = Sample.objects.none()
        else:
            qs = Sample.objects.none()

        results = {}
        qs1 = qs.values_list(
            "projectid",
            "projectid__status",
            "projectid__project_name",
            "sample_name",
            "projectid__project_desc",
            "projectid__start_date",
            "projectid__end_date",
            "projectid__pi_last",
            "projectid__pi_first",
            "projectid__pi_affiliation",
            "projectid__pi_email",
            "projectid__pi_phone"
        )

        qs1 = [[u'nan' if x is None else x for x in c] for c in qs1]
        results['data'] = list(qs1)
        myJson = ujson.dumps(results, ensure_ascii=False)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "PROJ_TABLE")


def sampleTableJSON(request):
    if request.is_ajax():
        jsonSamples = request.GET['key']
        selSamples = json.loads(jsonSamples)

        if selSamples:
            if request.user.is_superuser:
                qs = Sample.objects.filter(sampleid__in=selSamples)
            elif request.user.is_authenticated():
                path_list = Sample.objects.filter(refid__author=request.user).values_list('sampleid', flat=True)
                qs = Sample.objects.filter( Q(sampleid__in=path_list) | Q(projectid__status='public') ).filter(sampleid__in=selSamples)
            else:
                qs = Sample.objects.none()
        else:
            qs = Sample.objects.none()

        results = {}
        qs1 = qs.values_list(
            "projectid__project_name",
            "refid",
            "sampleid",
            "sample_name",
            "organism",
            "collection_date",
            "depth",
            "elev",
            "seq_platform",
            "seq_gene",
            "seq_gene_region",
            "seq_barcode",
            "seq_for_primer",
            "seq_rev_primer",
            "env_biome",
            "env_feature",
            "env_material",
            "geo_loc_country",
            "geo_loc_state",
            "geo_loc_city",
            "geo_loc_farm",
            "geo_loc_plot",
            "latitude",
            "longitude",
            "annual_season_precpt",
            "annual_season_temp"
        )

        qs1 = [[u'nan' if x is None else x for x in c] for c in qs1]
        results['data'] = list(qs1)
        myJson = ujson.dumps(results, ensure_ascii=False)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "SAMP_TABLE")


def referenceTableJSON(request):
    if request.is_ajax():
        jsonSamples = request.GET['key']
        selSamples = json.loads(jsonSamples)

        if selSamples:
            if request.user.is_superuser:
                qs = Reference.objects.filter(sample__sampleid__in=selSamples).distinct()
            elif request.user.is_authenticated():
                qs = Reference.objects.filter( Q(author=request.user) | Q(projectid__status='public') ).distinct()
            else:
                qs = Reference.objects.none()
        else:
            qs = Reference.objects.none()

        results = {}
        qs1 = qs.values_list(
            "projectid__project_name",
            "projectid",
            "refid",
            "raw",
            "path",
            "source",
            "author_id__username"
        )

        qs1 = [[u'nan' if x is None else x for x in c] for c in qs1]
        results['data'] = list(qs1)
        myJson = ujson.dumps(results, ensure_ascii=False)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "REF_TABLE")


def airTableJSON(request):
    if request.is_ajax():
        jsonSamples = request.GET['key']
        selSamples = json.loads(jsonSamples)

        if selSamples:
            if request.user.is_superuser:
                qs = Air.objects.filter(sampleid__in=selSamples)
            elif request.user.is_authenticated():
                path_list = Air.objects.filter(refid__author=request.user).values_list('sampleid', flat=True)
                qs = Air.objects.filter( Q(sampleid__in=path_list) | Q(projectid__status='public') ).filter(sampleid__in=selSamples)
            else:
                qs = Air.objects.none()
        else:
            qs = Air.objects.none()

        results = {}
        qs1 = qs.values_list(
            "projectid__project_name",
            "sampleid__sample_name",
            "barometric_press",
            "carb_dioxide",
            "carb_monoxide",
            "chem_admin_term",
            "chem_admin_time",
            "elev",
            "humidity",
            "methane",
            "organism_type",
            "organism_count",
            "oxy_stat_samp",
            "oxygen",
            "perturbation_type",
            "perturbation_interval",
            "pollutants_type",
            "pollutants_concentration",
            "rel_to_oxygen",
            "resp_part_matter_substance",
            "resp_part_matter_concentration",
            "samp_collect_device",
            "samp_mat_process",
            "samp_salinity",
            "samp_size",
            "samp_store_dur",
            "samp_store_loc",
            "samp_store_temp",
            "solar_irradiance",
            "temp",
            "ventilation_rate",
            "ventilation_type",
            "volatile_org_comp_name",
            "volatile_org_comp_concentration",
            "wind_direction",
            "wind_speed",
        )

        qs1 = [[u'nan' if x is None else x for x in c] for c in qs1]
        results['data'] = list(qs1)
        myJson = ujson.dumps(results, ensure_ascii=False)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "AIR_TABLE")


def associatedTableJSON(request):
    if request.is_ajax():
        jsonSamples = request.GET['key']
        selSamples = json.loads(jsonSamples)

        if selSamples:
            if request.user.is_superuser:
                qs = Human_Associated.objects.filter(sampleid__in=selSamples)
            elif request.user.is_authenticated():
                path_list = Human_Associated.objects.filter(refid__author=request.user).values_list('sampleid', flat=True)
                qs = Human_Associated.objects.filter( Q(sampleid__in=path_list) | Q(projectid__status='public') ).filter(sampleid__in=selSamples)
            else:
                qs = Human_Associated.objects.none()
        else:
            qs = Human_Associated.objects.none()

        results = {}
        qs1 = qs.values_list(
            "projectid__project_name",
            "sampleid__sample_name",
            "samp_collect_device",
            "samp_mat_process",
            "samp_size",
            "samp_store_temp",
            "samp_store_dur",
            "samp_type",
            "samp_location",
            "samp_temp",
            "samp_ph",
            "samp_oxy_stat",
            "samp_salinity",
            "host_subject_id",
            "host_age",
            "host_pulse",
            "host_gender",
            "host_ethnicity",
            "host_height",
            "host_weight",
            "host_bmi",
            "host_weight_loss_3_month",
            "host_body_temp",
            "host_occupation",
            "pet_farm_animal",
            "smoker",
            "diet_type",
            "diet_duration",
            "diet_frequency",
            "diet_last_six_month",
            "last_meal",
            "medic_hist_perform",
            "disease_type",
            "disease_location",
            "disease_duration",
            "organism_count",
            "tumor_location",
            "tumor_mass",
            "tumor_stage",
            "drug_usage",
            "drug_type",
            "drug_duration",
            "drug_frequency",
            "perturbation",
            "pert_type",
            "pert_duration",
            "pert_frequency",
            "fetal_health_stat",
            "amniotic_fluid_color",
            "gestation_stat",
            "maternal_health_stat"
        )

        qs1 = [[u'nan' if x is None else x for x in c] for c in qs1]
        results['data'] = list(qs1)
        myJson = ujson.dumps(results, ensure_ascii=False)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "ASSOC_TABLE")


def microbialTableJSON(request):
    if request.is_ajax():
        jsonSamples = request.GET['key']
        selSamples = json.loads(jsonSamples)

        if selSamples:
            if request.user.is_superuser:
                qs = Microbial.objects.filter(sampleid__in=selSamples)
            elif request.user.is_authenticated():
                path_list = Microbial.objects.filter(refid__author=request.user).values_list('sampleid', flat=True)
                qs = Microbial.objects.filter( Q(sampleid__in=path_list) | Q(projectid__status='public') ).filter(sampleid__in=selSamples)
            else:
                qs = Microbial.objects.none()
        else:
            qs = Microbial.objects.none()

        results = {}
        qs1 = qs.values_list(
            "projectid__project_name",
            "sampleid__sample_name",
            "alkalinity",
            "alkyl_diethers",
            "altitude",
            "aminopept_act",
            "ammonium",
            "bacteria_carb_prod",
            "biomass_part_name",
            "biomass_amount",
            "bishomohopanol",
            "bromide",
            "calcium",
            "carb_nitro_ratio",
            "chem_administration_term",
            "chem_administration_time",
            "chloride",
            "chlorophyll",
            "diether_lipids_name",
            "diether_lipids_concentration",
            "diss_carb_dioxide",
            "diss_hydrogen",
            "diss_inorg_carb",
            "diss_org_carb",
            "diss_org_nitro",
            "diss_oxygen",
            "glucosidase_act",
            "magnesium",
            "mean_frict_vel",
            "mean_peak_frict_vel",
            "methane",
            "n_alkanes_name",
            "n_alkanes_concentration",
            "nitrate",
            "nitrite",
            "nitro",
            "org_carb",
            "org_matter",
            "org_nitro",
            "organism_name",
            "organism_count",
            "oxy_stat_samp",
            "part_org_carb",
            "perturbation_type",
            "perturbation_interval",
            "petroleum_hydrocarb",
            "ph",
            "phaeopigments_type",
            "phaeopigments_concentration",
            "phosphate",
            "phosplipid_fatt_acid_name",
            "phosplipid_fatt_acid_conc",
            "potassium",
            "pressure",
            "redox_potential",
            "rel_to_oxygen",
            "salinity",
            "samp_collect_device",
            "samp_mat_process",
            "samp_size",
            "samp_store_dur",
            "samp_store_loc",
            "samp_store_temp",
            "silicate",
            "sodium",
            "sulfate",
            "sulfide",
            "temp",
            "tot_carb",
            "tot_nitro",
            "tot_org_carb",
            "turbidity",
            "water_content"
        )

        qs1 = [[u'nan' if x is None else x for x in c] for c in qs1]
        results['data'] = list(qs1)
        myJson = ujson.dumps(results, ensure_ascii=False)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "MICRO_TABLE")


def soilTableJSON(request):
    if request.is_ajax():
        jsonSamples = request.GET['key']
        selSamples = json.loads(jsonSamples)

        if selSamples:
            if request.user.is_superuser:
                qs = Soil.objects.filter(sampleid__in=selSamples)
            elif request.user.is_authenticated():
                path_list = Soil.objects.filter(refid__author=request.user).values_list('sampleid', flat=True)
                qs = Soil.objects.filter( Q(sampleid__in=path_list) | Q(projectid__status='public') ).filter(sampleid__in=selSamples)
            else:
                qs = Soil.objects.none()
        else:
            qs = Soil.objects.none()

        results = {}
        qs1 = qs.values_list(
            "projectid__project_name",
            "sampleid__sample_name",
            "samp_collection_device",
            "samp_size",
            "samp_depth",
            "samp_prep",
            "samp_sieve_size",
            "samp_store_dur",
            "samp_store_loc",
            "samp_store_temp",
            "samp_weight_dna_ext",
            "pool_dna_extracts",
            "fao_class",
            "local_class",
            "texture_class",
            "porosity",
            "profile_position",
            "slope_aspect",
            "slope_gradient",
            "bulk_density",
            "drainage_class",
            "water_content_soil",
            "cur_land_use",
            "cur_vegetation",
            "cur_crop",
            "cur_cultivar",
            "crop_rotation",
            "cover_crop",
            "fert_amendment_class",
            "fert_placement",
            "fert_type",
            "fert_tot_amount",
            "fert_N_tot_amount",
            "fert_P_tot_amount",
            "fert_K_tot_amount",
            "irrigation_type",
            "irrigation_tot_amount",
            "residue_removal",
            "residue_growth_stage",
            "residue_removal_percent",
            "tillage_event",
            "tillage_event_depth",
            "amend1_class",
            "amend1_active_ingredient",
            "amend1_tot_amount",
            "amend2_class",
            "amend2_active_ingredient",
            "amend2_tot_amount",
            "amend3_class",
            "amend3_active_ingredient",
            "amend3_tot_amount",
            "rRNA_copies",
            "microbial_biomass_C",
            "microbial_biomass_N",
            "microbial_respiration",
            "soil_pH",
            "soil_EC",
            "soil_C",
            "soil_OM",
            "soil_N",
            "soil_NO3_N",
            "soil_NH4_N",
            "soil_P",
            "soil_K",
            "soil_S",
            "soil_Zn",
            "soil_Fe",
            "soil_Cu",
            "soil_Mn",
            "soil_Ca",
            "soil_Mg",
            "soil_Na",
            "soil_B",
            "plant_C",
            "plant_N",
            "plant_P",
            "plant_K",
            "plant_Ca",
            "plant_Mg",
            "plant_S",
            "plant_Na",
            "plant_Cl",
            "plant_Al",
            "plant_B",
            "plant_Cu",
            "plant_Fe",
            "plant_Mn",
            "plant_Zn",
            "crop_tot_biomass_fw",
            "crop_tot_biomass_dw",
            "crop_tot_above_biomass_fw",
            "crop_tot_above_biomass_dw",
            "crop_tot_below_biomass_fw",
            "crop_tot_below_biomass_dw",
            "harv_fraction",
            "harv_fresh_weight",
            "harv_dry_weight",
            "ghg_chamber_placement",
            "ghg_N2O",
            "ghg_CO2",
            "ghg_NH4",
            "soil_texture_sand",
            "soil_texture_silt",
            "soil_texture_clay",
            "soil_water_cap",
            "soil_water_cap_rating",
            "soil_surf_hardness",
            "soil_surf_hardness_rating",
            "soil_subsurf_hardness",
            "soil_subsurf_hardness_rating",
            "soil_agg_stability",
            "soil_agg_stability_rating",
            "soil_organic_matter",
            "soil_organic_matter_rating",
            "soil_ACE_protein_index",
            "soil_ACE_protein_index_rating",
            "soil_root_pathogen_pressure",
            "soil_root_pathogen_pressure_rating",
            "soil_respiration_four_day",
            "soil_respiration_four_day_rating",
            "soil_active_C",
            "soil_active_C_rating",
            "soil_pH_rating",
            "soil_p_rating",
            "soil_k_rating",
            "soil_minor_elements_rating",
            "CASH_SHI_rating"
        )

        qs1 = [[u'nan' if x is None else x for x in c] for c in qs1]
        results['data'] = list(qs1)
        myJson = ujson.dumps(results, ensure_ascii=False)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "SOIL_TABLE")


def waterTableJSON(request):
    if request.is_ajax():
        jsonSamples = request.GET['key']
        selSamples = json.loads(jsonSamples)

        if selSamples:
            if request.user.is_superuser:
                qs = Water.objects.filter(sampleid__in=selSamples)
            elif request.user.is_authenticated():
                path_list = Water.objects.filter(refid__author=request.user).values_list('sampleid', flat=True)
                qs = Water.objects.filter( Q(sampleid__in=path_list) | Q(projectid__status='public') ).filter(sampleid__in=selSamples)
            else:
                qs = Water.objects.none()
        else:
            qs = Water.objects.none()

        results = {}
        qs1 = qs.values_list(
            "projectid__project_name",
            "sampleid__sample_name",
            "alkalinity",
            "alkyl_diethers",
            "altitude",
            "aminopept_act",
            "ammonium",
            "atmospheric_data",
            "bac_prod",
            "bac_resp",
            "bacteria_carb_prod",
            "biomass_part_name",
            "biomass_amount",
            "bishomohopanol",
            "bromide",
            "calcium",
            "carb_nitro_ratio",
            "chem_administration_name",
            "chem_administration_time",
            "chloride",
            "chlorophyll",
            "conduc",
            "density",
            "diether_lipids",
            "diss_carb_dioxide",
            "diss_hydrogen",
            "diss_inorg_carb",
            "diss_inorg_nitro",
            "diss_inorg_phosp",
            "diss_org_carb",
            "diss_org_nitro",
            "diss_oxygen",
            "down_par",
            "elev",
            "fluor",
            "glucosidase_act",
            "light_intensity",
            "magnesium",
            "mean_frict_vel",
            "mean_peak_frict_vel",
            "n_alkanes",
            "nitrate",
            "nitrite",
            "nitro",
            "org_carb",
            "org_matter",
            "org_nitro",
            "organism_name",
            "organism_count",
            "oxy_stat_samp",
            "part_org_carb",
            "part_org_nitro",
            "perturbation_type",
            "perturbation_interval",
            "pretroleum_hydrocarb",
            "ph",
            "phaeopigments",
            "phosphate",
            "phosplipid_fatt_acid",
            "photon_flux",
            "potassium",
            "pressure",
            "primary_prod",
            "redox_potential",
            "rel_to_oxygen",
            "samp_mat_process",
            "samp_salinity",
            "samp_size",
            "samp_store_dur",
            "samp_store_loc",
            "samp_store_temp",
            "samp_vol_we_dna_ext",
            "silicate",
            "sodium",
            "soluble_react_phosp",
            "source_material_id",
            "sulfate",
            "sulfide",
            "suspen_part_matter",
            "temp",
            "tidal_stage",
            "tot_depth_water_col",
            "tot_diss_nitro",
            "tot_inorg_nitro",
            "tot_nitro",
            "tot_part_carb",
            "tot_phosp",
            "water_current_direction",
            "water_current_magnitude"
        )

        qs1 = [[u'nan' if x is None else x for x in c] for c in qs1]
        results['data'] = list(qs1)
        myJson = ujson.dumps(results, ensure_ascii=False)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "WATER_TABLE")


def userTableJSON(request):
    if request.is_ajax():
        jsonSamples = request.GET['key']
        selSamples = json.loads(jsonSamples)

        if selSamples:
            if request.user.is_superuser:
                qs = UserDefined.objects.filter(sampleid__in=selSamples)
            elif request.user.is_authenticated():
                path_list = UserDefined.objects.filter(refid__author=request.user).values_list('sampleid', flat=True)
                qs = UserDefined.objects.filter( Q(sampleid__in=path_list) | Q(projectid__status='public') ).filter(sampleid__in=selSamples)
            else:
                qs = UserDefined.objects.none()
        else:
            qs = UserDefined.objects.none()

        results = {}
        qs1 = qs.values_list(
            "projectid__project_name",
            "sampleid__sample_name",
            "usr_cat1",
            "usr_cat2",
            "usr_cat3",
            "usr_cat4",
            "usr_cat5",
            "usr_cat6",
            "usr_quant1",
            "usr_quant2",
            "usr_quant3",
            "usr_quant4",
            "usr_quant5",
            "usr_quant6"
        )

        qs1 = [[u'nan' if x is None else x for x in c] for c in qs1]
        results['data'] = list(qs1)
        myJson = ujson.dumps(results, ensure_ascii=False)
        return HttpResponse(myJson)
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "USER_TABLE")


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def select(request):
    # send Locations to page, should be based on visible samples and projects
    # loop through all visible samples and store locations via dictionary
    # key off each location, save the samples and the projects they are from
    # on page, display a marker at each unique location, with related projects on mouseover, and samples on select
    # populate a tree with samples from selected marker, checkbox links to actual sample list, mirrors clicks
    # data format: locationDict, keys are lat-lon combo (+- not NSEW, rounded to tenths or hundredths)
    # values are project titles, in a semi-colon separated string
    # project titles are appended by a list of sample ids which have approximately this key's coordinates

    # need to send both project names and sample ids up to page, associated with coordinates, need parent-child nodes up
    # locationDict['lat-lon'] = projectName[0]+':'+P0sample[1]+':'+P0sample[4]+';'+projectName[2]+':'+P2sample[3]

    functions.log(request, "PAGE", "SELECT")
    if request.method == 'POST':
        uploaded = request.FILES["normFile"]
        # we specifically want myphylodb.biom here
        if uploaded.name.split('.')[-1] != "biom":
            print "Wrong file type for pre normalized data"
            functions.securityLog(request, "SelectNormalized", "incorrect file type, wanted 'biom' got '"+str(uploaded.name.split('.')[-1])+"'")
            return render(
                request,
                'select.html',
                {'form9': UploadForm9,
                 'selList': '',
                 'normpost': 'error',
                 'method': 'POST',
                 'locationBasedSampleDict': {}}
            )
        # derive directory from user and make certain it exists
        myDir = 'myPhyloDB/media/usr_temp/' + str(request.user) + '/'
        if not os.path.exists(myDir):
            os.makedirs(myDir)
        else:
            shutil.rmtree(myDir)

        # save file user handed us to disc, was just in memory before
        functions.handle_uploaded_file(uploaded, myDir, uploaded.name)

        # parse given file for data via exploding panda
        filename = str(myDir) + uploaded.name
        try:
            # get sampleIDs here, read the file, go to 'columns' then 'id' for each column entry. Metadata is there too but not needed now
            # since this should have been built my phylodb, we can just double check samples exist and match perms, then
            # save pkl file with selection in it, rename file in correct spot and skip to analyses
            with open(filename) as f:
                myJson = json.load(f)
            selList = []
            for col in myJson['columns']:
                selList.append(col['id'])
            functions.write_taxa_summary(myDir+"available_taxa_summary.pkl", myJson['rows'])

        except Exception:
            return render(
                request,
                'select.html',
                {'form9': UploadForm9,
                 'selList': '',
                 'normpost': 'error',
                 'method': 'POST',
                 'locationBasedSampleDict': {}}
            )

        # save selected samples file to users temp/ folder
        #selList = list(set(savedDF['sampleid']))
        path = str(myDir) + 'usr_sel_samples.pkl'

        with open(path, 'wb') as f:
            pickle.dump(selList, f)

        # save as norm samples because we are skipping normalization
        path = str(myDir) + 'usr_norm_samples.pkl'

        with open(path, 'wb') as f:
            pickle.dump(selList, f)

        # Update dataID to prevent expired selection bugs
        thisUser = request.user.profile
        thisUser.dataID = uuid4().hex
        thisUser.save()
        # so now we have samples selected and dataID updated, now we functionally skip norm by creating a phyloseq.biom

        #projectList = list(set(savedDF['projectid']))   # missing projectid in savedDF, can add based on samples
        projectList = []
        permList = perms.getViewProjects(request)
        allValid = True
        for samp in selList:
            myProj = Sample.objects.get(sampleid=samp).projectid  # TODO 1.3 refactor ID chain, id is actually a project
            if myProj not in projectList:   # avoid duplicates
                if myProj in permList:  # verify user has permission to view this project
                    projectList.append(myProj.projectid)
                else:
                    allValid = False    # invalid permissions
                    functions.securityLog(request, "SelectProjects", "permissions abuse attempt")

        #print "ProjectList:", projectList
        if not allValid: #Project.objects.filter(projectid__in=projectList).exists():
            return render(
                request,
                'select.html',
                {'form9': UploadForm9,
                 'selList': '',
                 'normpost': 'error',
                 'method': 'POST',
                 'locationBasedSampleDict': {}}

            )
        #print "Projects exist"
        if not Reference.objects.filter(projectid__in=projectList).exists():
            return render(
                request,
                'select.html',
                {'form9': UploadForm9,
                 'selList': '',
                 'normpost': 'error',
                 'method': 'POST',
                 'locationBasedSampleDict': {}}
            )
        #print "References exist"
        if not Sample.objects.filter(sampleid__in=selList).exists():
            return render(
                request,
                'select.html',
                {'form9': UploadForm9,
                 'selList': '',
                 'normpost': 'error',
                 'method': 'POST',
                 'locationBasedSampleDict': {}}
            )

        # at this point, we've done everything we need to based on the contents of the biom file
        # so now we just make sure its saved in the right place via a file rename
        myDir = 'myPhyloDB/media/usr_temp/' + str(request.user) + '/'
        path = str(myDir) + 'myphylodb.biom'
        try:
            os.remove(path)
        except Exception as ex:
            pass
        os.rename(filename, path)

        # TODO 1.3 either here or in JS, restrict taxaLevel select based on data (ON ALL ANALYSES ACTUALLY)
        # some projects don't have otu or such, we should probably just check the biom file after norm to see what level we can go to
        # set some sort of flag every time norm finishes (and here as an equivalent) for this user, which is loaded by all pages?
        # how to load this in browser without adding to each page, would need to be part of the shared function
        # which controls the taxa level selector in the first place


        # just need myphylodb.biom to be present in usr_temp/usr/ for analyses to find
        # the file they gave us is in this directory already, just need to rename it (is it phyloseq type? or myphylodb?)
        #with open(path, 'w') as outfile:
        #    json.dump(biom, outfile, ensure_ascii=True, indent=4)
        # TODO 1.3 normalize gets skipped with this process, but the norm button is still visible. HIDE
        # when does this phyloseq.biom get used? I can't find it. We use myphylodb.biom that norm makes using select pkl
        # so in theory we just need to parse out the select pkl and go to norm with it, though that won't preserve taxa
        # since the goal of this is to preserve CORE taxa filtering via biom format, we should skip the norm step by
        # saving a myphylodb.biom with CORE, parsing it for selected samples here before going straight to analyses

        # main difference between current CORE biom export and myphylodb.biom is the id:name pair, whereas phyloseq is
        # exclusively names, and CORE export is currently only id.
        # if we do just make CORE export a myphylodb.biom, we can dramatically simplify this pipeline by parsing samples
        # and saving the biom file in the right location alongside usr_sel_samples.pkl, should contain all needed info
        # between those two files

        # We need to export phyloseq style and just read the samples, then save the file for later use
        # If we put this in norm (ie don't allow normalization from a biom file), we can just export myphylodb.biom
        # files from analysis pages and such for quick subsetting (CORE in particular)
        # for page visibility reasons, this is select instead of norm
        # but its functionally both select and normalize, so we need to show analyses afterwards instead of norm (norm shouldn't work)

        return render(
            request,
            'select.html',
            {'form9': UploadForm9,
             'selList': selList,
             'normpost': 'Success',
             'method': 'GET'
             }
             #'locationBasedSampleDict': {}}
        )

    else:
        return render(
            request,
            'select.html',
            {'form9': UploadForm9,
             'selList': '',
             'normpost': '',
             'method': 'GET'
             }
        )
        # TODO 1.4 re-implement the map, properly, using esri instead of google maps (likely no changes on this end)
        ''''# insert map population logic here
        visibleProjects = perms.getViewProjects(request)
        # get all samples for visibleProjects, put into dictionary keyed by rounded coordinates, link projectname?
        sampleSetList = []
        locationBasedSampleDict = {}
        for proj in visibleProjects:
            projSamples = Sample.objects.filter(projectid=proj.projectid)
            sampleSetList.append(projSamples)
            # print "proj:", proj.project_name
            for samp in projSamples:
                try:
                    # .00001 ~ 1 meter
                    # .01  ~ 1 km
                    # round to hundredths place for the sake of grouping, made roundOff val a constant for fine tuning
                    # 1/roundOff = decimal place to floor (always rounded down for simplicity/speed)
                    # need to get roundOff val based on 10^input from slider, somehow push request back to page?

                    # disabled map code for now browser-side

                    roundOff = 100  # bigger numbers here mean more precise coordinates, smaller for cleaner grouping
                    myLat = samp.latitude
                    myLon = samp.longitude
                    if myLat is not None and myLon is not None:
                        # print "Samp:", samp.sample_name, "lat:lon ~~~", myLat, ":", myLon
                        # round and concatenate into colon separable string
                        coordKey = str(math.floor(myLat*roundOff)/roundOff)+":"+str(math.floor(myLon*roundOff)/roundOff)
                        # store in dictionary with key of coordKey, value should be sample id, name, and project name
                        myValues = str(samp.sampleid)+":"+str(samp.sample_name)+":"+str(proj.project_name)
                        if coordKey in locationBasedSampleDict.keys():
                            locationBasedSampleDict[coordKey] += ";"+myValues
                        else:
                            locationBasedSampleDict[coordKey] = myValues
                except Exception as e:
                    print "Error with samp:", e
        # group by location data, key search?
        return render(
            request,
            'select.html',
            {'form9': UploadForm9,
             'selList': '',
             'normpost': '',
             'method': 'GET',
             'locationBasedSampleDict': locationBasedSampleDict
             }
        )'''


def taxaJSON(request):
    results = {}
    qs1 = OTU_99.objects.values_list('kingdomid', 'kingdomid__kingdomName', 'phylaid', 'phylaid__phylaName', 'classid', 'classid__className', 'orderid', 'orderid__orderName', 'familyid', 'familyid__familyName', 'genusid', 'genusid__genusName', 'speciesid', 'speciesid__speciesName', 'otuid', 'otuName', 'otuSeq')
    results['data'] = list(qs1)
    myJson = ujson.dumps(results, ensure_ascii=False)
    return HttpResponse(myJson)


# TODO 1.3 html check
def taxa(request):
    functions.log(request, "PAGE", "TAXA")
    return render(
        request,
        'taxa.html'
    )


def pathJSON(request):
    results = {}
    qs1 = ko_entry.objects.using('picrust').values_list('ko_lvl1_id', 'ko_lvl1_id__ko_lvl1_name', 'ko_lvl2_id', 'ko_lvl2_id__ko_lvl2_name', 'ko_lvl3_id', 'ko_lvl3_id__ko_lvl3_name', 'ko_lvl4_id', 'ko_orthology', 'ko_name', 'ko_desc')
    results['data'] = list(qs1)
    myJson = ujson.dumps(results, ensure_ascii=False)
    return HttpResponse(myJson)


def printKoList(koList):
    print "Printing koList: ", koList, " (L: ", len(koList), " )"
    allthenames = []
    fullOtuList = koOtuList.objects.all()
    for thing in fullOtuList:
        allthenames.append(thing.koID)
    print "Printing koOtuList: ", allthenames

    print "First entry list:"
    first = True
    for otuList in fullOtuList:
        if first:
            first = False
            myList = otuList.otuList
            print myList
            print "Entries in list:"
            for entry in myList:
                print entry


def printKoOtuList():
    print "Printing koOtuLists..."
    fullOtuList = koOtuList.objects.all()
    for thing in fullOtuList:
        print thing.koID, ":", thing.otuList


def getOtuFromKoList(koList):   # new qsFunc for speed sake, requires KoOtuLists to be fully populated
    # printKoList(koList)
    finalotuList = []
    otuDict = {}
    for ko in koList:
        try:
            curOtuList = koOtuList.objects.get(koID=ko).otuList.split(",")
            for otu in curOtuList:
                otuDict[otu] = otu
        except:
            pass
    for entry in otuDict:
        if entry != "":
            finalotuList.append(str(entry.replace(" u\'", "").replace("\'", "")))
    # need to strip the " u'" and "'" off the string, because unicode into exact string match

    return finalotuList


# if you actually need to call this for some reason, remove the return comment
def populateKoOtuList():    # for initial population, similar to old qsFunc. currently no UI hook (runtime is absurd)
    print "PopulateKoOtuList, currently disabled"
    return

    from django.db import connection

    print "Connection imported"

    # used for debugging, for the most part this whole function should NOT be called outside of specific maintenance
    # printKoOtuList()
    # return

    with connection.cursor() as cursor:
        print "Altering connection settings. Cursor ", cursor
        cursor.execute("PRAGMA synchronous = OFF")
        cursor.execute("PRAGMA temp_store = MEMORY")
        #cursor.execute("PRAGMA default_cache_size = 10000")
        cursor.execute("PRAGMA journal_mode = WAL")
        cursor.execute("PRAGMA synchronous")
        print "S: ", cursor.fetchall()
        cursor.execute("PRAGMA temp_store")
        print "T: ", cursor.fetchall()
        #cursor.execute("PRAGMA default_cache_size")
        #print "C: ", cursor.fetchall()
        cursor.execute("PRAGMA journal_mode")
        print "J: ", cursor.fetchall()
        print "Changed SQLite settings"  # verify these are temporary
        print "Tables:", connection.introspection.table_names()
        try:
            print "Recreating full koOtuList... this will take some time"

            koList = []
            records = ko_lvl1.objects.using('picrust').all()
            for record in records:
                koList.extend(record.ko_entry_set.values_list('ko_orthology', flat=True))

            print "ko1:", len(koList)

            records = ko_lvl2.objects.using('picrust').all()
            for record in records:
                koList.extend(record.ko_entry_set.values_list('ko_orthology', flat=True))

            print "ko2:", len(koList)

            records = ko_lvl3.objects.using('picrust').all()
            for record in records:
                koList.extend(record.ko_entry_set.values_list('ko_orthology', flat=True))

            print "ko3:", len(koList)

            koList.extend(ko_entry.objects.using('picrust').all())
            koLen = len(koList)
            print "KoList length:", koLen

            delIter = 0
            missed = 0
            delStart = time.time()
            # slower delete function, grants feedback
            for delKo in koList:
                try:
                    koOtuList.objects.get(koID=delKo).delete()
                except Exception:
                    missed += 1
                delIter += 1
                if delIter % 50 == 0:
                    print "Deleted", delIter, "missed", missed
                    print "Total time for deletion:", time.time()-delStart, "at", (delIter*100.0)/(1.0*koLen), "% complete"
                    missed = 0

            print "Deleting remaining entries"
            koOtuList.objects.all().delete()
            print "Old records wiped, here we go"

            otuList = OTU_99.objects.all().values_list('otuid', flat=True)
            otuLen = len(otuList)
            print "OtuList length:", otuLen
            curOtu = 0
            totalKos = 0
            if koList:
                koDict = {}
                koOtuDict = {}
                for ko in koList:
                    koDict[ko] = ko
                outputFreq = 50
                iterTime = 0
                totalTime = 0
                iter = 0
                for otu in otuList:
                    startTime = time.time()
                    curTotal = 0
                    creTotal = 0
                    missTotal = 0
                    try:
                        qs = PICRUSt.objects.using('picrust').filter(otuid=otu)
                        geneList = qs[0].geneList
                        geneList = geneList.replace("[", "")
                        geneList = geneList.replace("]", "")
                        geneList = geneList.replace("'", "")
                        geneList = geneList.replace(" ", "")
                        genes = geneList.split(",")
                        checkTime = 0
                        existsTime = 0
                        appendTime = 0
                        otherAppTime = 0
                        for gene in genes:
                            prevTime = time.time()
                            try:
                                if koDict[gene] == gene:  # check if gene is in koDict
                                    checkTime += time.time()-prevTime
                                    prevTime = time.time()
                                    # get or make koOtuList object for this ko
                                    if gene not in koOtuDict.keys():
                                        existsTime += time.time()-prevTime
                                        prevTime = time.time()

                                        koOtuDict[gene] = []

                                        appendTime += time.time()-prevTime
                                        prevTime = time.time()
                                        creTotal += 1
                                        totalKos += 1
                                    else:
                                        existsTime += time.time()-prevTime
                                        prevTime = time.time()

                                    koOtuDict[gene].append(otu)

                                    otherAppTime += time.time()-prevTime
                                    curTotal += 1
                                else:
                                    missTotal += 1
                            except:
                                missTotal += 1
                                pass  # can be a keyerror from gene not being in koDict

                    except Exception as e:
                        pass
                    curOtu += 1
                    thisTime = time.time()-startTime
                    iterTime += thisTime
                    totalTime += thisTime
                    iter += 1
                    if iter % outputFreq == 0:
                        print "Past", outputFreq, " ", iterTime/outputFreq, " vs Total ", totalTime/curOtu, " \tOtu: ", curOtu, " / ", otuLen
                        iterTime = 0

                    # early break statement for testing
                    #if iter == 50:
                    #    break

                # save to database from koOtuDict
                print "Completed main processing loop, saving to database"
                koSet = koOtuDict.keys()
                mass_create(koSet)
                print "Created new objects, saving values"
                iter = 0
                upTime = time.time()
                fiftyTime = time.time()
                for ko in koSet:
                    try:
                        thisKoOtuList = koOtuList.objects.get(koID=ko)
                        thisKoOtuList.otuList = koOtuDict[ko]   # save as list, ie comma delimiter with u'' encaps
                        # saving like this takes all of 5 minutes, changing reader code won't make a significant impact
                        thisKoOtuList.save()
                        iter += 1
                        if iter % 50 == 0:
                            print "Saved 50, took", time.time()-fiftyTime, "seconds. Total save time is", time.time()-upTime
                            fiftyTime = time.time()
                            connection.commit()
                            print "Committed"
                    except Exception as ex:
                        print "Error during save:", ex

            # get all ko genelists and search for matching otus (out of entire set)
            # add matching otus to related koOtuList entries
        except Exception as er:
            print "Problem with popKOL (M) ", er


# create koOtuList objects from a gene list
@transaction.atomic
def mass_create(koGenes):
    for newGene in koGenes:
        newList = koOtuList.objects.create(koID=newGene)
        newList.otuList = ""
        newList.save()


def qsFunc(koList):  # works, but is still slow, still leaks (although only maybe 100 MB at a time vs 10 GB)
    finalotuList = []
    otuList = OTU_99.objects.all().values_list('otuid', flat=True)
    if koList:
        for otu in otuList:
            try:
                qs = PICRUSt.objects.using('picrust').filter(otuid=otu)
                if any(i in qs[0].geneList for i in koList):
                    finalotuList.append(otu)
            except:
                pass
    return finalotuList


def compareMethods(ko):

    koList = []
    koList.append(ko)
    newOtu = getOtuFromKoList(koList)
    print "New:", len(newOtu)

    oldOtu = qsFunc(koList)
    print "Old:", len(oldOtu)

    shared = []
    uniqueNew = []
    uniqueOld = []

    for otu in oldOtu:
        if otu in newOtu:
            shared.append(otu)
        else:
            uniqueOld.append(otu)
    for otu in newOtu:
        if otu in oldOtu:
            pass
        else:
            uniqueNew.append(otu)

    print "Finished comparison"
    print len(shared), ";", len(uniqueNew), ";", len(uniqueOld)

    # new otu is missing some objects, picrust database bigger than otu_99?
    '''print "Checking genes of uniqueOld:"
    checkRes = checkGenes(uniqueOld, ko)
    print checkRes'''


def checkGenes(otuList, ko):
    ret = "Valid"
    errco = 0
    for otu in otuList:
        genes = PICRUSt.objects.using('picrust').filter(otuid=otu)[0].geneList
        if ko in genes:
            pass
        else:
            ret = "Invalid"
            errco += 1
    return ret+str(errco)


def checkOtuListIntegrity():
    otuLists = koOtuList.objects.all()
    print "KoOtuListList length:", len(otuLists)
    koList = []
    records = ko_lvl1.objects.using('picrust').all()
    for record in records:
        koList.extend(record.ko_entry_set.values_list('ko_orthology', flat=True))

    print "ko1:", len(koList)

    records = ko_lvl2.objects.using('picrust').all()
    for record in records:
        koList.extend(record.ko_entry_set.values_list('ko_orthology', flat=True))

    print "ko2:", len(koList)

    records = ko_lvl3.objects.using('picrust').all()
    for record in records:
        koList.extend(record.ko_entry_set.values_list('ko_orthology', flat=True))

    print "ko3:", len(koList)

    koList.extend(ko_entry.objects.using('picrust').all())

    print "KoList length:", len(koList)


def pathTaxaJSON(request):
    try:

        if request.is_ajax():
            wanted = request.GET['key']
            wanted = wanted.replace('"', '')
            koList = []

            if ko_lvl1.objects.using('picrust').filter(ko_lvl1_id=wanted).exists():
                record = ko_lvl1.objects.using('picrust').get(ko_lvl1_id=wanted)
                koList = record.ko_entry_set.values_list('ko_orthology', flat=True)
            elif ko_lvl2.objects.using('picrust').filter(ko_lvl2_id=wanted).exists():
                record = ko_lvl2.objects.using('picrust').get(ko_lvl2_id=wanted)
                koList = record.ko_entry_set.values_list('ko_orthology', flat=True)
            elif ko_lvl3.objects.using('picrust').filter(ko_lvl3_id=wanted).exists():
                record = ko_lvl3.objects.using('picrust').get(ko_lvl3_id=wanted)
                koList = record.ko_entry_set.values_list('ko_orthology', flat=True)
            elif ko_entry.objects.using('picrust').filter(ko_lvl4_id=wanted).exists():
                koList = ko_entry.objects.using('picrust').filter(ko_lvl4_id=wanted).values_list('ko_orthology', flat=True)

            startTime = time.time()
            finalotuList = getOtuFromKoList(koList)
            print "Finished kegg_path query:", time.time()-startTime, "seconds elapsed.", len(finalotuList), "entries found"

            results = {}
            qs1 = OTU_99.objects.filter(otuid__in=finalotuList).values_list('kingdomid', 'kingdomid__kingdomName', 'phylaid', 'phylaid__phylaName', 'classid', 'classid__className', 'orderid', 'orderid__orderName', 'familyid', 'familyid__familyName', 'genusid', 'genusid__genusName', 'speciesid', 'speciesid__speciesName', 'otuName', 'otuid')
            results['data'] = list(qs1)
            myJson = ujson.dumps(results, ensure_ascii=False)

            return HttpResponse(myJson)
        else:
            # this is a slightly concerning state to get to, so we're logging it
            print "Request was not ajax!"
            functions.log(request, "NON_AJAX", "PATH_TAXA")
    except Exception as e:
        print "Error during kegg path: ", e


# TODO 1.3 html check
def kegg_path(request):
    functions.log(request, "PAGE", "KEGG_PATH")
    return render(
        request,
        'kegg_path.html'
    )


def nzJSON(request):
    qs1 = nz_entry.objects.using('picrust').values_list('nz_lvl1_id','nz_lvl1_id__nz_lvl1_name', 'nz_lvl2_id', 'nz_lvl2_id__nz_lvl2_name', 'nz_lvl3_id', 'nz_lvl3_id__nz_lvl3_name', 'nz_lvl4_id', 'nz_lvl4_id__nz_lvl4_name', 'nz_lvl5_id', 'nz_orthology', 'nz_name', 'nz_desc')
    results = {}
    results['data'] = list(qs1)
    myJson = ujson.dumps(results, ensure_ascii=False)
    return HttpResponse(myJson)


def nzTaxaJSON(request):
    try:
        if request.is_ajax():
            wanted = request.GET['key']
            wanted = wanted.replace('"', '')
            koList = []

            if nz_lvl1.objects.using('picrust').filter(nz_lvl1_id=wanted).exists():
                record = nz_lvl1.objects.using('picrust').get(nz_lvl1_id=wanted)
                koList = record.nz_entry_set.values_list('nz_orthology', flat=True)
            elif nz_lvl2.objects.using('picrust').filter(nz_lvl2_id=wanted).exists():
                record = nz_lvl2.objects.using('picrust').get(nz_lvl2_id=wanted)
                koList = record.nz_entry_set.values_list('nz_orthology', flat=True)
            elif nz_lvl3.objects.using('picrust').filter(nz_lvl3_id=wanted).exists():
                record = nz_lvl3.objects.using('picrust').get(nz_lvl3_id=wanted)
                koList = record.nz_entry_set.values_list('nz_orthology', flat=True)
            elif nz_lvl4.objects.using('picrust').filter(nz_lvl4_id=wanted).exists():
                record = nz_lvl4.objects.using('picrust').get(nz_lvl4_id=wanted)
                koList = record.nz_entry_set.values_list('nz_orthology', flat=True)
            elif nz_entry.objects.using('picrust').filter(nz_lvl5_id=wanted).exists():
                koList = nz_entry.objects.using('picrust').filter(nz_lvl5_id=wanted).values_list('nz_orthology', flat=True)

            startTime = time.time()
            finalotuList = getOtuFromKoList(koList)
            print "Finished kegg_enzyme query:", time.time()-startTime, "seconds elapsed. Starting data filter"
            filterTime = time.time()
            results = {}
            qs1 = OTU_99.objects.filter(otuid__in=finalotuList).values_list('kingdomid', 'kingdomid__kingdomName', 'phylaid', 'phylaid__phylaName', 'classid', 'classid__className', 'orderid', 'orderid__orderName', 'familyid', 'familyid__familyName', 'genusid', 'genusid__genusName', 'speciesid', 'speciesid__speciesName', 'otuid', 'otuName')
            results['data'] = list(qs1)
            myJson = ujson.dumps(results, ensure_ascii=False)
            print "Finished kegg_enzyme formatting:", time.time()-filterTime, "seconds elapsed. Returning"
            return HttpResponse(myJson)
        else:
            print "Request was not ajax!"
            functions.log(request, "NON_AJAX", "NZ_TAXA")
    except Exception as e:
        print "Error during kegg enzyme: ", e


# TODO 1.3 html check
def kegg_enzyme(request):
    functions.log(request, "PAGE", "KEGG_NZ")
    return render(
        request,
        'kegg_enzyme.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def norm(request):
    functions.log(request, "PAGE", "NORM")
    functions.cleanup('myPhyloDB/media/temp/norm')

    return render(
        request,
        'norm.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def ANOVA(request):
    functions.log(request, "PAGE", "ANOVA")
    functions.cleanup('myPhyloDB/media/temp/anova')

    return render(
        request,
        'anova.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def CORR(request):
    functions.log(request, "PAGE", "CORR")
    functions.cleanup('myPhyloDB/media/temp/corr')

    return render(
        request,
        'corr.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def core(request):
    functions.log(request, "PAGE", "CORE")
    functions.cleanup('myPhyloDB/media/temp/core')
    return render(
        request,
        'core.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def spac(request):  # refactored from rich to spac, not a typo
    functions.log(request, "PAGE", "SpAC")
    functions.cleanup('myPhyloDB/media/temp/spac')

    return render(
        request,
        'SpAC.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def soil_index(request):
    functions.log(request, "PAGE", "SOIL_INDEX")
    functions.cleanup('myPhyloDB/media/temp/soil_index')

    return render(
        request,
        'soil_index.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def DiffAbund(request):
    functions.log(request, "PAGE", "DIFFABUND")
    functions.cleanup('myPhyloDB/media/temp/diffabund')

    return render(
        request,
        'diff_abund.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def GAGE(request):
    functions.log(request, "PAGE", "GAGE")
    functions.cleanup('myPhyloDB/media/temp/gage')

    return render(
        request,
        'gage.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def PCA(request):
    functions.log(request, "PAGE", "PCA+")
    functions.cleanup('myPhyloDB/media/temp/pca')

    return render(
        request,
        'pca.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def PCoA(request):
    functions.log(request, "PAGE", "PCoA")
    functions.cleanup('myPhyloDB/media/temp/pcoa')

    return render(
        request,
        'pcoa.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def RF(request):
    functions.log(request, "PAGE", "RF")
    functions.cleanup('myPhyloDB/media/temp/rf')

    return render(
        request,
        'rf.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def SPLS(request):
    functions.log(request, "PAGE", "sPLS")
    functions.cleanup('myPhyloDB/media/temp/spls')

    return render(
        request,
        'spls.html'
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def WGCNA(request):
    functions.log(request, "PAGE", "WGCNA")
    functions.cleanup('myPhyloDB/media/temp/wgcna')

    return render(
        request,
        'wgcna.html'
    )


def saveSampleList(request):
    if request.is_ajax():
        allJson = request.GET["all"]
        selList = json.loads(allJson)
        # selList has the list of selected sampleIDs, need to verify they all belong to projects in getViewProjects
        debug("Selection:", selList)    # selList is just a list of sampleIDs
        myProjects = perms.getViewProjects(request)
        # sample object has projectid foreignkey field, which is actually a project (not a project id)
        mySamples = Sample.objects.filter(sampleid__in=selList)
        # TODO 1.4 tweak queue to only allow one high sample count analysis at a time, like a global max selected count?
        # essentially just need a way to make sure none of the large normalizations occur simultaneously
        allMatch = True
        for samp in mySamples:
            if samp.projectid in myProjects:
                continue
            else:
                allMatch = False
                break
        if len(mySamples) > maxSamples:
            text = "Maximum sample count exceeded: You selected ", len(mySamples), " while the maximum is ", maxSamples
            return HttpResponse(text)   # TODO 1.3 this is not redboxing user page, and norm is still lit
        if allMatch:
            # remove old files
            try:
                dirpath = 'myPhyloDB/media/usr_temp/' + str(request.user)
                shutil.rmtree(dirpath, ignore_errors=True)
            except Exception as e:
                print "Error during sample save, rmtree:", e

            # save file
            myDir = 'myPhyloDB/media/usr_temp/' + str(request.user) + '/'
            path = str(myDir) + 'usr_sel_samples.pkl'

            if not os.path.exists(myDir):
                os.makedirs(myDir)

            with open(path, 'wb') as f:
                pickle.dump(selList, f)

            thisUser = request.user.profile
            thisUser.dataID = uuid4().hex
            thisUser.save()

            text = 'Selected sample(s) have been recorded!\nYou may proceed directly to data normalization.'
            return HttpResponse(text)
        else:
            # samples in selList do not match permitted viewing samples
            text = 'Error with selecting samples: Invalid Permissions'
            # technically this can also occur when non-existent sampleIDs are given, but
            # that also likely stems from a security breach attempt, as the tree only shows real samples
            functions.securityLog(request, "SelectSamples", "permissions abuse attempt")
            return HttpResponse(text)  # TODO 1.3 this is not redboxing user page, and norm is still lit
    else:
        print "Request was not ajax!"
        functions.log(request, "NON_AJAX", "SAVE_SAMPLE")


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def reprocess(request):
    functions.log(request, "PAGE", "REPROCESS")
    return render(
        request,
        'reprocess.html',
        {'form4': UploadForm4,
         'mform': UploadForm10},
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def update(request):
    functions.log(request, "PAGE", "UPDATE")
    state = ''
    return render(
        request,
        'update.html',
        {'form5': UploadForm5,
         'state': state}
    )


# TODO 1.3 html check
def updaStop(request):
    file1 = request.FILES['docfile11']
    p_uuid, pType, num_samp = functions.projectid(file1)
    proj = Project.objects.get(projectid=p_uuid)
    proj.wip = False
    proj.save()
    functions.log(request, "STOP", "UPDATE")
    state = "Update stopped"
    return render(
        request,
        'update.html',
        {'form5': UploadForm5,
         'state': state}
    )


# TODO 1.3 html check
def updateFunc(request, stopList):  # TODO 1.3 update should use meta file from user uploads (atm using upload form)
    form5 = UploadForm5(request.POST, request.FILES)
    state = ''
    curUser = User.objects.get(username=request.user.username)

    PID = 0
    RID = request.POST['RID']

    if stopList[PID] == RID:
        return updaStop(request)

    if form5.is_valid():
        refid = request.POST['refid']
        file1 = request.FILES['docfile11']

        if stopList[PID] == RID:
            return updaStop(request)

        ref = Reference.objects.get(refid=refid)    # TODO 1.3 why is this called from processFunc???
        dest = ref.path
        p_uuid, pType, num_samp = functions.projectid(file1)
        if not num_samp:
            return render(
                request,
                'update.html',
                {'form5': UploadForm5,
                 'state': "Error: Please open and save your meta file in Excel/OpenOffice in order to perform the necessary calculations..."
                 }
            )

        if stopList[PID] == RID:
            return updaStop(request)

        try:
            metaFile = '/'.join([dest, file1.name])
            functions.handle_uploaded_file(file1, dest, file1.name)
            functions.parse_project(metaFile, p_uuid, curUser)
        except Exception as e:
            state = "There was an error parsing your metafile: " + str(file1.name) + "\nError info: "+str(e)
            return render(
                request,
                'update.html',
                {'form5': UploadForm5,
                 'state': state}
            )

        if stopList[PID] == RID:
            return updaStop(request)

        try:
            reference = Reference.objects.get(refid=refid)
            raw = reference.raw
            source = reference.source
            userID = str(request.user.id)
            functions.parse_sample(metaFile, p_uuid, pType, num_samp, dest, raw, source, userID, stopList, RID, PID)
        except Exception as e:
            state = "There was an error parsing your metafile: " + str(file1.name) + "\nError info: "+str(e)
            return render(
                request,
                'update.html',
                {'form5': UploadForm5,
                 'state': state}
            )

        if stopList[PID] == RID:
            return updaStop(request)

        proj = Project.objects.get(projectid=p_uuid)
        proj.wip = False
        proj.save()

        state = 'Path: ' + str(dest) + ' is finished parsing!'

    if stopList[PID] == RID:
        return updaStop(request)

    return render(
        request,
        'update.html',
        {'form5': UploadForm5,
         'state': state}
    )


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
@user_passes_test(lambda u: u.is_superuser)
def pybake(request):
    functions.log(request, "PAGE", "PYBAKE")

    return render(
        request,
        'pybake.html',
        {'form6': UploadForm6,
         'form7': UploadForm7,
         'form8': UploadForm8}
    )


# TODO 1.3 html check
def uploadNorm(request):
    uploaded = request.FILES["normFile"]
    savedDF = pd.read_csv(uploaded, index_col=0, sep='\t')
    selList = list(set(savedDF['sampleid']))

    # save selected samples file to users temp/ folder
    myDir = 'myPhyloDB/media/usr_temp/' + str(request.user) + '/'
    path = str(myDir) + 'usr_sel_samples.pkl'

    if not os.path.exists(myDir):
        os.makedirs(myDir)

    with open(path, 'wb') as f:
        pickle.dump(selList, f)

    # save normalized file to users temp/ folder
    myDir = 'myPhyloDB/media/usr_temp/' + str(request.user) + '/'
    path = str(myDir) + 'usr_norm_data.csv'

    if not os.path.exists(myDir):
        os.makedirs(myDir)
    savedDF.to_csv(path, sep='\t')

    projectList = list(set(savedDF['projectid']))

    if Project.objects.filter(projectid__in=projectList).exists():
        projects = Project.objects.filter(projectid__in=projectList)

    else:
        return render(
            request,
            'select.html',
            {'normpost': 'One or more projects were not found in your database!'}
        )

    if Reference.objects.filter(projectid__in=projectList).exists():
        refs = Reference.objects.filter(projectid__in=projectList)
    else:
        return render(
            request,
            'select.html',
            {'normpost': 'One or more projects were not found in your database!'}
        )

    if Sample.objects.filter(sampleid__in=selList).exists():
        samples = Sample.objects.filter(sampleid__in=selList)
    else:
        return render(
            request,
            'select.html',
            {'normpost': 'One or more samples were not found in your database!'}
        )

    biome = {}
    tempDF = savedDF.drop_duplicates(subset='sampleid', keep='last')
    tempDF.set_index('sampleid', drop=True, inplace=True)

    metaDF = tempDF.drop(['kingdomName', 'phylaName', 'className', 'orderName', 'familyName', 'genusName', 'speciesName', 'otuName', 'otuid', 'abund', 'rel_abund', 'abund_16S', 'rich', 'diversity'], axis=1)
    myList = list(tempDF.index.values)

    nameList = []
    for i in myList:
        nameList.append({"id": str(i), "metadata": metaDF.loc[i].to_dict()})

    # get list of lists with abundances
    taxaOnlyDF = savedDF.loc[:, ['sampleid', 'kingdomName', 'phylaName', 'className', 'orderName', 'familyName', 'genusName', 'speciesName', 'otuName', 'otuid', 'abund']]
    taxaOnlyDF = taxaOnlyDF.pivot(index='otuid', columns='sampleid', values='abund')
    dataList = taxaOnlyDF.values.tolist()

    # get list of taxa
    namesDF = savedDF.loc[:, ['sampleid', 'otuid']]
    namesDF['taxa'] = savedDF.loc[:, ['kingdomName', 'phylaName', 'className', 'orderName', 'familyName', 'genusName', 'speciesName', 'otuName']].values.tolist()
    namesDF = namesDF.pivot(index='otuid', columns='sampleid', values='taxa')

    taxaList = []
    for index, row in namesDF.iterrows():
        metaDict = {}
        metaDict['taxonomy'] = row[0]
        taxaList.append({"id": index, "metadata": metaDict})

    biome['format'] = 'Biological Observation Matrix 0.9.1-dev'
    biome['format_url'] = 'http://biom-format.org/documentation/format_versions/biom-1.0.html'
    biome['type'] = 'OTU table'
    biome['generated_by'] = 'myPhyloDB'
    biome['date'] = str(datetime.datetime.now())
    biome['matrix_type'] = 'dense'
    biome['matrix_element_type'] = 'float'
    biome['rows'] = taxaList
    biome['columns'] = nameList
    biome['data'] = dataList

    myDir = 'myPhyloDB/media/usr_temp/' + str(request.user) + '/'
    path = str(myDir) + 'usr_norm_data.biom'
    with open(path, 'w') as outfile:
        ujson.dump(biome, outfile, ensure_ascii=True, indent=4, sort_keys=True)

    return render(
        request,
        'select.html',
        {'form9': UploadForm9,
         'projects': projects,
         'refs': refs,
         'samples': samples,
         'selList': selList,
         'normpost': 'Success'},
    )


# Function to create a user folder at login
def login_usr_callback(sender, user, request, **kwargs):
    user = request.user
    if not os.path.exists('myPhyloDB/media/usr_temp/'+str(user)):
        os.makedirs('myPhyloDB/media/usr_temp/'+str(user))
    functions.cleanup('myPhyloDB/media/usr_temp/'+str(user))

user_logged_in.connect(login_usr_callback)


# Function has been added to context processor in settings file
def usrFiles(request):
    selFiles = os.path.exists('myPhyloDB/media/usr_temp/' + str(request.user) + '/usr_sel_samples.pkl')
    normFiles = os.path.exists('myPhyloDB/media/usr_temp/' + str(request.user) + '/myphylodb.biom')

    try:
        dataID = UserProfile.objects.get(user=request.user).dataID
    except Exception:
        dataID = "nodataidfound"

    return {
        'selFiles': selFiles,
        'normFiles': normFiles,
        'dataID': dataID
    }


# TODO 1.3 html check
@login_required(login_url='/myPhyloDB/accounts/login/')
def profile(request):
    functions.log(request, "PAGE", "PROFILE")
    projects = Project.objects.none()
    if request.user.is_superuser:
        projects = Project.objects.all().order_by('project_name')
    elif request.user.is_authenticated():
        path_list = Reference.objects.filter(Q(author=request.user)).values_list('projectid_id')
        projects = Project.objects.all().filter( Q(projectid__in=path_list) | Q(status='public') ).order_by('project_name')
    if not request.user.is_superuser and not request.user.is_authenticated():
        projects = Project.objects.all().filter( Q(status='public') ).order_by('project_name')

    return render(
        request,
        'profile.html',
        {'projects': projects}
    )


# TODO 1.3 html check
def changeuser(request):
    functions.log(request, "PAGE", "CHANGEUSER")
    return render(
        request,
        'changeuser.html',
        {"form": UserUpdateForm,
            "error": "none"}
    )


# TODO 1.3 html check
def updateInfo(request):
    functions.log(request, "FUNCTION", "UPDATEINFO")
    stuff = request.POST
    user = request.user
    pword = stuff['pword']
    error = "none"
    try:
        verified = check_password(pword, user.password)

        if verified:
            user.first_name = stuff['firstName']
            user.last_name = stuff['lastName']
            user.email = stuff['email']

            # kept out of loop to avoid changing things which aren't meant to change so easily (ie whitelist perms)
            up = request.user.profile
            up.firstName = stuff['firstName']
            up.lastName = stuff['lastName']
            up.affiliation = stuff['affiliation']
            up.city = stuff['city']
            up.state = stuff['state']
            up.country = stuff['country']
            up.zip = stuff['zip']
            up.phone = stuff['phone']
            up.reference = stuff['reference']
            up.purpose = stuff['purpose']
            up.dataID = uuid4().hex
            user.save()
            up.save()
            messages.success(request, 'Your profile was updated.')
        else:
            messages.error(request, 'Password did not match')
    except Exception:
        messages.error(request, 'Something went wrong. Please check your entered values')

    return render(
        request,
        'changeuser.html',
        {"form": UserUpdateForm, "error": error}
    )


def checkSamples(metaFile, source, fileName):
    wb = openpyxl.load_workbook(metaFile, data_only=True, read_only=True)
    ws = wb['MIMARKs']
    metaList = []
    for row in xrange(7, ws.max_row+1):
        val = ws.cell(row=row, column=3).value
        if val is not None:
            metaList.append(str(val))

    # now get list from contig or oligo files
    mothurList = []
    # if 454_sff or 454_fastq check oligos
    if source == "454_sff":
        # multiple oligos named in given file column index 2
        lc = 0
        with open(fileName, 'rb') as myFile:
            oligosList = []
            for line in myFile:
                segments = line.strip('\n').strip('\r').split("\t")
                if len(segments) >= 2:
                    if '\n' not in segments[1] and '\r' not in segments[1] and str(segments[1]) != '':
                        oligosList.append(str(segments[1]))
            for oligos in oligosList:
                with open("mothur/temp/"+oligos, 'rb') as myOligos:
                    for line in myOligos:
                        if lc == 0:
                            lc += 1
                            continue
                        segments = line.strip('\n').strip('\r').split("\t")
                        if len(segments) >= 3:
                            if '\n' not in segments[2] and '\r' not in segments[2] and str(segments[2]) != '':
                                mothurList.append(str(segments[2]))
    if source == "454_fastq":
        # single oligos file to parse, column index 2
        lc = 0
        with open(fileName, 'rb') as myFile:
            for line in myFile:
                if lc == 0:
                    lc += 1
                    continue
                segments = line.strip('\n').strip('\r').split("\t")
                if len(segments) >= 3:
                    if '\n' not in segments[2] and '\r' not in segments[2] and str(segments[2]) != '':
                        mothurList.append(str(segments[2]))
    # if miseq check contig
    if source == "miseq":
        # open temp.files, check column 0
        with open(fileName, 'rb') as myFile:
            for line in myFile:
                segments = line.strip('\n').strip('\r').split("\t")
                if len(segments) >= 1:
                    if '\n' not in segments[0] and '\r' not in segments[0] and str(segments[0]) != '':
                        mothurList.append(str(segments[0]))

    # match two lists and return result
    foundMothurDict = {}
    foundMetaDict = {}
    duplicateMothurList = []
    duplicateMetaList = []
    problemMothurList = []
    problemMetaList = []
    for thing in mothurList:
        if thing in foundMothurDict.keys():
            duplicateMothurList.append(thing)
        else:
            foundMothurDict[thing] = thing
        if thing not in metaList:
            problemMothurList.append(thing)
    for thing in metaList:
        if thing in foundMetaDict.keys():
            duplicateMetaList.append(thing)
        else:
            foundMetaDict[thing] = thing
        if thing not in mothurList:
            problemMetaList.append(thing)

    errorString = ""

    if len(duplicateMothurList) > 0:
        errorString += "The following samples were found more than once in your contig file:\n"
        errorString += str(duplicateMothurList) + "\n"
    if len(duplicateMetaList) > 0:
        errorString += "The following samples were found more than once in your meta file:\n"
        errorString += str(duplicateMetaList) + "\n"

    if len(problemMothurList) > 0:
        errorString += "The following samples are in your contig file but not in your meta file:\n"
        errorString += str(problemMothurList) + "\n"

    if len(problemMetaList) > 0:
        errorString += "The following samples are in your meta file but not in your contig file:\n"
        errorString += str(problemMetaList) + "\n"

    return errorString


def getAdminLog():      # this gets called by server_utils with a specific command, could probably just move the code
    logs = LogEntry.objects.all()
    for log in logs:
        logFile = open('admin_log.txt', 'a')
        logFile.write(str(log.action_time)+" : "+str(log.user)+" : "+str(log)+"\n")
        logFile.close()
